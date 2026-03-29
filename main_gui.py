import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, colorchooser, simpledialog
import pandas as pd
import numpy as np
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.figure import Figure
from matplotlib.gridspec import GridSpec
from matplotlib.patches import Ellipse
import scipy.cluster.hierarchy as sch
import os
import re
import pickle
import datetime
from collections import defaultdict
import itertools
try:
    from openpyxl.styles import PatternFill
except ImportError:
    PatternFill = None # Fallback if openpyxl not installed
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler, LabelEncoder, OneHotEncoder
from sklearn.cross_decomposition import PLSRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import cross_val_score, cross_val_predict
from scipy import stats
from visualization import VisualizationManager
from export_utils import ExportManager
from mgf_handler import MGFSearchEngine


class MetabolomicsApp(tk.Tk):
    """Main application window for herbal metabolomics analysis"""
    
    def __init__(self):
        super().__init__()
        
        # Window configuration
        self.title("Herbal Metabolomics Analyzer v2.15.1")
        self.geometry("1400x900")
        self.configure(bg="#8df109")
        
        # Data storage
        self.data = None
        self.preprocessed_data = None
        self.screened_data = None  # Store dataset after Univariate Screening
        self.data_before_norm = None  # Store data before normalization for comparison
        self.feature_metadata = None
        self.sample_names = None
        self.group_labels = None
        self.replicate_mapping = {}  # Store replicate relationships
        self.group_mapping = {}  # Store sample -> group mapping
        self.group_colors = {}  # Store group -> color mapping
        self.qc_samples = []  # Store QC sample names
        
        # Analysis results storage
        self.pca_result = None
        self.plsda_result = None
        self.volcano_result = None
        self.rf_result = None
        self.heatmap_data = None
        self.generated_plots = {}

        # Global plot settings
        self.plot_width_var = tk.DoubleVar(value=8.0)
        self.plot_height_var = tk.DoubleVar(value=6.0)
        self.plot_dpi_var = tk.IntVar(value=100)
        
        # Create GUI
        self.create_menu_bar()
        self.create_main_layout()
        self.create_status_bar()

        # Visualization manager
        self.vis_manager = VisualizationManager(self)
        self.export_manager = ExportManager()
        self.mgf_engine = MGFSearchEngine()
        
    def create_menu_bar(self):
        """Create menu bar"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Load CSV", command=self.load_csv, accelerator="Ctrl+O")
        file_menu.add_command(label="Load Session", command=self.load_session, accelerator="Ctrl+L")
        file_menu.add_command(label="Save Session", command=self.save_session, accelerator="Ctrl+S")
        file_menu.add_separator()
        
        # Export submenu
        export_menu = tk.Menu(file_menu, tearoff=0)
        file_menu.add_cascade(label="Export", menu=export_menu)
        export_menu.add_command(label="Export Preprocessed Data", command=self.export_preprocessed)
        export_menu.add_command(label="Export PCA Results", command=self.export_pca)
        export_menu.add_command(label="Export Volcano Results", command=self.export_volcano_results)
        export_menu.add_command(label="Export PLS-DA VIPs", command=self.export_plsda_vips)
        export_menu.add_command(label="Export RF Features", command=self.export_rf_features)
        export_menu.add_command(label="Export Consensus Biomarkers", command=self.export_consensus_features)
        export_menu.add_command(label="Export Heatmap Data", command=self.export_heatmap_data)
        export_menu.add_separator()
        export_menu.add_command(label="Export Current Plot (High Res)", command=self.export_current_plot)
        export_menu.add_command(label="Export All Plots to PDF Report", command=self.export_pdf_report)

        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit, accelerator="Ctrl+Q")
        
        # Edit menu
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        edit_menu.add_command(label="Rename Samples", command=self.rename_samples_dialog)
        edit_menu.add_command(label="Configure Groups", command=self.configure_groups)
        edit_menu.add_command(label="Manage Group Names", command=self.manage_group_names)
        edit_menu.add_command(label="Select QC Samples", command=self.select_qc_samples)
        edit_menu.add_separator()
        edit_menu.add_command(label="Plot Settings...", command=self.configure_plot_settings)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="Quick Start Guide", command=self.show_help)
        help_menu.add_command(label="About", command=self.show_about)
        
        # Keyboard shortcuts
        self.bind('<Control-o>', lambda e: self.load_csv())
        self.bind('<Control-q>', lambda e: self.quit())
        self.bind('<Control-s>', lambda e: self.save_session())
        self.bind('<Control-l>', lambda e: self.load_session())
        
    def create_main_layout(self):
        """Create main application layout with notebook tabs"""
        
        # Create notebook (tab container)
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Create tabs logically sorted for sequential workflow
        self.create_data_tab()
        self.create_preprocessing_tab()
        self.create_pca_tab()
        self.create_volcano_tab()  # Univariate screening now precedes predictive ML
        self.create_plsda_tab()
        self.create_pls_val_tab()
        self.create_rf_tab()
        self.create_heatmap_tab()
        self.create_venn_tab()
        self.create_upset_tab()
        self.create_feature_viewer_tab()
        self.create_spectrum_tab()
        
    def create_data_tab(self):
        """Tab 1: Data Loading and Preview"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📂 Data")
        
        # Load button and group config
        load_frame = ttk.Frame(tab)
        load_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(
            load_frame,
            text="📂 Load CSV File",
            command=self.load_csv,
            width=18
        ).pack(side='left', padx=3)
        
        ttk.Button(
            load_frame,
            text="🏷️ Rename Samples",
            command=self.rename_samples_dialog,
            width=18
        ).pack(side='left', padx=3)
        
        ttk.Button(
            load_frame,
            text="⚙️ Configure Groups",
            command=self.configure_groups,
            width=18
        ).pack(side='left', padx=3)
        
        ttk.Button(
            load_frame,
            text="✏️ Manage Group Names",
            command=self.manage_group_names,
            width=20
        ).pack(side='left', padx=3)
        
        ttk.Button(
            load_frame,
            text="🔬 Select QC Samples",
            command=self.select_qc_samples,
            width=18
        ).pack(side='left', padx=3)
        
        self.file_label = ttk.Label(load_frame, text="No file loaded", foreground="gray")
        self.file_label.pack(side='left', padx=10)
        
        # Data summary
        summary_frame = ttk.LabelFrame(tab, text="Data Summary", padding=10)
        summary_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.summary_text = scrolledtext.ScrolledText(
            summary_frame,
            height=10,
            font=('Courier', 10)
        )
        self.summary_text.pack(fill='both', expand=True)
        
        # Data preview
        preview_frame = ttk.LabelFrame(tab, text="Data Preview (First 10 Rows)", padding=10)
        preview_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Create treeview for table display
        self.preview_tree = ttk.Treeview(preview_frame, show='headings')
        self.preview_tree.pack(side='left', fill='both', expand=True)
        
        # Scrollbars
        vsb = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_tree.yview)
        vsb.pack(side='right', fill='y')
        hsb = ttk.Scrollbar(tab, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.preview_tree.bind("<Button-3>", self.show_preview_context_menu)
        
    def show_preview_context_menu(self, event):
        iid = self.preview_tree.identify_row(event.y)
        if iid:
            self.preview_tree.selection_set(iid)
            feature_id = self.preview_tree.item(iid, "values")[0]
            
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="Search Spectrum for this Feature", 
                             command=lambda: self.send_to_spectrum_search(feature_id))
            menu.add_command(label="View in Feature Viewer", 
                             command=lambda: self.open_in_feature_viewer(feature_id))
            menu.tk_popup(event.x_root, event.y_root)

    def send_to_spectrum_search(self, feature_id):
        for i in range(self.notebook.index("end")):
            if "Spectrum Search" in self.notebook.tab(i, "text"):
                self.notebook.select(i)
                break
        self.spectrum_search_type_var.set("Feature ID")
        self.spectrum_search_value_var.set(feature_id)
        self.search_spectra()

    def open_in_feature_viewer(self, feature_id):
        for i in range(self.notebook.index("end")):
            if "Feature Viewer" in self.notebook.tab(i, "text"):
                self.notebook.select(i)
                self.feature_viewer_id_var.set(feature_id)
                self.run_feature_plot()
                break
        
    def create_preprocessing_tab(self):
        """Tab 2: Preprocessing with Feature Filtering and Distribution Plots"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔧 Preprocessing")
        
        # Top frame for options and results
        top_frame = ttk.Frame(tab)
        top_frame.pack(fill='x', padx=10, pady=10)
        
        # Left: Options frame
        options_frame = ttk.LabelFrame(top_frame, text="Preprocessing Options", padding=10)
        options_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        # Technical replicate averaging
        row = 0
        self.avg_replicates_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            options_frame,
            text="Keep original replicates + Add averaged value (Skips QCs)",
            variable=self.avg_replicates_var
        ).grid(row=row, column=0, columnspan=3, sticky='w', pady=5)
        
        # Missing Value Imputation Section
        row += 1
        ttk.Separator(options_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky='ew', pady=10)
        
        row += 1
        ttk.Label(
            options_frame, 
            text="🩹 Missing Value Imputation:", 
            font=('Arial', 10, 'bold')
        ).grid(row=row, column=0, columnspan=3, sticky='w', pady=(5,2))
        
        row += 1
        self.impute_lod_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            options_frame,
            text="LOD Imputation (replace zeros/missing)",
            variable=self.impute_lod_var
        ).grid(row=row, column=0, sticky='w', padx=(20,5))
        
        ttk.Label(options_frame, text="Fraction of Min:").grid(row=row, column=1, sticky='e', padx=2)
        self.lod_fraction_var = tk.StringVar(value="1/5 (20%)")
        ttk.Combobox(
            options_frame,
            textvariable=self.lod_fraction_var,
            values=["1/2 (50%)", "1/3 (33%)", "1/4 (25%)", "1/5 (20%)"],
            state='readonly',
            width=10
        ).grid(row=row, column=2, sticky='w', padx=2)
        
        # Missing Value Imputation Section
        row += 1
        ttk.Separator(options_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky='ew', pady=10)
        
        row += 1
        ttk.Label(
            options_frame, 
            text="🩹 Missing Value Imputation:", 
            font=('Arial', 10, 'bold')
        ).grid(row=row, column=0, columnspan=3, sticky='w', pady=(5,2))
        
        row += 1
        self.impute_lod_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            options_frame,
            text="LOD Imputation (replace zeros/missing)",
            variable=self.impute_lod_var
        ).grid(row=row, column=0, sticky='w', padx=(20,5))
        
        ttk.Label(options_frame, text="Fraction of Min:").grid(row=row, column=1, sticky='e', padx=2)
        self.lod_fraction_var = tk.DoubleVar(value=0.2)
        ttk.Spinbox(
            options_frame,
            from_=0.1,
            to=0.5,
            increment=0.1,
            textvariable=self.lod_fraction_var,
            width=8
        ).grid(row=row, column=2, sticky='w', padx=2)
        
        # Feature Filtering Section
        row += 1
        ttk.Separator(options_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky='ew', pady=10)
        
        row += 1
        ttk.Label(
            options_frame, 
            text="🔍 Feature Filtering:", 
            font=('Arial', 10, 'bold')
        ).grid(row=row, column=0, columnspan=3, sticky='w', pady=(5,2))
        
        # Detection rate filter
        row += 1
        self.filter_detection_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            options_frame,
            text="Detection rate filter",
            variable=self.filter_detection_var
        ).grid(row=row, column=0, sticky='w', padx=(20,5))
        
        ttk.Label(options_frame, text="Min %:").grid(row=row, column=1, sticky='e', padx=2)
        self.detection_threshold_var = tk.DoubleVar(value=80.0)
        ttk.Spinbox(
            options_frame,
            from_=50,
            to=100,
            increment=5,
            textvariable=self.detection_threshold_var,
            width=8
        ).grid(row=row, column=2, sticky='w', padx=2)
        
        # Minimum intensity filter
        row += 1
        self.filter_intensity_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            options_frame,
            text="Minimum intensity filter",
            variable=self.filter_intensity_var
        ).grid(row=row, column=0, sticky='w', padx=(20,5))
        
        ttk.Label(options_frame, text="Threshold:").grid(row=row, column=1, sticky='e', padx=2)
        self.min_intensity_var = tk.DoubleVar(value=1000.0)
        ttk.Entry(
            options_frame,
            textvariable=self.min_intensity_var,
            width=10
        ).grid(row=row, column=2, sticky='w', padx=2)
        
        # RSD filter (for QC samples)
        row += 1
        self.filter_rsd_var = tk.BooleanVar(value=False)
        filter_rsd_check = ttk.Checkbutton(
            options_frame,
            text="QC RSD filter",
            variable=self.filter_rsd_var,
            command=self.on_rsd_filter_toggle
        )
        filter_rsd_check.grid(row=row, column=0, sticky='w', padx=(20,5))
        
        ttk.Label(options_frame, text="Max RSD %:").grid(row=row, column=1, sticky='e', padx=2)
        self.rsd_threshold_var = tk.DoubleVar(value=30.0)
        ttk.Spinbox(
            options_frame,
            from_=10,
            to=50,
            increment=5,
            textvariable=self.rsd_threshold_var,
            width=8
        ).grid(row=row, column=2, sticky='w', padx=2)
        
        # QC sample selection button (initially hidden)
        row += 1
        self.qc_select_frame = ttk.Frame(options_frame)
        self.qc_select_frame.grid(row=row, column=0, columnspan=3, sticky='w', padx=(40,5), pady=2)
        self.qc_select_frame.grid_remove()  # Hide initially
        
        ttk.Button(
            self.qc_select_frame,
            text="🔬 Select QC Samples",
            command=self.select_qc_samples,
            width=20
        ).pack(side='left', padx=2)
        
        self.qc_status_label = ttk.Label(
            self.qc_select_frame,
            text="No QC samples selected",
            foreground="gray",
            font=('Arial', 8)
        )
        self.qc_status_label.pack(side='left', padx=5)
        
        # IQR filter
        row += 1
        self.filter_iqr_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            options_frame,
            text="IQR variance filter",
            variable=self.filter_iqr_var
        ).grid(row=row, column=0, sticky='w', padx=(20,5))
        
        ttk.Label(options_frame, text="Factor:").grid(row=row, column=1, sticky='e', padx=2)
        self.iqr_factor_var = tk.DoubleVar(value=0.5)
        ttk.Spinbox(
            options_frame,
            from_=0.1,
            to=2.0,
            increment=0.1,
            textvariable=self.iqr_factor_var,
            width=8,
            format="%.1f"
        ).grid(row=row, column=2, sticky='w', padx=2)
        
        # Normalization section
        row += 1
        ttk.Separator(options_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky='ew', pady=10)
        
        row += 1
        ttk.Label(
            options_frame,
            text="Normalization Method:",
            font=('Arial', 10, 'bold')
        ).grid(row=row, column=0, sticky='w', pady=5)
        
        row += 1
        self.norm_method_var = tk.StringVar(value="Log")
        norm_combo = ttk.Combobox(
            options_frame,
            textvariable=self.norm_method_var,
            values=["TIC", "Median", "Log", "Pareto", "Auto", "None"],
            state='readonly',
            width=15
        )
        norm_combo.grid(row=row, column=0, columnspan=3, sticky='w', padx=20, pady=5)
        
        # Run button
        row += 1
        ttk.Button(
            options_frame,
            text="▶️ Run Preprocessing",
            command=self.run_preprocessing
        ).grid(row=row, column=0, columnspan=3, pady=10)
        
        # Visualization button
        row += 1
        self.show_dist_button = ttk.Button(
            options_frame,
            text="📊 Show Distribution Plots",
            command=self.show_distribution_plots,
            state='disabled'
        )
        self.show_dist_button.grid(row=row, column=0, columnspan=3, pady=5)
        
        # Right: Results display
        results_frame = ttk.LabelFrame(top_frame, text="Preprocessing Results", padding=10)
        results_frame.pack(side='left', fill='both', expand=True, padx=(5, 0))
        
        self.preprocess_text = scrolledtext.ScrolledText(
            results_frame,
            height=18,
            width=40,
            font=('Courier', 9)
        )
        self.preprocess_text.pack(fill='both', expand=True)
        
        # Bottom: Plot area for distribution visualization
        plot_label_frame = ttk.LabelFrame(tab, text="Distribution Visualization", padding=10)
        plot_label_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.preprocessing_plot_frame = ttk.Frame(plot_label_frame)
        self.preprocessing_plot_frame.pack(fill='both', expand=True)
        
    def create_pca_tab(self):
        """Tab 3: PCA Analysis with Grouping"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📊 PCA")
        
        # Controls frame
        control_frame = ttk.LabelFrame(tab, text="PCA Settings", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)
        
        # Split into left (settings) and right (groups)
        settings_frame = ttk.Frame(control_frame)
        settings_frame.pack(side='left', fill='y', padx=(0, 20))
        
        groups_frame = ttk.LabelFrame(control_frame, text="Select Groups (All if none selected)", padding=5)
        groups_frame.pack(side='left', fill='both', expand=True)

        # Settings
        row = 0
        ttk.Label(settings_frame, text="Number of Components:").grid(row=row, column=0, sticky='w', padx=5, pady=5)
        self.pca_components_var = tk.IntVar(value=2)
        ttk.Spinbox(
            settings_frame,
            from_=2,
            to=10,
            textvariable=self.pca_components_var,
            width=10
        ).grid(row=row, column=1, sticky='w', padx=5, pady=5)
        
        # Log Transform Option
        row += 1
        self.pca_log_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            settings_frame,
            text="Apply Log10 Transform (fixes severe skew)",
            variable=self.pca_log_var
        ).grid(row=row, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        # Show ellipses option
        row += 1
        self.show_ellipses_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            settings_frame,
            text="Show 95% confidence ellipses",
            variable=self.show_ellipses_var
        ).grid(row=row, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        # Show labels option
        row += 1
        self.show_labels_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            settings_frame,
            text="Show sample labels on plot",
            variable=self.show_labels_var
        ).grid(row=row, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        # PERMANOVA option
        row += 1
        self.pca_permanova_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            settings_frame,
            text="Calculate & display PERMANOVA p-value",
            variable=self.pca_permanova_var
        ).grid(row=row, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        # PERMDISP option
        row += 1
        self.pca_permdisp_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            settings_frame,
            text="Calculate & display PERMDISP p-value",
            variable=self.pca_permdisp_var
        ).grid(row=row, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        # Run button
        row += 1
        ttk.Button(
            settings_frame,
            text="▶️ Run PCA",
            command=self.run_pca
        ).grid(row=row, column=0, columnspan=2, pady=10)
        
        # Group Selection Listbox
        self.pca_group_listbox = tk.Listbox(groups_frame, selectmode='multiple', height=6, exportselection=0)
        self.pca_group_listbox.pack(side='left', fill='both', expand=True)
        sb = ttk.Scrollbar(groups_frame, orient='vertical', command=self.pca_group_listbox.yview)
        sb.pack(side='right', fill='y')
        self.pca_group_listbox.config(yscrollcommand=sb.set)
        
        # Plot area
        self.pca_plot_frame = ttk.Frame(tab)
        self.pca_plot_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
    def create_volcano_tab(self):
        """Tab 4: Univariate Screening (Volcano Plot)"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📊 Univariate Screening")
        
        # Controls
        control_frame = ttk.LabelFrame(tab, text="Screening Settings & Filtering", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)
        
        # Explicit Group Selection
        group_frame = ttk.Frame(control_frame)
        group_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(group_frame, text="Group 1 (Control):").pack(side='left', padx=5)
        self.volcano_group1_var = tk.StringVar()
        self.volcano_group1_combo = ttk.Combobox(group_frame, textvariable=self.volcano_group1_var, state='readonly', width=15)
        self.volcano_group1_combo.pack(side='left', padx=5)
        
        ttk.Label(group_frame, text="Group 2 (Treatment):").pack(side='left', padx=15)
        self.volcano_group2_var = tk.StringVar()
        self.volcano_group2_combo = ttk.Combobox(group_frame, textvariable=self.volcano_group2_var, state='readonly', width=15)
        self.volcano_group2_combo.pack(side='left', padx=5)
        
        # Thresholds
        # Add Correction Method Selection
        correction_frame = ttk.Frame(control_frame)
        correction_frame.pack(fill='x', pady=(0, 5))
        
        ttk.Label(correction_frame, text="Correction Method:").pack(side='left', padx=5)
        self.correction_method_var = tk.StringVar(value="None")
        self.correction_method_combo = ttk.Combobox(
            correction_frame, textvariable=self.correction_method_var, 
            values=["None", "Bonferroni", "FDR (Benjamini-Hochberg)"], state='readonly', width=25
        )
        self.correction_method_combo.pack(side='left', padx=5)

        thresh_frame = ttk.Frame(control_frame)
        thresh_frame.pack(fill='x')
        
        ttk.Label(thresh_frame, text="P-value threshold:").pack(side='left', padx=5)
        self.pvalue_var = tk.DoubleVar(value=0.05)
        ttk.Entry(thresh_frame, textvariable=self.pvalue_var, width=10).pack(side='left', padx=5)
        
        ttk.Label(thresh_frame, text="Fold-change threshold:").pack(side='left', padx=5)
        self.fc_var = tk.DoubleVar(value=2.0)
        ttk.Entry(thresh_frame, textvariable=self.fc_var, width=10).pack(side='left', padx=5)
        
        ttk.Button(
            thresh_frame,
            text="▶️ Create Volcano Plot",
            command=self.run_volcano
        ).pack(side='left', padx=15)
        
        ttk.Button(
            thresh_frame,
            text="💾 Filter Data for Downstream ML",
            command=self.apply_screening_filter
        ).pack(side='left', padx=5)
        
        # Plot area
        self.volcano_plot_frame = ttk.Frame(tab)
        self.volcano_plot_frame.pack(fill='both', expand=True, padx=10, pady=5)

    def create_plsda_tab(self):
        """Tab 5: PLS-DA Analysis"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📈 PLS-DA & VIP")
        
        # Controls
        control_frame = ttk.LabelFrame(tab, text="PLS-DA Settings", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)
        
        # Number of components
        ttk.Label(control_frame, text="Number of Components:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.plsda_components_var = tk.IntVar(value=2)
        ttk.Spinbox(
            control_frame,
            from_=2,
            to=10,
            textvariable=self.plsda_components_var,
            width=10
        ).grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        # Log Transform Option
        self.plsda_log_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            control_frame,
            text="Apply Log10 Transform",
            variable=self.plsda_log_var
        ).grid(row=1, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        # Exclude QC Option
        self.plsda_exclude_qc_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            control_frame,
            text="Exclude QC Samples",
            variable=self.plsda_exclude_qc_var
        ).grid(row=2, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        # --- Pairwise Comparison Option ---
        self.plsda_pairwise_var = tk.BooleanVar(value=False)
        
        def toggle_pairwise():
            if self.plsda_pairwise_var.get():
                self.plsda_group_frame.grid()
                self.update_group_dropdowns()
            else:
                self.plsda_group_frame.grid_remove()

        ttk.Checkbutton(
            control_frame,
            text="Pairwise Comparison (Select 2 Groups)",
            variable=self.plsda_pairwise_var,
            command=toggle_pairwise
        ).grid(row=3, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        self.plsda_group_frame = ttk.Frame(control_frame)
        self.plsda_group_frame.grid(row=4, column=0, columnspan=3, sticky='w', padx=5, pady=5)
        self.plsda_group_frame.grid_remove() # Hidden by default
        
        ttk.Label(self.plsda_group_frame, text="Group 1:").pack(side='left', padx=2)
        self.plsda_group1_var = tk.StringVar()
        self.plsda_group1_combo = ttk.Combobox(self.plsda_group_frame, textvariable=self.plsda_group1_var, state='readonly', width=15)
        self.plsda_group1_combo.pack(side='left', padx=5)

        ttk.Label(self.plsda_group_frame, text="Group 2:").pack(side='left', padx=2)
        self.plsda_group2_var = tk.StringVar()
        self.plsda_group2_combo = ttk.Combobox(self.plsda_group_frame, textvariable=self.plsda_group2_var, state='readonly', width=15)
        self.plsda_group2_combo.pack(side='left', padx=5)
        
        # Button frame
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=5, column=0, columnspan=3, pady=10)
        ttk.Button(
            button_frame,
            text="▶️ Run PLS-DA",
            command=self.run_plsda
        ).pack(side='left', padx=5)
        ttk.Button(
            button_frame,
            text="💾 Export VIPs",
            command=self.export_plsda_vips
        ).pack(side='left', padx=5)
        
        # Results area
        results_frame = ttk.LabelFrame(tab, text="PLS-DA Results", padding=10)
        results_frame.pack(fill='x', padx=10, pady=5)
        
        self.plsda_results_text = scrolledtext.ScrolledText(
            results_frame,
            height=6,
            font=('Courier', 10)
        )
        self.plsda_results_text.pack(fill='both', expand=True)
        
        # Plot area
        self.plsda_plot_frame = ttk.Frame(tab)
        self.plsda_plot_frame.pack(fill='both', expand=True, padx=10, pady=5)

    def create_pls_val_tab(self):
        """Tab 6: PLS Validation (Permutation Test)"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="✅ PLS Validation")
        
        control_frame = ttk.LabelFrame(tab, text="Permutation Test Settings", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(control_frame, text="Number of Permutations:").pack(side='left', padx=5)
        self.n_perms_var = tk.IntVar(value=200)
        ttk.Spinbox(
            control_frame, from_=20, to=500, increment=20,
            textvariable=self.n_perms_var, width=10
        ).pack(side='left', padx=5)
        
        self.pls_val_exclude_qc_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            control_frame, text="Exclude QC Samples", variable=self.pls_val_exclude_qc_var
        ).pack(side='left', padx=10)
        
        ttk.Button(
            control_frame, text="▶️ Run Permutation Validation", command=self.run_pls_val
        ).pack(side='left', padx=10)
        
        self.pls_val_plot_frame = ttk.Frame(tab)
        self.pls_val_plot_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
    def create_rf_tab(self):
        """Tab 7: Random Forest"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🌲 Random Forest")
        
        # Controls
        control_frame = ttk.LabelFrame(tab, text="Random Forest Settings", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(control_frame, text="Number of trees:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.ntrees_var = tk.IntVar(value=500)
        ttk.Spinbox(
            control_frame,
            from_=100,
            to=1000,
            increment=100,
            textvariable=self.ntrees_var,
            width=10
        ).grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(control_frame, text="Top N features:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.top_n_var = tk.IntVar(value=20)
        ttk.Spinbox(
            control_frame,
            from_=10,
            to=50,
            increment=5,
            textvariable=self.top_n_var,
            width=10
        ).grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        self.rf_exclude_qc_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            control_frame, text="Exclude QC Samples", variable=self.rf_exclude_qc_var
        ).grid(row=2, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        # --- Pairwise Comparison Option ---
        self.rf_pairwise_var = tk.BooleanVar(value=False)
        
        def toggle_rf_pairwise():
            if self.rf_pairwise_var.get():
                self.rf_group_frame.grid()
                self.update_group_dropdowns()
            else:
                self.rf_group_frame.grid_remove()

        ttk.Checkbutton(
            control_frame,
            text="Pairwise Comparison (Select 2 Groups)",
            variable=self.rf_pairwise_var,
            command=toggle_rf_pairwise
        ).grid(row=3, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        self.rf_group_frame = ttk.Frame(control_frame)
        self.rf_group_frame.grid(row=4, column=0, columnspan=3, sticky='w', padx=5, pady=5)
        self.rf_group_frame.grid_remove() # Hidden by default
        
        ttk.Label(self.rf_group_frame, text="Group 1:").pack(side='left', padx=2)
        self.rf_group1_var = tk.StringVar()
        self.rf_group1_combo = ttk.Combobox(self.rf_group_frame, textvariable=self.rf_group1_var, state='readonly', width=15)
        self.rf_group1_combo.pack(side='left', padx=5)

        ttk.Label(self.rf_group_frame, text="Group 2:").pack(side='left', padx=2)
        self.rf_group2_var = tk.StringVar()
        self.rf_group2_combo = ttk.Combobox(self.rf_group_frame, textvariable=self.rf_group2_var, state='readonly', width=15)
        self.rf_group2_combo.pack(side='left', padx=5)
        
        self.rf_roc_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            control_frame, text="Generate ROC Curve (Pairwise only)", variable=self.rf_roc_var
        ).grid(row=5, column=0, columnspan=2, sticky='w', padx=5, pady=5)
        
        # Buttons
        btn_frame = ttk.Frame(control_frame)
        btn_frame.grid(row=6, column=0, columnspan=2, pady=10, sticky='w')
        
        ttk.Button(
            btn_frame,
            text="▶️ Run Random Forest",
            command=self.run_rf
        ).pack(side='left', padx=10)
        
        ttk.Button(
            btn_frame,
            text="💾 Export Features",
            command=self.export_rf_features
        ).pack(side='left', padx=5)
        
        # Results text
        results_frame = ttk.LabelFrame(tab, text="Random Forest Results", padding=10)
        results_frame.pack(fill='x', padx=10, pady=5)
        
        self.rf_results_text = scrolledtext.ScrolledText(
            results_frame,
            height=8,
            font=('Courier', 10)
        )
        self.rf_results_text.pack(fill='both', expand=True)
        
        # Plot area
        self.rf_plot_frame = ttk.Frame(tab)
        self.rf_plot_frame.pack(fill='both', expand=True, padx=10, pady=5)

    def create_heatmap_tab(self):
        """Tab 8: Heatmap Visualization"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔥 Heatmap (HCA)")
        
        control_frame = ttk.LabelFrame(tab, text="Heatmap Settings", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)
        
        # Pack right-side buttons first
        ttk.Button(
            control_frame, text="▶️ Create Clustered Heatmap", command=self.run_heatmap
        ).pack(side='right', padx=10)
        ttk.Button(
            control_frame, text="💾 Export Heatmap Data", command=self.export_heatmap_data
        ).pack(side='right', padx=5)

        # Pack left-side controls sequentially
        ttk.Label(control_frame, text="Top N (per model):").pack(side='left', padx=(5, 2))
        self.heatmap_top_n_var = tk.IntVar(value=15)
        ttk.Spinbox(
            control_frame,
            from_=5,
            to=50,
            increment=5,
            textvariable=self.heatmap_top_n_var,
            width=5
        ).pack(side='left', padx=2)
        
        ttk.Label(control_frame, text="Sort Features By:").pack(side='left', padx=(10, 2))
        self.heatmap_sort_var = tk.StringVar(value="Clustering")
        ttk.Combobox(control_frame, textvariable=self.heatmap_sort_var, 
                     values=["Retention Time", "Clustering"], state='readonly', width=14).pack(side='left', padx=2)

        ttk.Label(control_frame, text="Color Map:").pack(side='left', padx=(10, 2))
        self.heatmap_cmap_var = tk.StringVar(value="coolwarm")
        ttk.Combobox(control_frame, textvariable=self.heatmap_cmap_var, 
                     values=["coolwarm", "bwr", "RdBu_r", "viridis", "magma"], state='readonly', width=10).pack(side='left', padx=2)
        
        self.heatmap_plot_frame = ttk.Frame(tab)
        self.heatmap_plot_frame.pack(fill='both', expand=True, padx=10, pady=5)

    def create_venn_tab(self):
        """Tab 9: Venn Diagram Analysis"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔵 Comparative Analysis")
        
        # Controls
        control_frame = ttk.LabelFrame(tab, text="Venn Diagram Settings", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)
        
        # Group Selection
        ttk.Label(control_frame, text="Select Groups (2 to 6):").grid(row=0, column=0, sticky='w', padx=5)
        
        # Listbox with scrollbar
        list_frame = ttk.Frame(control_frame)
        list_frame.grid(row=1, column=0, rowspan=3, padx=5, pady=5, sticky='ns')
        
        self.venn_group_listbox = tk.Listbox(list_frame, selectmode='multiple', height=5, exportselection=0, width=25)
        self.venn_group_listbox.pack(side='left', fill='y')
        sb = ttk.Scrollbar(list_frame, orient='vertical', command=self.venn_group_listbox.yview)
        sb.pack(side='right', fill='y')
        self.venn_group_listbox.config(yscrollcommand=sb.set)
        
        # Threshold Value
        ttk.Label(control_frame, text="Detection Frequency Threshold (%):").grid(row=1, column=1, sticky='w', padx=15)
        self.venn_thresh_var = tk.DoubleVar(value=80.0)
        ttk.Spinbox(control_frame, from_=0, to=100, increment=5, textvariable=self.venn_thresh_var, width=10).grid(row=2, column=1, sticky='nw', padx=15)
        
        self.venn_desc_label = ttk.Label(control_frame, text="(Feature must be > 0 in ≥ X% of group samples)", foreground="gray")
        self.venn_desc_label.grid(row=3, column=1, sticky='nw', padx=15)
        
        # Buttons
        btn_frame = ttk.Frame(control_frame)
        btn_frame.grid(row=1, column=2, rowspan=3, padx=20)
        
        ttk.Button(btn_frame, text="▶️ Generate Venn Diagram", command=self.run_venn_analysis).pack(fill='x', pady=2)
        ttk.Button(btn_frame, text="💾 Export Intersections", command=self.export_venn_intersections).pack(fill='x', pady=2)
        
        # Split Plot and Table
        self.venn_paned = ttk.PanedWindow(tab, orient=tk.VERTICAL)
        self.venn_paned.pack(fill='both', expand=True, padx=10, pady=5)

        # Plot Area
        self.venn_plot_frame = ttk.Frame(self.venn_paned)
        self.venn_paned.add(self.venn_plot_frame, weight=3)

        # Table Area
        self.venn_table_frame = ttk.LabelFrame(self.venn_paned, text="Intersection Features")
        self.venn_paned.add(self.venn_table_frame, weight=1)

        table_control = ttk.Frame(self.venn_table_frame)
        table_control.pack(fill='x', padx=5, pady=5)
        ttk.Label(table_control, text="Select Intersection:").pack(side='left', padx=5)
        
        self.venn_intersection_var = tk.StringVar()
        self.venn_intersection_combo = ttk.Combobox(table_control, textvariable=self.venn_intersection_var, state='readonly', width=50)
        self.venn_intersection_combo.pack(side='left', padx=5)
        self.venn_intersection_combo.bind('<<ComboboxSelected>>', self.update_venn_table)

        columns = ("ID", "m/z", "RT", "FeatureID")
        self.venn_tree = ttk.Treeview(self.venn_table_frame, columns=columns, show='headings', selectmode='browse')
        self.venn_tree.heading("ID", text="ID", command=lambda: self.treeview_sort_column(self.venn_tree, "ID", False))
        self.venn_tree.column("ID", width=120)
        self.venn_tree.heading("m/z", text="m/z", command=lambda: self.treeview_sort_column(self.venn_tree, "m/z", False))
        self.venn_tree.column("m/z", width=100)
        self.venn_tree.heading("RT", text="RT", command=lambda: self.treeview_sort_column(self.venn_tree, "RT", False))
        self.venn_tree.column("RT", width=100)
        self.venn_tree.heading("FeatureID", text="Full Feature ID", command=lambda: self.treeview_sort_column(self.venn_tree, "FeatureID", False))
        self.venn_tree.column("FeatureID", width=250)

        sb = ttk.Scrollbar(self.venn_table_frame, orient="vertical", command=self.venn_tree.yview)
        self.venn_tree.configure(yscrollcommand=sb.set)
        
        self.venn_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        sb.pack(side=tk.RIGHT, fill=tk.Y, pady=5)
        
        self.venn_tree.bind("<Double-1>", self.on_venn_table_double_click)
        
        # Update listbox on tab change
        self.notebook.bind('<<NotebookTabChanged>>', self.on_tab_change)

    def create_status_bar(self):
        """Create status bar at bottom"""
        self.status_bar = ttk.Label(
            self,
            text="Ready",
            relief=tk.SUNKEN,
            anchor=tk.W,
            padding=(5, 2)
        )
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
    # ==================== Data Loading & Renaming ====================
    
    def on_tab_change(self, event):
        """Handle tab change events"""
        selected_tab = self.notebook.select()
        tab_text = self.notebook.tab(selected_tab, "text")
        
        if "Comparative" in tab_text or "Venn" in tab_text:
            self.update_venn_groups()
        elif "UpSet" in tab_text:
            self.update_upset_groups()
        elif "PCA" in tab_text:
            self.update_pca_groups()
        elif "Feature Viewer" in tab_text:
            self.update_feature_viewer_groups()
            
    def load_csv(self):
        """Load CSV file"""
        file_path = filedialog.askopenfilename(
            title="Select MZmine Quantification CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
            
        try:
            self.update_status("Loading CSV file...")
            self.data = pd.read_csv(file_path)
            
            # --- NEW: Create custom Feature ID format -> ID_m/z_RT ---
            col_id = next((c for c in self.data.columns if 'id' in c.lower() and 'row' in c.lower()), 'row ID')
            col_mz = next((c for c in self.data.columns if 'm/z' in c.lower()), 'row m/z')
            col_rt = next((c for c in self.data.columns if 'retention time' in c.lower() or 'rt' in c.lower().split()), 'row retention time')
            
            if col_id in self.data.columns and col_mz in self.data.columns and col_rt in self.data.columns:
                def make_id(row):
                    try:
                        rid = str(row[col_id]).split('.')[0]
                        mz = float(row[col_mz])
                        rt = float(row[col_rt])
                        return f"{rid}_{mz:.2f}_{rt:.2f}"
                    except:
                        return str(row[col_id])
                
                feature_ids = self.data.apply(make_id, axis=1)
                
                # Prevent duplicate Feature_ID columns if loaded multiple times
                if 'Feature_ID' in self.data.columns:
                    self.data.drop('Feature_ID', axis=1, inplace=True)
                
                self.data.insert(0, 'Feature_ID', feature_ids)
            
            # Detect and store replicate structure
            self.detect_replicates()
            
            # Initialize groups as "Ungrouped" - user must manually configure
            self.initialize_default_groups()
            
            # Auto-detect QC samples based on group name containing "QC"
            self.auto_detect_qc_samples()
            
            # Update file label
            filename = os.path.basename(file_path)
            self.file_label.config(text=f"Loaded: {filename}", foreground="green")
            
            # Display summary
            self.display_data_summary()
            
            # Display preview
            self.display_data_preview()
            
            # Populate Feature Viewer with all raw features
            self.update_feature_viewer_options()
            
            self.update_status(f"✓ Loaded {len(self.data)} features from {filename}")
            
            # Prompt user to configure groups
            messagebox.showinfo(
                "Data Loaded", 
                f"✓ Loaded {len(self.data)} features successfully!\n\n"
                f"📋 Detected {len(self.sample_names)} unique samples\n\n"
                "⚙️ Next Steps:\n"
                "   • (Optional) Rename your samples using the 'Rename Samples' button.\n"
                "   • Configure sample groups by clicking 'Configure Groups'."
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load CSV:\n{str(e)}")
            self.update_status("Error loading file")
    
    def detect_replicates(self):
        """
        Detect technical replicates based on naming pattern
        Format: any pattern ending with _1, _2, etc.
        """
        if self.data is None:
            return
        
        # Find all sample columns (containing .mzML)
        sample_cols = [col for col in self.data.columns if '.mzML' in col]
        
        # Pattern: extract base name and replicate number
        # More flexible pattern that works with any prefix
        pattern = r'^(.+?)_(\d+)\.mzML'
        
        replicate_groups = defaultdict(list)
        
        for col in sample_cols:
            match = re.match(pattern, col)
            if match:
                base_name = match.group(1)
                rep_num = match.group(2)
                replicate_groups[base_name].append((col, int(rep_num)))
            else:
                # If no replicate pattern, treat as single sample
                base_name = col.replace('.mzML', '').replace(' Peak area', '')
                replicate_groups[base_name].append((col, 1))
        
        # Sort replicates within each group
        for base_name in replicate_groups:
            replicate_groups[base_name].sort(key=lambda x: x[1])
        
        self.replicate_mapping = dict(replicate_groups)
        
        # Apply Natural Alphanumeric Sorting for sample names
        self.sample_names = sorted(
            list(self.replicate_mapping.keys()), 
            key=lambda x: [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', x)]
        )
    
    def initialize_default_groups(self):
        """Initialize all samples as 'Ungrouped' - user must manually configure"""
        if not self.sample_names:
            return
        
        # Set all samples to "Ungrouped" initially
        self.group_mapping = {}
        for sample in self.sample_names:
            self.group_mapping[sample] = "Ungrouped"
        
        # Default color for ungrouped
        self.group_colors = {"Ungrouped": "#CCCCCC"}

    def rename_samples_dialog(self):
        """Dialog to rename MZmine sample identifiers across the dataset"""
        if not self.sample_names:
            messagebox.showwarning("Warning", "Please load data first!")
            return
            
        rename_window = tk.Toplevel(self)
        rename_window.title("Rename Samples")
        rename_window.geometry("600x500")
        
        # Instructions
        instruction_frame = ttk.Frame(rename_window)
        instruction_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(
            instruction_frame,
            text="🏷️ Rename Sample IDs:",
            font=('Arial', 12, 'bold')
        ).pack(anchor='w')
        
        ttk.Label(
            instruction_frame,
            text="Type a new name for any sample. This will automatically update all \n"
                 "underlying CSV column names and technical duplicate mappings.",
            font=('Arial', 9),
            foreground='#555555'
        ).pack(anchor='w', pady=5)
        
        # Scrollable list
        canvas = tk.Canvas(rename_window)
        scrollbar = ttk.Scrollbar(rename_window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Header
        header_frame = ttk.Frame(scrollable_frame)
        header_frame.pack(fill='x', padx=10, pady=5)
        ttk.Label(header_frame, text="Current MZmine ID", width=30, font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5)
        ttk.Label(header_frame, text="→", width=5, font=('Arial', 10, 'bold')).grid(row=0, column=1)
        ttk.Label(header_frame, text="New Custom ID", width=30, font=('Arial', 10, 'bold')).grid(row=0, column=2, padx=5)
        
        # Entries
        rename_vars = {}
        for sample in self.sample_names:
            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill='x', padx=10, pady=2)
            
            ttk.Label(row_frame, text=sample, width=30).grid(row=0, column=0, padx=5, sticky='w')
            ttk.Label(row_frame, text="→", width=5).grid(row=0, column=1)
            
            new_name_var = tk.StringVar(value=sample)
            rename_vars[sample] = new_name_var
            ttk.Entry(row_frame, textvariable=new_name_var, width=30).grid(row=0, column=2, padx=5)
            
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        def apply_sample_rename():
            """Execute renaming logic across the dataframe columns and mappings"""
            rename_map = {}
            for old_name, new_name_var in rename_vars.items():
                new_name = new_name_var.get().strip()
                if new_name and new_name != old_name:
                    rename_map[old_name] = new_name
            
            if not rename_map:
                messagebox.showinfo("Info", "No name changes detected.")
                return
                
            # Dictionary for pandas rename
            col_rename_dict = {}
            
            for old_name, new_name in rename_map.items():
                reps = self.replicate_mapping[old_name]
                new_reps = []
                
                # Fix column names while retaining technical suffixes like "_1.mzML Peak area"
                for col_name, rep_num in reps:
                    new_col_name = col_name.replace(old_name, new_name)
                    col_rename_dict[col_name] = new_col_name
                    new_reps.append((new_col_name, rep_num))
                    
                # Update Mapping Dicts
                self.replicate_mapping[new_name] = new_reps
                del self.replicate_mapping[old_name]
                
                if old_name in self.group_mapping:
                    self.group_mapping[new_name] = self.group_mapping.pop(old_name)
                    
                if old_name in self.qc_samples:
                    self.qc_samples.remove(old_name)
                    self.qc_samples.append(new_name)
                    
            # Actually rename columns in the core DataFrame so everything else stays intact
            self.data.rename(columns=col_rename_dict, inplace=True)
            
            # Apply Natural Alphanumeric Sorting after renaming
            self.sample_names = sorted(
                list(self.replicate_mapping.keys()), 
                key=lambda x: [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', x)]
            )
            
            self.display_data_summary()
            self.display_data_preview()
            
            messagebox.showinfo("Success", f"✓ Renamed {len(rename_map)} sample(s) successfully!")
            rename_window.destroy()

        button_frame = ttk.Frame(rename_window)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(button_frame, text="💾 Apply Name Changes", command=apply_sample_rename).pack(side='left', padx=5)
        ttk.Button(button_frame, text="❌ Cancel", command=rename_window.destroy).pack(side='left', padx=5)
        
    def auto_detect_qc_samples(self):
        """
        Auto-detect QC samples based on group name containing 'QC' (case-insensitive)
        """
        if not self.group_mapping:
            return
        
        qc_samples_auto = []
        for sample, group in self.group_mapping.items():
            if 'qc' in group.lower():
                qc_samples_auto.append(sample)
        
        if qc_samples_auto:
            self.qc_samples = qc_samples_auto
            self.update_qc_status_label()
            
    def update_venn_groups(self):
        """Update the group list in the Venn tab"""
        if hasattr(self, 'venn_group_listbox') and self.group_mapping:
            current_selection = [self.venn_group_listbox.get(i) for i in self.venn_group_listbox.curselection()]
            self.venn_group_listbox.delete(0, tk.END)
            
            unique_groups = sorted(list(set(self.group_mapping.values())))
            for g in unique_groups:
                self.venn_group_listbox.insert(tk.END, g)
                
    def update_upset_groups(self):
        """Update the group list in the UpSet tab"""
        if hasattr(self, 'upset_group_listbox') and self.group_mapping:
            self.upset_group_listbox.delete(0, tk.END)
            unique_groups = sorted(list(set(self.group_mapping.values())))
            for g in unique_groups:
                self.upset_group_listbox.insert(tk.END, g)
                
    def update_pca_groups(self):
        """Update the group list in the PCA tab"""
        if hasattr(self, 'pca_group_listbox') and self.group_mapping:
            current_selection = [self.pca_group_listbox.get(i) for i in self.pca_group_listbox.curselection()]
            self.pca_group_listbox.delete(0, tk.END)
            
            unique_groups = sorted(list(set(self.group_mapping.values())))
            for i, g in enumerate(unique_groups):
                self.pca_group_listbox.insert(tk.END, g)
                if g in current_selection:
                    self.pca_group_listbox.selection_set(i)
            
    def update_feature_viewer_groups(self):
        """Update the group list in the Feature Viewer tab"""
        if hasattr(self, 'fv_group_listbox') and self.group_mapping:
            current_selection = [self.fv_group_listbox.get(i) for i in self.fv_group_listbox.curselection()]
            self.fv_group_listbox.delete(0, tk.END)
            
            unique_groups = sorted(list(set(self.group_mapping.values())))
            for i, g in enumerate(unique_groups):
                self.fv_group_listbox.insert(tk.END, g)
                if g in current_selection:
                    self.fv_group_listbox.selection_set(i)
            
    def update_group_dropdowns(self):
        """Update the group selection dropdowns in the Volcano Plot tab, excluding QCs"""
        if hasattr(self, 'volcano_group1_combo') and hasattr(self, 'volcano_group2_combo'):
            if self.group_mapping:
                unique_groups = sorted(set(self.group_mapping.values()))
                
                # Exclude any group containing 'qc' from the Univariate Screening dropdowns
                non_qc_groups = [g for g in unique_groups if 'qc' not in g.lower() and g != 'Ungrouped']
                
                self.volcano_group1_combo['values'] = non_qc_groups
                self.volcano_group2_combo['values'] = non_qc_groups
                
                # Try to auto-select if currently empty or invalid
                current_g1 = self.volcano_group1_var.get()
                current_g2 = self.volcano_group2_var.get()
                
                if current_g1 not in non_qc_groups or current_g2 not in non_qc_groups:
                    if len(non_qc_groups) >= 2:
                        self.volcano_group1_var.set(non_qc_groups[0])
                        self.volcano_group2_var.set(non_qc_groups[1])
                    elif len(non_qc_groups) == 1:
                        self.volcano_group1_var.set(non_qc_groups[0])
                        self.volcano_group2_var.set('')
                
                # --- Update PLS-DA Pairwise Dropdowns ---
                if hasattr(self, 'plsda_group1_combo') and hasattr(self, 'plsda_group2_combo'):
                    self.plsda_group1_combo['values'] = non_qc_groups
                    self.plsda_group2_combo['values'] = non_qc_groups
                    
                    # Auto-select for PLS-DA if unset or invalid
                    p_g1 = self.plsda_group1_var.get()
                    p_g2 = self.plsda_group2_var.get()
                    
                    if p_g1 not in non_qc_groups or p_g2 not in non_qc_groups:
                        if len(non_qc_groups) >= 2:
                            self.plsda_group1_var.set(non_qc_groups[0])
                            self.plsda_group2_var.set(non_qc_groups[1])
                        elif len(non_qc_groups) == 1:
                            self.plsda_group1_var.set(non_qc_groups[0])

                # --- Update RF Pairwise Dropdowns ---
                if hasattr(self, 'rf_group1_combo') and hasattr(self, 'rf_group2_combo'):
                    self.rf_group1_combo['values'] = non_qc_groups
                    self.rf_group2_combo['values'] = non_qc_groups
                    
                    # Auto-select for RF if unset or invalid
                    r_g1 = self.rf_group1_var.get()
                    r_g2 = self.rf_group2_var.get()
                    
                    if r_g1 not in non_qc_groups or r_g2 not in non_qc_groups:
                        if len(non_qc_groups) >= 2:
                            self.rf_group1_var.set(non_qc_groups[0])
                            self.rf_group2_var.set(non_qc_groups[1])
                        elif len(non_qc_groups) == 1:
                            self.rf_group1_var.set(non_qc_groups[0])

    def display_data_summary(self):
        """Display data summary statistics"""
        if self.data is None:
            return
            
        summary = []
        summary.append("=" * 60)
        summary.append("DATA SUMMARY")
        summary.append("=" * 60)
        summary.append(f"Total Features: {len(self.data)}")
        summary.append(f"Total Columns: {len(self.data.columns)}")
        
        sample_cols = [col for col in self.data.columns if '.mzML' in col]
        summary.append(f"Sample Columns: {len(sample_cols)}")
        
        if self.replicate_mapping:
            summary.append(f"\nUnique Samples: {len(self.replicate_mapping)}")
            rep_counts = [len(reps) for reps in self.replicate_mapping.values()]
            summary.append(f"Replicates per sample: {min(rep_counts)}-{max(rep_counts)}")
        
        if self.group_mapping:
            unique_groups = set(self.group_mapping.values())
            summary.append(f"\nCurrent Groups: {len(unique_groups)}")
            for group_name in sorted(unique_groups):
                group_samples = [s for s, g in self.group_mapping.items() if g == group_name]
                summary.append(f"  {group_name}: {len(group_samples)} samples")
        
        if self.qc_samples:
            summary.append(f"\nQC Samples: {len(self.qc_samples)}")
            for qc in self.qc_samples[:5]:  # Show first 5
                summary.append(f"  • {qc}")
            if len(self.qc_samples) > 5:
                summary.append(f"  ... and {len(self.qc_samples)-5} more")
        
        summary.append("\n" + "=" * 60)
        summary.append("⚠️  Please configure sample groups before analysis!")
        summary.append("=" * 60)
        
        self.summary_text.delete('1.0', tk.END)
        self.summary_text.insert('1.0', '\n'.join(summary))
        
        # Update dropdowns whenever summary updates (handles group renames/reconfigs)
        self.update_group_dropdowns()
        self.update_venn_groups()
        self.update_upset_groups()
        self.update_pca_groups()
        self.update_feature_viewer_groups()
        
    def display_data_preview(self):
        """Display first 10 rows"""
        if self.data is None:
            return
            
        self.preview_tree.delete(*self.preview_tree.get_children())
        preview_data = self.data.head(10).iloc[:, :10]
        
        self.preview_tree['columns'] = list(preview_data.columns)
        for col in preview_data.columns:
            self.preview_tree.heading(col, text=col, command=lambda _col=col: self.treeview_sort_column(self.preview_tree, _col, False))
            self.preview_tree.column(col, width=100, anchor='center')
        
        for idx, row in preview_data.iterrows():
            self.preview_tree.insert('', 'end', values=list(row))
    
    # ==================== QC Sample Selection ====================
    
    def on_rsd_filter_toggle(self):
        """Show/hide QC selection button when RSD filter is toggled"""
        if self.filter_rsd_var.get():
            self.qc_select_frame.grid()
        else:
            self.qc_select_frame.grid_remove()
    
    def select_qc_samples(self):
        """Dialog to select QC samples"""
        if not self.sample_names:
            messagebox.showwarning("Warning", "Please load data first!")
            return
        
        qc_window = tk.Toplevel(self)
        qc_window.title("Select QC Samples")
        qc_window.geometry("700x600")
        
        # Instructions
        instruction_frame = ttk.Frame(qc_window)
        instruction_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(
            instruction_frame,
            text="🔬 Select QC Samples for RSD Filtering:",
            font=('Arial', 12, 'bold')
        ).pack(anchor='w')
        
        ttk.Label(
            instruction_frame,
            text="Check the samples that should be used as Quality Control (QC) samples.\n"
                 "QC samples are used to filter features with high technical variation (RSD > threshold).",
            font=('Arial', 9),
            foreground='#555555'
        ).pack(anchor='w', pady=5)
        
        # Auto-detect button
        button_frame = ttk.Frame(instruction_frame)
        button_frame.pack(fill='x', pady=5)
        
        ttk.Button(
            button_frame,
            text="🎯 Auto-Detect (groups containing 'QC')",
            command=lambda: self.auto_select_qc_in_dialog(qc_checkboxes)
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="✓ Select All",
            command=lambda: self.select_all_qc(qc_checkboxes, True)
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="✗ Deselect All",
            command=lambda: self.select_all_qc(qc_checkboxes, False)
        ).pack(side='left', padx=5)
        
        # Scrollable sample list
        canvas = tk.Canvas(qc_window)
        scrollbar = ttk.Scrollbar(qc_window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Create checkboxes for each sample
        qc_checkboxes = {}
        
        for i, sample in enumerate(self.sample_names):
            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill='x', padx=10, pady=2)
            
            # Checkbox variable
            var = tk.BooleanVar(value=(sample in self.qc_samples))
            qc_checkboxes[sample] = var
            
            # Get group name for display
            group = self.group_mapping.get(sample, "Ungrouped")
            
            # Checkbox
            cb = ttk.Checkbutton(
                row_frame,
                text=f"{sample} (Group: {group})",
                variable=var
            )
            cb.pack(side='left')
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        def save_qc_selection():
            """Save QC sample selection"""
            self.qc_samples = [sample for sample, var in qc_checkboxes.items() if var.get()]
            
            if len(self.qc_samples) == 0:
                messagebox.showwarning(
                    "Warning",
                    "No QC samples selected!\n\n"
                    "QC RSD filtering will be skipped if no QC samples are selected."
                )
            
            # Update status label
            self.update_qc_status_label()
            
            # Update summary
            self.display_data_summary()
            
            messagebox.showinfo("Success", f"✓ Selected {len(self.qc_samples)} QC sample(s)")
            qc_window.destroy()
        
        # Save/Cancel buttons
        button_frame = ttk.Frame(qc_window)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(
            button_frame,
            text="💾 Save Selection",
            command=save_qc_selection
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="❌ Cancel",
            command=qc_window.destroy
        ).pack(side='left', padx=5)
    
    def auto_select_qc_in_dialog(self, qc_checkboxes):
        """Auto-select samples whose group name contains 'QC'"""
        count = 0
        for sample, var in qc_checkboxes.items():
            group = self.group_mapping.get(sample, "")
            if 'qc' in group.lower():
                var.set(True)
                count += 1
            else:
                var.set(False)
        
        messagebox.showinfo("Auto-Detect", f"Auto-selected {count} sample(s) from groups containing 'QC'")
    
    def select_all_qc(self, qc_checkboxes, select):
        """Select or deselect all QC samples"""
        for var in qc_checkboxes.values():
            var.set(select)
    
    def update_qc_status_label(self):
        """Update QC status label in preprocessing tab"""
        if hasattr(self, 'qc_status_label'):
            if self.qc_samples:
                self.qc_status_label.config(
                    text=f"{len(self.qc_samples)} QC sample(s) selected",
                    foreground="green"
                )
            else:
                self.qc_status_label.config(
                    text="No QC samples selected",
                    foreground="gray"
                )
    
    # ==================== Group Configuration ====================
    
    def configure_groups(self):
        """Open dialog to configure sample groups and colors"""
        if not self.sample_names:
            messagebox.showwarning("Warning", "Please load data first!")
            return
        
        config_window = tk.Toplevel(self)
        config_window.title("Configure Sample Groups")
        config_window.geometry("900x700")
        
        # Instructions
        instruction_frame = ttk.Frame(config_window)
        instruction_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(
            instruction_frame,
            text="📋 Assign Samples to Groups:",
            font=('Arial', 12, 'bold')
        ).pack(anchor='w')
        
        ttk.Label(
            instruction_frame,
            text="1. Select group from dropdown for each sample\n"
                 "2. Type new group name in dropdown to create new group\n"
                 "3. Click color button to customize group color\n"
                 "4. Click 'Save' when finished\n"
                 "💡 Tip: Name QC samples' group as 'QC' for auto-detection",
            font=('Arial', 9),
            foreground='#555555'
        ).pack(anchor='w', pady=5)
        
        # Create scrollable area
        canvas = tk.Canvas(config_window)
        scrollbar = ttk.Scrollbar(config_window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Header
        header_frame = ttk.Frame(scrollable_frame)
        header_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(header_frame, text="Sample Name", width=35, font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5)
        ttk.Label(header_frame, text="Assign to Group", width=20, font=('Arial', 10, 'bold')).grid(row=0, column=1, padx=5)
        ttk.Label(header_frame, text="Group Color", width=12, font=('Arial', 10, 'bold')).grid(row=0, column=2, padx=5)
        
        # Get unique existing groups
        unique_groups = sorted(set(self.group_mapping.values()))
        
        group_vars = {}
        color_buttons = {}
        
        for i, sample in enumerate(self.sample_names):
            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill='x', padx=10, pady=2)
            
            # Sample name
            ttk.Label(row_frame, text=sample, width=35).grid(row=0, column=0, padx=5, sticky='w')
            
            # Group dropdown (editable combobox)
            current_group = self.group_mapping.get(sample, "Ungrouped")
            group_var = tk.StringVar(value=current_group)
            group_vars[sample] = group_var
            
            group_combo = ttk.Combobox(
                row_frame,
                textvariable=group_var,
                values=unique_groups,
                width=20
            )
            group_combo.grid(row=0, column=1, padx=5)
            
            # Update unique_groups when new value is entered
            def update_groups(event, combo=group_combo):
                current_values = list(combo['values'])
                new_value = combo.get()
                if new_value and new_value not in current_values:
                    current_values.append(new_value)
                    combo['values'] = current_values
                    # Assign default color to new group
                    if new_value not in self.group_colors:
                        default_colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', 
                                         '#F7DC6F', '#BB8FCE', '#85C1E2', '#F8B739', '#52B788']
                        self.group_colors[new_value] = default_colors[len(self.group_colors) % len(default_colors)]
            
            group_combo.bind('<<ComboboxSelected>>', update_groups)
            group_combo.bind('<Return>', update_groups)
            
            # Color button
            group_color = self.group_colors.get(current_group, '#CCCCCC')
            color_btn = tk.Button(
                row_frame,
                text="    ",
                width=8,
                bg=group_color,
                command=lambda g=current_group, btn=None: self.choose_group_color_inline(g, config_window, group_vars, color_buttons)
            )
            color_btn.grid(row=0, column=2, padx=5)
            color_buttons[sample] = color_btn
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        def save_groups():
            """Save group configuration"""
            # Update group mapping
            for sample, group_var in group_vars.items():
                new_group = group_var.get().strip()
                if new_group:
                    self.group_mapping[sample] = new_group
                else:
                    self.group_mapping[sample] = "Ungrouped"
            
            # Ensure all groups have colors
            for group in set(self.group_mapping.values()):
                if group not in self.group_colors:
                    default_colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8']
                    self.group_colors[group] = default_colors[len(self.group_colors) % len(default_colors)]
            
            # Auto-detect QC samples after group assignment
            self.auto_detect_qc_samples()
            
            # Update summary
            self.display_data_summary()
            
            messagebox.showinfo("Success", "✓ Group configuration saved successfully!")
            config_window.destroy()
        
        # Button frame
        button_frame = ttk.Frame(config_window)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(
            button_frame,
            text="💾 Save Configuration",
            command=save_groups
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame,
            text="❌ Cancel",
            command=config_window.destroy
        ).pack(side='left', padx=5)
    
    def choose_group_color_inline(self, group_name, parent_window, group_vars, color_buttons):
        """Choose color for a group"""
        current_color = self.group_colors.get(group_name, '#4ECDC4')
        color = colorchooser.askcolor(
            title=f"Choose color for {group_name}",
            initialcolor=current_color,
            parent=parent_window
        )
        
        if color[1]:
            self.group_colors[group_name] = color[1]
            # Update all buttons with this group
            for sample, btn in color_buttons.items():
                if group_vars[sample].get() == group_name:
                    btn.config(bg=color[1])
    
    def manage_group_names(self):
        """Dialog to rename existing groups"""
        if not self.group_mapping:
            messagebox.showwarning("Warning", "Please configure groups first!")
            return
        
        manage_window = tk.Toplevel(self)
        manage_window.title("Manage Group Names")
        manage_window.geometry("600x400")
        
        ttk.Label(
            manage_window,
            text="✏️ Rename Groups:",
            font=('Arial', 12, 'bold')
        ).pack(pady=10)
        
        # Get unique groups
        unique_groups = sorted(set(self.group_mapping.values()))
        
        # Create frame for group entries
        groups_frame = ttk.Frame(manage_window)
        groups_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Header
        ttk.Label(groups_frame, text="Current Name", font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=10, pady=5)
        ttk.Label(groups_frame, text="→", font=('Arial', 10, 'bold')).grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(groups_frame, text="New Name", font=('Arial', 10, 'bold')).grid(row=0, column=2, padx=10, pady=5)
        ttk.Label(groups_frame, text="Sample Count", font=('Arial', 10, 'bold')).grid(row=0, column=3, padx=10, pady=5)
        
        rename_vars = {}
        
        for idx, group in enumerate(unique_groups, start=1):
            sample_count = sum(1 for g in self.group_mapping.values() if g == group)
            
            ttk.Label(groups_frame, text=group).grid(row=idx, column=0, padx=10, pady=5, sticky='w')
            ttk.Label(groups_frame, text="→").grid(row=idx, column=1, padx=5, pady=5)
            
            new_name_var = tk.StringVar(value=group)
            rename_vars[group] = new_name_var
            ttk.Entry(groups_frame, textvariable=new_name_var, width=25).grid(row=idx, column=2, padx=10, pady=5)
            
            ttk.Label(groups_frame, text=f"{sample_count} samples").grid(row=idx, column=3, padx=10, pady=5)
        
        def apply_rename():
            """Apply group name changes"""
            rename_map = {}
            for old_name, new_name_var in rename_vars.items():
                new_name = new_name_var.get().strip()
                if new_name and new_name != old_name:
                    rename_map[old_name] = new_name
            
            if not rename_map:
                messagebox.showinfo("Info", "No changes to apply")
                return
            
            # Update group_mapping
            for sample, group in self.group_mapping.items():
                if group in rename_map:
                    self.group_mapping[sample] = rename_map[group]
            
            # Update group_colors
            new_colors = {}
            for old_group, new_group in rename_map.items():
                if old_group in self.group_colors:
                    new_colors[new_group] = self.group_colors[old_group]
            
            # Remove old colors and add new
            for old_group in rename_map.keys():
                if old_group in self.group_colors:
                    del self.group_colors[old_group]
            self.group_colors.update(new_colors)
            
            # Auto-detect QC samples after renaming
            self.auto_detect_qc_samples()
            
            # Update summary
            self.display_data_summary()
            
            messagebox.showinfo("Success", f"✓ Renamed {len(rename_map)} group(s) successfully!")
            manage_window.destroy()
        
        # Buttons
        button_frame = ttk.Frame(manage_window)
        button_frame.pack(fill='x', padx=20, pady=10)
        
        ttk.Button(button_frame, text="✓ Apply Changes", command=apply_rename).pack(side='left', padx=5)
        ttk.Button(button_frame, text="❌ Cancel", command=manage_window.destroy).pack(side='left', padx=5)
    
    # ==================== Preprocessing ====================
    
    def average_technical_replicates_keep_original(self, data):
        """Keep original replicates AND add averaged value (skips QCs)"""
        if not self.replicate_mapping:
            return data
        
        sample_cols = [col for col in data.columns if '.mzML' in col]
        metadata_cols = [col for col in data.columns if col not in sample_cols]
        
        result_df = data[metadata_cols].copy()
        
        # Keep all original replicates
        for col in sample_cols:
            result_df[col] = data[col]
        
        # Add averaged replicates ONLY for non-QC samples
        for base_name, replicates in self.replicate_mapping.items():
            if base_name in self.qc_samples:
                continue # Do not average QC samples
                
            rep_columns = [col for col, _ in replicates]
            existing_cols = [col for col in rep_columns if col in data.columns]
            
            if len(existing_cols) >= 2:
                avg_col_name = f"{base_name}_avg.mzML Peak area"
                result_df[avg_col_name] = data[existing_cols].mean(axis=1)
        
        return result_df
    
    def apply_lod_imputation(self, data, fraction=0.2):
        """Replace 0 and NaN with fraction of the minimum positive value per feature"""
        sample_cols = [col for col in data.columns if '.mzML' in col]
        metadata_cols = [col for col in data.columns if col not in sample_cols]
        
        result_df = data[metadata_cols].copy()
        sample_data = data[sample_cols].copy()
        
        # Replace 0 with NaN for easier handling
        sample_data = sample_data.replace(0, np.nan)
        
        # Find minimum positive value per feature (row)
        min_vals = sample_data.min(axis=1)
        
        # If a feature is all NaNs, min_val will be NaN. Fill with global min.
        global_min = sample_data.min().min()
        if pd.isna(global_min):
            global_min = 1.0 # fallback if dataset is entirely empty
        min_vals = min_vals.fillna(global_min)
        
        # Multiply min_vals by fraction and broadcast across columns to fill NaNs
        impute_vals = min_vals * fraction
        imputed_data = sample_data.T.fillna(impute_vals).T
        
        result_df = pd.concat([result_df, imputed_data], axis=1)
        return result_df

    def filter_by_detection_rate(self, data, threshold=80.0):
        """
        Filter features based on detection rate per group
        Reference: Gowda et al. (2014) Metabolomics
        Keep features detected in ≥threshold% of samples in at least one group
        """
        sample_cols = [col for col in data.columns if '.mzML' in col]
        metadata_cols = [col for col in data.columns if col not in sample_cols]
        
        # Group samples
        groups = {}
        for col in sample_cols:
            base_name = re.sub(r'_(\d+|avg)\.mzML.*', '', col)
            group = self.group_mapping.get(base_name, "Ungrouped")
            if group not in groups:
                groups[group] = []
            groups[group].append(col)
        
        # Calculate detection rate per group
        keep_features = []
        for idx in range(len(data)):
            feature_row = data.iloc[idx]
            keep = False
            
            for group_name, group_cols in groups.items():
                group_values = feature_row[group_cols].values
                detected = np.sum(group_values > 0)
                detection_rate = (detected / len(group_values)) * 100
                
                if detection_rate >= threshold:
                    keep = True
                    break
            
            if keep:
                keep_features.append(idx)
        
        filtered_data = data.iloc[keep_features].reset_index(drop=True)
        return filtered_data, len(data) - len(filtered_data)
    
    def filter_by_min_intensity(self, data, threshold=1000.0):
        """
        Filter features with mean intensity below threshold
        Reference: Xia et al. (2015) - MetaboAnalyst approach
        """
        sample_cols = [col for col in data.columns if '.mzML' in col]
        metadata_cols = [col for col in data.columns if col not in sample_cols]
        
        mean_intensities = data[sample_cols].mean(axis=1)
        keep_mask = mean_intensities >= threshold
        
        filtered_data = data[keep_mask].reset_index(drop=True)
        removed = len(data) - len(filtered_data)
        
        return filtered_data, removed
    
    def filter_by_qc_rsd(self, data, max_rsd=30.0):
        """
        Filter features with RSD > threshold in selected QC samples
        Reference: Dunn et al. (2011) Nature Protocols
        """
        sample_cols = [col for col in data.columns if '.mzML' in col]
        
        # Get QC sample columns based on user selection
        qc_columns = []
        for qc_sample in self.qc_samples:
            # Find all columns for this QC sample (including replicates)
            if qc_sample in self.replicate_mapping:
                reps = self.replicate_mapping[qc_sample]
                qc_cols = [col for col, _ in reps if col in sample_cols]
                qc_columns.extend(qc_cols)
            
            # Also check for averaged column (if it somehow exists)
            avg_col = f"{qc_sample}_avg.mzML Peak area"
            if avg_col in sample_cols:
                qc_columns.append(avg_col)
        
        if len(qc_columns) == 0:
            messagebox.showinfo(
                "Info",
                "No QC samples selected or QC columns not found.\n"
                "QC RSD filtering skipped.\n\n"
                "Please select QC samples using 'Select QC Samples' button."
            )
            return data, 0
        
        # Calculate RSD for each feature in QC samples
        keep_features = []
        for idx in range(len(data)):
            qc_values = data.iloc[idx][qc_columns].values
            qc_values = qc_values[qc_values > 0]  # Exclude zeros
            
            if len(qc_values) < 2:
                continue  # Can't calculate RSD with <2 values
            
            mean_val = np.mean(qc_values)
            std_val = np.std(qc_values, ddof=1)
            rsd = (std_val / mean_val * 100) if mean_val > 0 else 100
            
            if rsd <= max_rsd:
                keep_features.append(idx)
        
        filtered_data = data.iloc[keep_features].reset_index(drop=True)
        removed = len(data) - len(filtered_data)
        
        return filtered_data, removed
    
    def filter_by_iqr(self, data, factor=0.5):
        """
        Filter features with low interquartile range (low variance)
        Reference: Hackstadt & Hess (2009) BMC Bioinformatics
        Keep features with IQR > median_IQR × factor
        """
        sample_cols = [col for col in data.columns if '.mzML' in col]
        
        # Calculate IQR for each feature
        iqr_values = []
        for idx in range(len(data)):
            feature_values = data.iloc[idx][sample_cols].values
            feature_values = feature_values[feature_values > 0]
            
            if len(feature_values) > 0:
                q75, q25 = np.percentile(feature_values, [75, 25])
                iqr = q75 - q25
                iqr_values.append(iqr)
            else:
                iqr_values.append(0)
        
        iqr_values = np.array(iqr_values)
        median_iqr = np.median(iqr_values)
        threshold = median_iqr * factor
        
        keep_mask = iqr_values >= threshold
        filtered_data = data[keep_mask].reset_index(drop=True)
        removed = len(data) - len(filtered_data)
        
        return filtered_data, removed
    
    def normalize_data_local(self, data, method='tic'):
        """Apply normalization to sample columns"""
        sample_cols = [col for col in data.columns if '.mzML' in col]
        metadata_cols = [col for col in data.columns if col not in sample_cols]
        
        if len(sample_cols) == 0:
            return data
        
        result_df = data[metadata_cols].copy()
        sample_data = data[sample_cols].copy()
        
        method = method.lower()
        
        if method == 'tic':
            column_sums = sample_data.sum(axis=0)
            normalized = sample_data / column_sums * column_sums.mean()
        elif method == 'median':
            column_medians = sample_data.median(axis=0)
            normalized = sample_data / column_medians * column_medians.mean()
        elif method == 'log':
            normalized = np.log2(sample_data + 1)
        elif method == 'pareto':
            means = sample_data.mean(axis=0)
            stds = sample_data.std(axis=0)
            normalized = (sample_data - means) / np.sqrt(stds)
        elif method == 'auto':
            means = sample_data.mean(axis=0)
            stds = sample_data.std(axis=0)
            normalized = (sample_data - means) / stds
        else:
            normalized = sample_data
        
        result_df = pd.concat([result_df, normalized], axis=1)
        return result_df
    
    def run_preprocessing(self):
        """Run preprocessing pipeline with feature filtering"""
        if self.data is None:
            messagebox.showwarning("Warning", "Please load data first!")
            return
            
        try:
            self.update_status("Running preprocessing...")
            
            results = []
            results.append("=" * 60)
            results.append("PREPROCESSING RESULTS")
            results.append("=" * 60)
            
            working_data = self.data.copy()
            initial_features = len(working_data)
            results.append(f"\nInitial features: {initial_features}")
            
            # Step 1: Replicate averaging
            if self.avg_replicates_var.get():
                results.append("\n✓ Keeping original replicates + adding averaged values (Skipping QCs)...")
                working_data = self.average_technical_replicates_keep_original(working_data)
                results.append(f"  After adding averages: {len(working_data.columns)} columns")
            else:
                results.append("\n✗ Skipped replicate averaging")
            
            # Step 1.5: Missing Value Imputation
            if getattr(self, 'impute_lod_var', tk.BooleanVar(value=False)).get():
                fraction_str = self.lod_fraction_var.get()
                if "1/2" in fraction_str: fraction = 0.5
                elif "1/3" in fraction_str: fraction = 1.0 / 3.0
                elif "1/4" in fraction_str: fraction = 0.25
                else: fraction = 0.2
                
                results.append("\n" + "─" * 60)
                results.append("MISSING VALUE IMPUTATION")
                results.append("─" * 60)
                results.append(f"✓ Applying LOD Imputation (fraction={fraction:.2f})")
                
                # Count missing before
                sample_cols_temp = [c for c in working_data.columns if '.mzML' in c]
                zero_count = (working_data[sample_cols_temp] == 0).sum().sum()
                nan_count = working_data[sample_cols_temp].isna().sum().sum()
                results.append(f"  Found {zero_count} zeros and {nan_count} NaNs")
                
                working_data = self.apply_lod_imputation(working_data, fraction)
                results.append("  Imputation complete")
            
            # Step 2: Feature Filtering
            results.append("\n" + "─" * 60)
            results.append("FEATURE FILTERING")
            results.append("─" * 60)
            
            total_removed = 0
            
            # Detection rate filter
            if self.filter_detection_var.get():
                threshold = self.detection_threshold_var.get()
                working_data, removed = self.filter_by_detection_rate(working_data, threshold)
                total_removed += removed
                results.append(f"✓ Detection rate filter (≥{threshold}%): removed {removed} features")
                results.append(f"  Remaining: {len(working_data)} features")
            
            # Minimum intensity filter
            if self.filter_intensity_var.get():
                threshold = self.min_intensity_var.get()
                working_data, removed = self.filter_by_min_intensity(working_data, threshold)
                total_removed += removed
                results.append(f"✓ Min intensity filter (≥{threshold}): removed {removed} features")
                results.append(f"  Remaining: {len(working_data)} features")
            
            # QC RSD filter
            if self.filter_rsd_var.get():
                max_rsd = self.rsd_threshold_var.get()
                working_data, removed = self.filter_by_qc_rsd(working_data, max_rsd)
                total_removed += removed
                if removed > 0:
                    results.append(f"✓ QC RSD filter (≤{max_rsd}%): removed {removed} features")
                    results.append(f"  Using {len(self.qc_samples)} QC sample(s)")
                    results.append(f"  Remaining: {len(working_data)} features")
                else:
                    results.append(f"✗ QC RSD filter: skipped (no QC samples or insufficient data)")
            
            # IQR variance filter
            if self.filter_iqr_var.get():
                factor = self.iqr_factor_var.get()
                working_data, removed = self.filter_by_iqr(working_data, factor)
                total_removed += removed
                results.append(f"✓ IQR filter (factor={factor}): removed {removed} features")
                results.append(f"  Remaining: {len(working_data)} features")
            
            if total_removed == 0:
                results.append("✗ No feature filtering applied")
            else:
                results.append(f"\n📊 Total features removed: {total_removed} ({total_removed/initial_features*100:.1f}%)")
            
            # Store data before normalization
            self.data_before_norm = working_data.copy()
            
            # Step 3: Normalization
            results.append("\n" + "─" * 60)
            results.append("NORMALIZATION")
            results.append("─" * 60)
            
            norm_method = self.norm_method_var.get()
            if norm_method != 'None':
                results.append(f"✓ Applying {norm_method} normalization...")
                self.preprocessed_data = self.normalize_data_local(working_data, method=norm_method.lower())
                results.append(f"  Normalization complete")
            else:
                self.preprocessed_data = working_data.copy()
                results.append("✗ Skipped normalization")
            
            # Initialize screened data default to all preprocessed data
            self.screened_data = self.preprocessed_data.copy()
            
            # Final summary
            sample_cols_final = [col for col in self.preprocessed_data.columns if '.mzML' in col]
            results.append("\n" + "=" * 60)
            results.append("PREPROCESSING COMPLETE!")
            results.append("=" * 60)
            results.append(f"Final features: {len(self.preprocessed_data)}")
            results.append(f"Final shape: {self.preprocessed_data.shape}")
            results.append(f"Sample columns: {len(sample_cols_final)}")
            results.append("\n💡 Click 'Show Distribution Plots' to")
            results.append("   visualize normalization effects")
            
            self.preprocess_text.delete('1.0', tk.END)
            self.preprocess_text.insert('1.0', '\n'.join(results))
            
            # Enable distribution plot button
            self.show_dist_button.config(state='normal')
            self.update_feature_viewer_options()
            
            self.update_status("✓ Preprocessing complete")
            messagebox.showinfo(
                "Success",
                f"Preprocessing complete!\n\n"
                f"Features: {initial_features} → {len(self.preprocessed_data)}\n"
                f"Removed: {total_removed} ({total_removed/initial_features*100:.1f}%)"
            )
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"Preprocessing failed:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error in preprocessing")
    
    def show_distribution_plots(self):
        """Show before/after normalization distribution plots"""
        if self.data_before_norm is None or self.preprocessed_data is None:
            messagebox.showwarning("Warning", "Please run preprocessing first!")
            return

        try:
            self.update_status("Generating distribution plots...")
            self.vis_manager.create_distribution_plots()
            self.update_status("✓ Distribution plots generated")

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"Failed to generate plots:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error generating plots")

    # ==================== PCA Analysis ====================
    
    def run_pca(self):
        """Run PCA analysis"""
        if self.preprocessed_data is None:
            messagebox.showwarning("Warning", "Please run preprocessing first!")
            return
        
        # Check if groups are configured
        ungrouped_count = sum(1 for g in self.group_mapping.values() if g == "Ungrouped")
        if ungrouped_count > 0:
            response = messagebox.askyesno(
                "Warning",
                f"⚠️ {ungrouped_count} samples are still 'Ungrouped'.\n\n"
                "PCA plot works best with properly configured groups.\n\n"
                "Do you want to continue anyway?"
            )
            if not response:
                return
            
        try:
            self.update_status("Running PCA...")
            
            selected_indices = self.pca_group_listbox.curselection()
            selected_groups = [self.pca_group_listbox.get(i) for i in selected_indices]
            
            n_components = self.pca_components_var.get()
            pca_result = self.perform_pca_local(self.preprocessed_data, n_components, selected_groups)
            
            # Assign groups to PCA samples
            pca_result = self.assign_groups_to_pca(pca_result)
            
            # PERMANOVA Calculation
            if getattr(self, 'pca_permanova_var', None) and self.pca_permanova_var.get():
                pca_result['permanova_pvalue'] = self.calculate_permanova(
                    self.preprocessed_data, 
                    pca_result['sample_names'], 
                    pca_result['groups']
                )
            
            # PERMDISP Calculation
            if getattr(self, 'pca_permdisp_var', None) and self.pca_permdisp_var.get():
                pca_result['permdisp_pvalue'] = self.calculate_permdisp(
                    self.preprocessed_data, 
                    pca_result['sample_names'], 
                    pca_result['groups']
                )
            
            self.pca_result = pca_result
            self.create_pca_plot_with_groups(pca_result)
            self.vis_manager.create_pca_plot(pca_result)
            
            self.update_status("✓ PCA complete")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"PCA failed:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error in PCA")
    
    def perform_pca_local(self, data, n_components=2, selected_groups=None):
        """Built-in PCA implementation"""
        all_sample_cols = [col for col in data.columns if '.mzML' in col]
        
        if selected_groups:
            sample_cols = []
            for col in all_sample_cols:
                base_name = re.sub(r'_(\d+|avg)\.mzML.*', '', col)
                group = self.group_mapping.get(base_name, "Ungrouped")
                if group in selected_groups:
                    sample_cols.append(col)
            if len(sample_cols) < 2:
                raise ValueError("Not enough samples in selected groups for PCA.")
        else:
            sample_cols = all_sample_cols

        X = data[sample_cols].T
        
        # Robust imputation for zero/missing values
        X = X.replace(0, np.nan)
        min_val = X.min().min() * 0.2
        if pd.isna(min_val): min_val = 0.1
        X = X.fillna(min_val)
        
        # Optional Log Transform to fix extreme skewing
        if hasattr(self, 'pca_log_var') and self.pca_log_var.get():
            X = np.log10(X + 1)
        
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        
        pca = PCA(n_components=n_components)
        scores = pca.fit_transform(X_scaled)
        variance_explained = pca.explained_variance_ratio_ * 100
        
        return {
            'scores': scores,
            'variance_explained': variance_explained,
            'loadings': pca.components_,
            'sample_names': sample_cols,
            'model': pca
        }
    
    def assign_groups_to_pca(self, pca_result):
        """Assign group labels to PCA samples"""
        sample_names = pca_result['sample_names']
        groups = []
        
        for sample_col in sample_names:
            # Extract base name by removing replicate suffix and .mzML extension
            base_name = re.sub(r'_(\d+|avg)\.mzML.*', '', sample_col)
            
            # Find matching group using exact dictionary lookup to prevent prefix overlaps
            group = self.group_mapping.get(base_name, "Ungrouped")
            
            groups.append(group)
        
        pca_result['groups'] = groups
        return pca_result

    def calculate_permanova(self, data, sample_names, groups):
        """Calculate PERMANOVA p-value using scikit-bio."""
        try:
            import pandas as pd
            import numpy as np
            from skbio.stats.distance import permanova, DistanceMatrix
            from scipy.spatial.distance import pdist, squareform
        except ImportError:
            messagebox.showwarning("Missing Library", "The 'scikit-bio' library is required for PERMANOVA.\nPlease install it using: pip install scikit-bio")
            return None
            
        valid_indices = [i for i, g in enumerate(groups) if g != "Ungrouped"]
        if len(set([groups[i] for i in valid_indices])) < 2:
            return None
            
        valid_samples = [sample_names[i] for i in valid_indices]
        valid_groups = [groups[i] for i in valid_indices]
        
        X = data[valid_samples].T
        X = X.replace(0, np.nan)
        min_val = X.min().min() * 0.2
        if pd.isna(min_val): min_val = 0.1
        X = X.fillna(min_val)
        
        if hasattr(self, 'pca_log_var') and self.pca_log_var.get():
            X = np.log10(X + 1)
            
        distances = pdist(X, metric='euclidean')
        dist_matrix = squareform(distances)
        
        dm = DistanceMatrix(dist_matrix, ids=valid_samples)
        metadata = pd.DataFrame({'Group': valid_groups}, index=valid_samples)
        
        res = permanova(dm, metadata, column='Group', permutations=999)
        return res['p-value']
        
    def calculate_permdisp(self, data, sample_names, groups):
        """Calculate PERMDISP p-value using scikit-bio."""
        try:
            import pandas as pd
            import numpy as np
            from skbio.stats.distance import permdisp, DistanceMatrix
            from scipy.spatial.distance import pdist, squareform
        except ImportError:
            messagebox.showwarning("Missing Library", "The 'scikit-bio' library is required for PERMDISP.\nPlease install it using: pip install scikit-bio")
            return None
            
        valid_indices = [i for i, g in enumerate(groups) if g != "Ungrouped"]
        if len(set([groups[i] for i in valid_indices])) < 2:
            return None
            
        valid_samples = [sample_names[i] for i in valid_indices]
        valid_groups = [groups[i] for i in valid_indices]
        
        X = data[valid_samples].T
        X = X.replace(0, np.nan)
        min_val = X.min().min() * 0.2
        if pd.isna(min_val): min_val = 0.1
        X = X.fillna(min_val)
        
        if hasattr(self, 'pca_log_var') and self.pca_log_var.get():
            X = np.log10(X + 1)
            
        distances = pdist(X, metric='euclidean')
        dist_matrix = squareform(distances)
        
        dm = DistanceMatrix(dist_matrix, ids=valid_samples)
        metadata = pd.DataFrame({'Group': valid_groups}, index=valid_samples)
        
        try:
            res = permdisp(dm, metadata, column='Group', permutations=999)
            return res['p-value']
        except Exception as e:
            print(f"PERMDISP calculation error: {e}")
            return None
            
    def create_pca_plot_with_groups(self, pca_result):
        """Create PCA score plot with group colors, confidence ellipses, and hover tooltips"""
        for widget in self.pca_plot_frame.winfo_children():
            widget.destroy()
            
        w, h, dpi = self.plot_width_var.get(), self.plot_height_var.get(), self.plot_dpi_var.get()
        fig = Figure(figsize=(w, h), dpi=dpi)
        self.generated_plots['PCA'] = fig
        ax = fig.add_subplot(111)
        
        groups = pca_result.get('groups', None)
        show_ellipses = self.show_ellipses_var.get() if hasattr(self, 'show_ellipses_var') else True
        show_labels = self.show_labels_var.get() if hasattr(self, 'show_labels_var') else True
        
        scatters = [] # Keep track of scatter objects for hover interactions
        
        if groups:
            unique_groups = sorted(set(groups))
            
            for group in unique_groups:
                group_indices = [i for i, g in enumerate(groups) if g == group]
                group_scores_pc1 = pca_result['scores'][group_indices, 0]
                group_scores_pc2 = pca_result['scores'][group_indices, 1]
                
                color = self.group_colors.get(group, '#4ECDC4')
                
                sc = ax.scatter(
                    group_scores_pc1,
                    group_scores_pc2,
                    s=100,
                    alpha=0.7,
                    c=color,
                    edgecolors='black',
                    linewidth=1.5,
                    label=group,
                    zorder=3
                )
                scatters.append((sc, group_indices))
                
                if show_ellipses and len(group_indices) >= 3:
                    self.draw_confidence_ellipse(
                        ax, 
                        group_scores_pc1, 
                        group_scores_pc2,
                        color=color,
                        alpha=0.2,
                        edgecolor=color,
                        linewidth=2
                    )
            
            ax.legend(loc='best', framealpha=0.9, fontsize=10)
        else:
            sc = ax.scatter(
                pca_result['scores'][:, 0],
                pca_result['scores'][:, 1],
                s=100,
                alpha=0.7,
                edgecolors='black',
                linewidth=1.5
            )
            scatters.append((sc, list(range(len(pca_result['sample_names'])))))
        
        # Display static labels if explicitly requested (removed 30-sample limit)
        if show_labels:
            for i, name in enumerate(pca_result['sample_names']):
                short_name = re.sub(r'\.mzML.*', '', name)
                ax.annotate(
                    short_name,
                    (pca_result['scores'][i, 0], pca_result['scores'][i, 1]),
                    fontsize=7,
                    alpha=0.6,
                    zorder=2
                )
        
        title = "PCA Score Plot (Hover for Sample IDs)"
        if show_ellipses and groups:
            title += " with 95% Confidence Ellipses"
        
        if 'permanova_pvalue' in pca_result and pca_result['permanova_pvalue'] is not None:
            title += f"\nPERMANOVA p-value: {pca_result['permanova_pvalue']:.3f}"
        
        ax.set_xlabel(f"PC1 ({pca_result['variance_explained'][0]:.1f}%)", fontsize=12)
        ax.set_ylabel(f"PC2 ({pca_result['variance_explained'][1]:.1f}%)", fontsize=12)
        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3, linestyle='--', zorder=1)
        ax.axhline(y=0, color='k', linewidth=0.5, linestyle='--', alpha=0.3, zorder=1)
        ax.axvline(x=0, color='k', linewidth=0.5, linestyle='--', alpha=0.3, zorder=1)
        
        # Interactive Tooltip Logic for PCA
        annot = ax.annotate("", xy=(0,0), xytext=(10,10), textcoords="offset points",
                            bbox=dict(boxstyle="round", fc="yellow", alpha=0.9),
                            arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=0.1"), zorder=10)
        annot.set_visible(False)
        
        def on_hover(event):
            if event.inaxes == ax:
                is_contained = False
                for sc, indices in scatters:
                    cont, ind = sc.contains(event)
                    if cont:
                        idx = indices[ind["ind"][0]]
                        pos = sc.get_offsets()[ind["ind"][0]]
                        annot.xy = pos
                        
                        # Get exact sample col name and clean it up for the tooltip
                        sample_col = pca_result['sample_names'][idx]
                        clean_name = re.sub(r'\.mzML.*', '', sample_col)
                        group_name = pca_result['groups'][idx] if 'groups' in pca_result else 'Ungrouped'
                        
                        annot.set_text(f"ID: {clean_name}\nGrp: {group_name}")
                        annot.set_visible(True)
                        canvas.draw_idle()
                        is_contained = True
                        break
                
                if not is_contained and annot.get_visible():
                    annot.set_visible(False)
                    canvas.draw_idle()
        
        canvas = FigureCanvasTkAgg(fig, master=self.pca_plot_frame)
        canvas.draw()
        canvas.mpl_connect("motion_notify_event", on_hover)
        canvas.get_tk_widget().pack(fill='both', expand=True)
        
        toolbar = NavigationToolbar2Tk(canvas, self.pca_plot_frame)
        toolbar.update()
    
    def draw_confidence_ellipse(self, ax, x, y, n_std=2.4477, facecolor='none', **kwargs):
        """Draw confidence ellipse for a group of points"""
        if len(x) < 2:
            return
        
        cov = np.cov(x, y)
        eigenvalues, eigenvectors = np.linalg.eig(cov)
        
        order = eigenvalues.argsort()[::-1]
        eigenvalues = eigenvalues[order]
        eigenvectors = eigenvectors[:, order]
        
        angle = np.degrees(np.arctan2(eigenvectors[1, 0], eigenvectors[0, 0]))
        width, height = 2 * n_std * np.sqrt(eigenvalues)
        
        mean_x = np.mean(x)
        mean_y = np.mean(y)
        
        ellipse = Ellipse(
            xy=(mean_x, mean_y),
            width=width,
            height=height,
            angle=angle,
            facecolor=facecolor,
            **kwargs
        )
        
        ax.add_patch(ellipse)
        return ellipse

    # ==================== Volcano Plot & Univariate Screening ====================
    
    def run_volcano(self):
        """Run volcano plot analysis"""
        if self.preprocessed_data is None:
            messagebox.showwarning("Warning", "Please run preprocessing first!")
            return
            
        group1 = self.volcano_group1_var.get()
        group2 = self.volcano_group2_var.get()
        
        if not group1 or not group2:
            messagebox.showwarning("Warning", "Please select both Group 1 and Group 2 from the dropdowns!")
            return
            
        if group1 == group2:
            messagebox.showwarning("Warning", "Group 1 and Group 2 must be different!")
            return
        
        try:
            self.update_status("Creating volcano plot...")
            
            pval_thresh = self.pvalue_var.get()
            fc_thresh = self.fc_var.get()
            correction_method = self.correction_method_var.get()
            
            volcano_result = self.perform_volcano_analysis(
                self.preprocessed_data, 
                group1,
                group2,
                pval_thresh, 
                fc_thresh,
                correction_method
            )
            
            self.volcano_result = volcano_result
            self.create_volcano_plot(volcano_result, pval_thresh, fc_thresh)
            self.vis_manager.create_volcano_plot(volcano_result, pval_thresh, fc_thresh)
            
            self.update_status("✓ Volcano plot complete")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"Volcano plot failed:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error in volcano plot")
    
    def perform_volcano_analysis(self, data, group1, group2, pval_thresh=0.05, fc_thresh=2.0, correction_method='None'):
        """Perform differential analysis for volcano plot"""
        sample_cols = [col for col in data.columns if '.mzML' in col]
        
        # Get sample columns for each group
        group1_cols = []
        group2_cols = []
        
        for base_name, reps in self.replicate_mapping.items():
            group = self.group_mapping.get(base_name, "Ungrouped")
            
            if group == group1:
                target_list = group1_cols
            elif group == group2:
                target_list = group2_cols
            else:
                continue
                
            # Prefer the averaged column to avoid inflating the T-test N (pseudo-replication)
            avg_col = f"{base_name}_avg.mzML Peak area"
            if avg_col in sample_cols:
                target_list.append(avg_col)
            else:
                # Fallback to individual replicates if they weren't averaged
                rep_cols = [c for c, _ in reps if c in sample_cols]
                target_list.extend(rep_cols)
                
        if len(group1_cols) == 0 or len(group2_cols) == 0:
             raise ValueError(f"Could not find valid samples for one or both selected groups.\nGroup 1 ({group1}): {len(group1_cols)} cols\nGroup 2 ({group2}): {len(group2_cols)} cols")
        
        # Calculate fold changes and p-values
        fold_changes = []
        pvalues = []
        
        for idx in range(len(data)):
            # Safely extract and clean numeric arrays (handles missing NaNs gracefully)
            g1_vals = pd.to_numeric(data.iloc[idx][group1_cols], errors='coerce').values
            g2_vals = pd.to_numeric(data.iloc[idx][group2_cols], errors='coerce').values
            
            g1_vals = g1_vals[~np.isnan(g1_vals)]
            g2_vals = g2_vals[~np.isnan(g2_vals)]
            
            mean1 = np.mean(g1_vals) if len(g1_vals) > 0 else 0
            mean2 = np.mean(g2_vals) if len(g2_vals) > 0 else 0
            
            # Fold change (log2)
            fc = np.log2((mean2 + 1) / (mean1 + 1))
            fold_changes.append(fc)
            
            # T-test (requires at least 2 non-NaN values per group)
            if len(g1_vals) >= 2 and len(g2_vals) >= 2:
                # Check for zero variance to avoid Runtime Errors
                if np.var(g1_vals) == 0 and np.var(g2_vals) == 0 and mean1 == mean2:
                    pvalues.append(1.0)
                else:
                    try:
                        _, pval = stats.ttest_ind(g1_vals, g2_vals, equal_var=False)
                        pvalues.append(pval if not np.isnan(pval) else 1.0)
                    except:
                        pvalues.append(1.0)
            else:
                pvalues.append(1.0)
        
        raw_pvalues = np.array(pvalues)
        final_pvalues = raw_pvalues.copy()

        if correction_method == 'Bonferroni':
            final_pvalues = np.minimum(raw_pvalues * len(raw_pvalues), 1.0)
        elif correction_method == 'FDR (Benjamini-Hochberg)':
            n_feat = len(raw_pvalues)
            indices = np.argsort(raw_pvalues)
            sorted_p = raw_pvalues[indices]
            correction_factors = n_feat / np.arange(1, n_feat + 1)
            corrected_p = sorted_p * correction_factors
            for i in range(n_feat - 2, -1, -1):
                corrected_p[i] = min(corrected_p[i], corrected_p[i + 1])
            final_pvalues[indices] = np.minimum(corrected_p, 1.0)

        return {
            'fold_changes': np.array(fold_changes),
            'pvalues': final_pvalues,
            'raw_pvalues': raw_pvalues,
            'group1': group1,
            'group2': group2,
            'feature_ids': data.iloc[:, 0].values,
            'correction_method': correction_method
        }
    
    def create_volcano_plot(self, result, pval_thresh, fc_thresh):
        """Create interactive volcano plot with hover tooltips"""
        for widget in self.volcano_plot_frame.winfo_children():
            widget.destroy()
        
        w, h, dpi = self.plot_width_var.get(), self.plot_height_var.get(), self.plot_dpi_var.get()
        fig = Figure(figsize=(w, h), dpi=dpi)
        self.generated_plots['Univariate Screening'] = fig
        ax = fig.add_subplot(111)
        
        fc = result['fold_changes']
        pvals = np.array(result['pvalues'], dtype=float)
        
        # Guard against p=0.0 causing log10(0) -> infinite crash
        pvals = np.clip(pvals, np.finfo(float).tiny, 1.0)
        log_pvals = -np.log10(pvals)
        
        # Classify points
        significant_up = (fc >= np.log2(fc_thresh)) & (pvals <= pval_thresh)
        significant_down = (fc <= -np.log2(fc_thresh)) & (pvals <= pval_thresh)
        not_significant = ~(significant_up | significant_down)
        
        # Plot
        ax.scatter(fc[not_significant], log_pvals[not_significant], 
                  c='gray', s=20, alpha=0.5, label='Not Significant')
                  
        sc_up = ax.scatter(fc[significant_up], log_pvals[significant_up], 
                  c='red', s=30, alpha=0.7, label=f'Up in {result["group2"]}')
                  
        sc_down = ax.scatter(fc[significant_down], log_pvals[significant_down], 
                  c='blue', s=30, alpha=0.7, label=f'Down in {result["group2"]}')
        
        # Threshold lines
        ax.axhline(-np.log10(pval_thresh), color='black', linestyle='--', linewidth=1, alpha=0.5)
        ax.axvline(np.log2(fc_thresh), color='black', linestyle='--', linewidth=1, alpha=0.5)
        ax.axvline(-np.log2(fc_thresh), color='black', linestyle='--', linewidth=1, alpha=0.5)
        
        ax.set_xlabel(f'log2(Fold Change) [{result["group2"]} / {result["group1"]}]', fontsize=12)
        
        ylabel = '-log10(p-value)'
        if result.get('correction_method', 'None') != 'None':
            ylabel = f'-log10({result["correction_method"]} p-value)'
        ax.set_ylabel(ylabel, fontsize=12)
        
        ax.set_title('Volcano Plot', fontsize=14, fontweight='bold')
        ax.legend(loc='best', framealpha=0.9)
        ax.grid(True, alpha=0.3)
        
        # Interactive Tooltip Logic
        annot = ax.annotate("", xy=(0,0), xytext=(10,10), textcoords="offset points",
                            bbox=dict(boxstyle="round", fc="yellow", alpha=0.8),
                            arrowprops=dict(arrowstyle="->"))
        annot.set_visible(False)
        
        up_idx = np.where(significant_up)[0]
        down_idx = np.where(significant_down)[0]
        
        def on_hover(event):
            if event.inaxes == ax:
                cont_up, ind_up = sc_up.contains(event)
                cont_down, ind_down = sc_down.contains(event)
                
                idx = None
                if cont_up:
                    idx = up_idx[ind_up["ind"][0]]
                    annot.xy = sc_up.get_offsets()[ind_up["ind"][0]]
                elif cont_down:
                    idx = down_idx[ind_down["ind"][0]]
                    annot.xy = sc_down.get_offsets()[ind_down["ind"][0]]
                
                if idx is not None:
                    feat_id = result['feature_ids'][idx]
                    raw_p = result.get('raw_pvalues', result['pvalues'])[idx]
                    text = f"{feat_id}\nRaw P: {raw_p:.2e}"
                    if result.get('correction_method', 'None') != 'None':
                        text += f"\nAdj P: {result['pvalues'][idx]:.2e}"
                    
                    annot.set_text(text)
                    annot.set_visible(True)
                    canvas.draw_idle()
                else:
                    if annot.get_visible():
                        annot.set_visible(False)
                        canvas.draw_idle()
        
        canvas = FigureCanvasTkAgg(fig, master=self.volcano_plot_frame)
        canvas.draw()
        canvas.mpl_connect("motion_notify_event", on_hover)
        canvas.get_tk_widget().pack(fill='both', expand=True)
        
        toolbar = NavigationToolbar2Tk(canvas, self.volcano_plot_frame)
        toolbar.update()
        
    def apply_screening_filter(self):
        """Applies the current Volcano Plot thresholds to filter the dataset for downstream PLS-DA and RF."""
        if getattr(self, 'volcano_result', None) is None or self.preprocessed_data is None:
            messagebox.showwarning("Warning", "Please run 'Create Volcano Plot' first!")
            return
            
        try:
            pval_thresh = self.pvalue_var.get()
            fc_thresh = self.fc_var.get()
            
            fc = self.volcano_result['fold_changes']
            pvals = self.volcano_result['pvalues']
            
            significant_up = (fc >= np.log2(fc_thresh)) & (pvals <= pval_thresh)
            significant_down = (fc <= -np.log2(fc_thresh)) & (pvals <= pval_thresh)
            sig_mask = significant_up | significant_down
            
            sig_features = self.volcano_result['feature_ids'][sig_mask]
            
            if len(sig_features) == 0:
                messagebox.showwarning("Warning", "No significant features found with the current thresholds. Filter not applied.")
                return
                
            # Subset the preprocessed data directly
            self.screened_data = self.preprocessed_data[self.preprocessed_data.iloc[:, 0].isin(sig_features)].reset_index(drop=True)
            
            # Update feature viewer with filtered list
            self.update_feature_viewer_options()

            self.update_status(f"✓ Filter applied: {len(sig_features)} features selected for downstream analysis.")
            messagebox.showinfo("Success", f"Filtering complete!\n\n{len(sig_features)} significant features retained for downstream analysis.")
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"Filtering failed:\n{str(e)}\n\nDetails:\n{error_detail}")

    # ==================== PLS-DA Analysis & Validation ====================

    def run_plsda(self):
        """Run PLS-DA analysis with VIP calculation"""
        if getattr(self, 'screened_data', None) is None:
            messagebox.showwarning("Warning", "Please run preprocessing (and optional Univariate Screening) first!")
            return
        
        try:
            self.update_status("Running PLS-DA...")
            
            n_components = self.plsda_components_var.get()
            exclude_qc = self.plsda_exclude_qc_var.get()
            
            # Check pairwise settings
            specific_groups = None
            if self.plsda_pairwise_var.get():
                g1 = self.plsda_group1_var.get()
                g2 = self.plsda_group2_var.get()
                if not g1 or not g2:
                    messagebox.showwarning("Warning", "Please select both groups for pairwise comparison.")
                    return
                if g1 == g2:
                    messagebox.showwarning("Warning", "Groups must be different.")
                    return
                specific_groups = [g1, g2]

            plsda_result = self.perform_plsda_local(self.screened_data, n_components, exclude_qc, specific_groups)
            
            self.plsda_result = plsda_result
            self.display_plsda_results(plsda_result)
            self.create_plsda_plot(plsda_result)
            self.vis_manager.create_plsda_plot(plsda_result)
            
            self.update_status("✓ PLS-DA complete")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"PLS-DA failed:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error in PLS-DA")
    
    def perform_plsda_local(self, data, n_components=2, exclude_qc=True, specific_groups=None):
        """Built-in PLS-DA implementation with VIP extraction and R2/Q2 metrics"""

        sample_cols = [col for col in data.columns if '.mzML' in col]
        
        if exclude_qc or specific_groups:
            filtered_cols = []
            for col in sample_cols:
                base_name = re.sub(r'_(\d+|avg)\.mzML.*', '', col)
                group_name = self.group_mapping.get(base_name, "Ungrouped")
                
                if specific_groups and group_name not in specific_groups:
                    continue
                
                # Exclude if explicitly in qc_samples OR if its group name suggests it's QC
                if exclude_qc and (base_name in getattr(self, 'qc_samples', []) or 'qc' in group_name.lower()):
                    continue
                    
                filtered_cols.append(col)
            sample_cols = filtered_cols
            
        if len(sample_cols) < 2:
            raise ValueError("Not enough samples for PLS-DA after excluding QCs. Please check your data.")

        X = data[sample_cols].T
        
        # Robust imputation
        X = X.replace(0, np.nan)
        min_val = X.min().min() * 0.2
        if pd.isna(min_val): min_val = 0.1
        X = X.fillna(min_val)
        
        # Optional Log Transform
        if hasattr(self, 'plsda_log_var') and self.plsda_log_var.get():
            X = np.log10(X + 1)
        
        # Prepare labels
        y_labels = []
        for col in sample_cols:
            base_name = re.sub(r'_(\d+|avg)\.mzML.*', '', col)
            group = self.group_mapping.get(base_name, "Ungrouped")
            y_labels.append(group)
            
        unique_groups = set(y_labels)
        if len(unique_groups) < 2:
            raise ValueError("PLS-DA requires at least 2 distinct groups. After excluding QCs, less than 2 groups remain.")
        
        # Encode labels for Random Forest CV (requires 1D integers)
        le = LabelEncoder()
        y_cls = le.fit_transform(y_labels)
        
        # Encode labels for PLS-DA (requires One-Hot matrix for multi-class)
        y_labels_2d = np.array(y_labels).reshape(-1, 1)
        try:
            encoder = OneHotEncoder(sparse_output=False)
            Y_mat = encoder.fit_transform(y_labels_2d)
        except TypeError:
            encoder = OneHotEncoder(sparse=False)
            Y_mat = encoder.fit_transform(y_labels_2d)

        # If strictly 2 groups, a 1D vector is mathematically identical
        if len(unique_groups) == 2:
            Y_pls = Y_mat[:, 1]
        else:
            Y_pls = Y_mat
        
        # Scale features
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        
        # Fit PLS-DA
        plsda = PLSRegression(n_components=n_components)
        scores = plsda.fit_transform(X_scaled, Y_pls)[0]

        # Calculate R2X
        X_hat = np.dot(scores, plsda.x_loadings_.T)
        ss_res_x = np.sum((X_scaled - X_hat)**2)
        ss_tot_x = np.sum(X_scaled**2)
        r2_x = 1 - (ss_res_x / ss_tot_x) if ss_tot_x > 0 else 0

        # Calculate R2Y and Q2 scores
        def calculate_block_variance(y_true, y_pred):
            ss_res = np.sum((y_true - y_pred) ** 2)
            mean_y = np.mean(y_true, axis=0)
            ss_tot = np.sum((y_true - mean_y) ** 2)
            if ss_tot == 0: return 0.0
            return 1.0 - (ss_res / ss_tot)

        Y_pred_actual = plsda.predict(X_scaled)
        if Y_pred_actual.ndim > 1 and Y_pls.ndim == 1:
            Y_pred_actual = Y_pred_actual.ravel()
        r2_y = calculate_block_variance(Y_pls, Y_pred_actual)
        
        # Calculate VIP scores
        vip_scores = self.calculate_vip_scores(plsda, X_scaled, Y_pls)
        
        # Cross-validation setup
        if len(y_cls) >= 10 and min(np.bincount(y_cls)) >= 2:
            from sklearn.model_selection import StratifiedKFold
            skf = StratifiedKFold(n_splits=min(5, min(np.bincount(y_cls))))
            cv_method = list(skf.split(X_scaled, y_cls))
        else:
            from sklearn.model_selection import LeaveOneOut
            cv_method = LeaveOneOut()

        Y_cv_actual = cross_val_predict(plsda, X_scaled, Y_pls, cv=cv_method)
        q2_y = calculate_block_variance(Y_pls, Y_cv_actual)

        cv_scores = cross_val_score(
            RandomForestClassifier(n_estimators=100, random_state=42),
            X_scaled, y_cls, cv=cv_method
        )
        
        return {
            'scores': scores,
            'vip_scores': vip_scores,
            'cv_accuracy': cv_scores.mean(),
            'cv_std': cv_scores.std(),
            'r2_x': r2_x,
            'r2_y': r2_y,
            'q2_y': q2_y,
            'sample_names': sample_cols,
            'groups': y_labels,
            'model': plsda
        }
    
    def calculate_vip_scores(self, model, X, y):
        """Calculate Variable Importance in Projection (VIP) scores"""
        import numpy as np
        t = model.x_scores_
        w = model.x_weights_
        q = model.y_loadings_
        
        p, h = w.shape
        vips = np.zeros((p,))
        
        # Calculate explained variance of Y for each component
        # Works for both single-class (1D) and multi-class (2D) Y
        s = np.sum(t ** 2, axis=0) * np.sum(q ** 2, axis=1)
        total_s = np.sum(s)
        
        w_norm_sq = np.sum(w ** 2, axis=0)
        
        for i in range(p):
            weight = (w[i, :] ** 2) / w_norm_sq
            vips[i] = np.sqrt(p * np.sum(s * weight) / total_s)
        
        return vips
    
    def display_plsda_results(self, result):
        """Display PLS-DA results"""
        import tkinter as tk
        self.plsda_results_text.delete('1.0', tk.END)
        
        text = []
        text.append("=" * 60)
        text.append("PLS-DA RESULTS")
        text.append("=" * 60)
        text.append(f"\nModel Performance (n_components={result['model'].n_components}):")
        text.append(f"R²X (Explained Var. X):   {result.get('r2_x', 0):.3f}")
        text.append(f"R²Y (Goodness of Fit):    {result.get('r2_y', 0):.3f}")
        text.append(f"Q²  (Predictability):     {result.get('q2_y', 0):.3f}")
        text.append(f"RF CV Accuracy:           {result['cv_accuracy']:.3f} ± {result['cv_std']:.3f}")
        text.append(f"\nTop 10 VIP Scores (VIP > 1.0 indicates important features):")
        text.append("-" * 60)
        
        # Get feature names (assuming first column is ID)
        feature_ids = self.screened_data.iloc[:, 0].values
        vip_with_ids = list(zip(feature_ids, result['vip_scores']))
        vip_sorted = sorted(vip_with_ids, key=lambda x: x[1], reverse=True)
        
        for i, (feat_id, vip) in enumerate(vip_sorted[:10], 1):
            text.append(f"{i:2d}. {feat_id}: {vip:.3f}")
        
        self.plsda_results_text.insert('1.0', '\n'.join(text))
    
    def create_plsda_plot(self, result):
        """Create PLS-DA score plot with 95% Confidence Ellipses AND VIP bar chart"""
        import matplotlib.transforms as transforms
        
        for widget in getattr(self, 'plsda_plot_frame', tk.Frame()).winfo_children():
            widget.destroy()
        
        w, h, dpi = self.plot_width_var.get(), self.plot_height_var.get(), self.plot_dpi_var.get()
        fig = Figure(figsize=(w * 2, h), dpi=dpi)
        self.generated_plots['PLS-DA'] = fig
        
        # === Left Subplot: PLS-DA Score Plot ===
        ax1 = fig.add_subplot(121)
        groups = result['groups']
        unique_groups = sorted(set(groups))
        
        def add_confidence_ellipse(x, y, ax, color):
            """Calculates and draws a 95% confidence ellipse"""
            if len(x) < 3:
                return  # Need at least 3 points for a reliable covariance matrix
            cov = np.cov(x, y)
            if cov[0, 0] == 0 or cov[1, 1] == 0:
                return
            
            pearson = cov[0, 1] / np.sqrt(cov[0, 0] * cov[1, 1])
            ell_radius_x = np.sqrt(1 + pearson)
            ell_radius_y = np.sqrt(1 - pearson)
            
            ellipse = Ellipse((0, 0), width=ell_radius_x * 2, height=ell_radius_y * 2,
                              facecolor=color, alpha=0.15, edgecolor=color, linewidth=1.5)
            
            # 2.4477 is the square root of 5.991 (the 95% confidence chi-square value for 2 DOF)
            scale_x = np.sqrt(cov[0, 0]) * 2.4477
            scale_y = np.sqrt(cov[1, 1]) * 2.4477
            
            transf = transforms.Affine2D() \
                .rotate_deg(45) \
                .scale(scale_x, scale_y) \
                .translate(np.mean(x), np.mean(y))
                
            ellipse.set_transform(transf + ax.transData)
            ax.add_patch(ellipse)
        
        for group in unique_groups:
            group_indices = [i for i, g in enumerate(groups) if g == group]
            group_scores_lv1 = result['scores'][group_indices, 0]
            group_scores_lv2 = result['scores'][group_indices, 1]
            
            color = getattr(self, 'group_colors', {}).get(group, '#4ECDC4')
            
            ax1.scatter(
                group_scores_lv1,
                group_scores_lv2,
                s=100,
                alpha=0.7,
                c=color,
                edgecolors='black',
                linewidth=1.5,
                label=group
            )
            # Add the 95% confidence ellipse for this group
            add_confidence_ellipse(group_scores_lv1, group_scores_lv2, ax1, color)
        
        r2_val = result.get('r2_y', 0)
        q2_val = result.get('q2_y', 0)
        
        ax1.set_xlabel('LV1', fontsize=12)
        ax1.set_ylabel('LV2', fontsize=12)
        ax1.set_title(f"PLS-DA Score Plot (95% Confidence Ellipses)\n"
                      f"R²Y: {r2_val:.3f} | Q²: {q2_val:.3f}", 
                      fontsize=12, fontweight='bold')
        ax1.legend(loc='best', framealpha=0.9)
        ax1.grid(True, alpha=0.3)
        
        # === Right Subplot: VIP Scores ===
        ax2 = fig.add_subplot(122)
        vips = result['vip_scores']
        feature_ids = self.screened_data.iloc[:, 0].values
        
        # Plot top 15 VIPs
        top_idx = np.argsort(vips)[-15:]
        ax2.barh(range(15), vips[top_idx], color='purple', alpha=0.7)
        ax2.set_yticks(range(15))
        ax2.set_yticklabels([feature_ids[i] for i in top_idx], fontsize=8)
        ax2.axvline(1.0, color='red', linestyle='--', label='VIP > 1.0 cutoff')
        ax2.set_xlabel('VIP Score', fontsize=12)
        ax2.set_title("Top 15 Variable Importance in Projection (VIP)", fontsize=12, fontweight='bold')
        ax2.legend(loc='lower right')
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, master=self.plsda_plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)
        
        toolbar = NavigationToolbar2Tk(canvas, self.plsda_plot_frame)
        toolbar.update()
        
    def run_pls_val(self):
        """Run Permutation Testing for PLS-DA Validation with Multi-Dimensional Y-Block"""
        
        if getattr(self, 'screened_data', None) is None:
            messagebox.showwarning("Warning", "Please run preprocessing (and optional Univariate Screening) first!")
            return
            
        try:
            self.update_status("Running PLS Validation Permutation Test (this may take a moment)...")
            # Force GUI update safely
            if hasattr(self, 'update_idletasks'):
                self.update_idletasks()
            elif hasattr(self, 'pls_val_plot_frame'):
                self.pls_val_plot_frame.update_idletasks()
            
            n_perms = self.n_perms_var.get()
            n_components = self.plsda_components_var.get()
            exclude_qc = getattr(self, 'pls_val_exclude_qc_var', tk.BooleanVar(value=True)).get()
            
            # --- 1. Data Preparation ---
            # Reuse samples from the generated PLS-DA model to ensure consistency (handles pairwise selection)
            if 'sample_names' in self.plsda_result:
                sample_cols = self.plsda_result['sample_names']
            else:
                sample_cols = [col for col in self.screened_data.columns if '.mzML' in col]
                # Note: This fallback path ignores exclude_qc logic, assuming sample_names is always present from PLS-DA

            if len(sample_cols) < 2:
                messagebox.showwarning("Warning", "Not enough samples for Validation after excluding QCs.")
                return
            
            X = self.screened_data[sample_cols].T
            X = X.replace(0, np.nan)
            min_val = X.min().min() * 0.2
            if pd.isna(min_val): min_val = 0.1
            X = X.fillna(min_val)
            
            if hasattr(self, 'plsda_log_var') and self.plsda_log_var.get():
                X = np.log10(X + 1)
            
            y_labels = []
            for col in sample_cols:
                base_name = re.sub(r'_(\d+|avg)\.mzML.*', '', col)
                group = self.group_mapping.get(base_name, "Ungrouped")
                y_labels.append(group)
            
            unique_groups = set(y_labels)
            if len(unique_groups) < 2:
                messagebox.showwarning("Warning", "Permutation Test requires at least 2 groups after excluding QCs.")
                return
                
            # --- 2. Proper Y-Block Encoding via sklearn ---
            y_labels_2d = np.array(y_labels).reshape(-1, 1)
            
            try:
                encoder = OneHotEncoder(sparse_output=False)
                y_dummies = encoder.fit_transform(y_labels_2d)
            except TypeError:
                encoder = OneHotEncoder(sparse=False)
                y_dummies = encoder.fit_transform(y_labels_2d)
            
            if len(unique_groups) == 2:
                Y_mat = y_dummies[:, 1] 
            else:
                Y_mat = y_dummies 
                
            # Keep a 1D encoded version strictly for calculating the X-axis correlation in the plot
            le = LabelEncoder()
            y_1d_corr = le.fit_transform(y_labels)
            
            scaler = StandardScaler()
            X_scaled = scaler.fit_transform(X)
            
            # Optimize CV approach based on sample size
            if len(y_labels) >= 15 and min(np.bincount(y_1d_corr)) >= 3:
                from sklearn.model_selection import StratifiedKFold
                skf = StratifiedKFold(n_splits=min(7, min(np.bincount(y_1d_corr))))
                cv = list(skf.split(X_scaled, y_1d_corr))
            else:
                from sklearn.model_selection import LeaveOneOut
                cv = LeaveOneOut()

            # --- 3. Block Metric Helper Function ---
            def calculate_block_variance(y_true, y_pred):
                """Calculates R2/Q2 for both 1D vectors and 2D dummy matrices"""
                ss_res = np.sum((y_true - y_pred) ** 2)
                mean_y = np.mean(y_true, axis=0)
                ss_tot = np.sum((y_true - mean_y) ** 2)
                if ss_tot == 0: return 0.0
                return 1.0 - (ss_res / ss_tot)

            # --- 4. Fit Actual Model ---
            pls_val = PLSRegression(n_components=n_components)
            pls_val.fit(X_scaled, Y_mat)
            
            # Get actual R2
            Y_pred_actual = pls_val.predict(X_scaled)
            if Y_pred_actual.ndim > 1 and Y_mat.ndim == 1:
                Y_pred_actual = Y_pred_actual.ravel()
            r2_actual = calculate_block_variance(Y_mat, Y_pred_actual)
            
            # Get actual Q2 via Cross-Validation
            Y_cv_actual = cross_val_predict(pls_val, X_scaled, Y_mat, cv=cv)
            q2_actual = calculate_block_variance(Y_mat, Y_cv_actual)
            
            r2_perms = []
            q2_perms = []
            corrs = []
            
            # --- 5. Permutation Loop ---
            for _ in range(n_perms):
                shuffle_idx = np.random.permutation(len(Y_mat))
                Y_mat_perm = Y_mat[shuffle_idx]
                y_1d_perm = y_1d_corr[shuffle_idx]
                
                corr = np.corrcoef(y_1d_corr, y_1d_perm)[0, 1]
                if np.isnan(corr): corr = 0
                corrs.append(corr)
                
                pls_val.fit(X_scaled, Y_mat_perm)
                Y_pred_perm = pls_val.predict(X_scaled)
                if Y_pred_perm.ndim > 1 and Y_mat_perm.ndim == 1:
                    Y_pred_perm = Y_pred_perm.ravel()
                    
                r2_perms.append(calculate_block_variance(Y_mat_perm, Y_pred_perm))
                
                Y_cv_perm = cross_val_predict(pls_val, X_scaled, Y_mat_perm, cv=cv)
                q2_perms.append(calculate_block_variance(Y_mat_perm, Y_cv_perm))
                
            # --- 6. Render Validation Plot ---
            for widget in getattr(self, 'pls_val_plot_frame', tk.Frame()).winfo_children():
                widget.destroy()
                
            w, h, dpi = self.plot_width_var.get(), self.plot_height_var.get(), self.plot_dpi_var.get()
            fig = Figure(figsize=(w, h), dpi=dpi)
            self.generated_plots['PLS-DA Validation'] = fig
            ax = fig.add_subplot(111)
            
            ax.scatter(corrs, r2_perms, color='#1f77b4', alpha=0.6, label='R² (Permuted)')
            ax.scatter(corrs, q2_perms, color='#2ca02c', alpha=0.6, label='Q² (Permuted)')
            
            ax.scatter([1.0], [r2_actual], color='#1f77b4', marker='*', s=150, label='R² (Actual)')
            ax.scatter([1.0], [q2_actual], color='#2ca02c', marker='*', s=150, label='Q² (Actual)')
            
            all_corrs = corrs + [1.0]
            all_r2 = r2_perms + [r2_actual]
            all_q2 = q2_perms + [q2_actual]

            if len(corrs) > 1:
                z_r2 = np.polyfit(all_corrs, all_r2, 1)
                p_r2 = np.poly1d(z_r2)
                z_q2 = np.polyfit(all_corrs, all_q2, 1)
                p_q2 = np.poly1d(z_q2)
                
                x_line = np.linspace(min(corrs + [0]), 1.0, 50)
                ax.plot(x_line, p_r2(x_line), color='#1f77b4', linestyle='--')
                ax.plot(x_line, p_q2(x_line), color='#2ca02c', linestyle='--')
                
                r2_int = p_r2(0)
                q2_int = p_q2(0)
            else:
                r2_int, q2_int = 0, 0
                
            p_val = (np.sum(np.array(q2_perms) >= q2_actual) + 1) / (n_perms + 1)

            perm_results = {
                'corrs': corrs, 'r2_perms': r2_perms, 'q2_perms': q2_perms,
                'r2_actual': r2_actual, 'q2_actual': q2_actual,
                'r2_int': r2_int, 'q2_int': q2_int,
                'p_val': p_val, 'n_perms': n_perms
            }
            
            title = (f"PLS-DA Permutation Test (n={n_perms})\n"
                     f"R²Y = {r2_actual:.3f}, Q² = {q2_actual:.3f}\n"
                     f"R² int = {r2_int:.3f}, Q² int = {q2_int:.3f} | p-value = {p_val:.3f}")
            ax.set_title(title, fontsize=12, fontweight='bold')
            ax.set_xlabel("Correlation between permuted and original Y", fontsize=11)
            ax.set_ylabel("Score (R² / Q²)", fontsize=11)
            ax.legend(loc='best')
            ax.grid(True, linestyle=':', alpha=0.6)
            fig.tight_layout()
            self.vis_manager.create_pls_val_plot(perm_results)
            
            canvas = FigureCanvasTkAgg(fig, master=self.pls_val_plot_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)
            toolbar = NavigationToolbar2Tk(canvas, self.pls_val_plot_frame)
            toolbar.update()
            
            self.update_status("✓ PLS Validation complete")
            messagebox.showinfo("Success", f"Validation complete!\n\nR²Y: {r2_actual:.3f}\nQ²: {q2_actual:.3f}\nEmpirical p-value: {p_val:.3f}")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"Validation failed:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error in PLS Validation")
    
    # ==================== Random Forest ====================
    
    def run_rf(self):
        """Run Random Forest analysis"""
        if getattr(self, 'screened_data', None) is None:
            messagebox.showwarning("Warning", "Please run preprocessing (and optional Univariate Screening) first!")
            return
        
        try:
            self.update_status("Running Random Forest...")
            
            n_trees = self.ntrees_var.get()
            top_n = self.top_n_var.get()
            exclude_qc = getattr(self, 'rf_exclude_qc_var', tk.BooleanVar(value=True)).get()
            
            rf_result = self.perform_rf_analysis(self.screened_data, n_trees, top_n, exclude_qc)
            
            self.rf_result = rf_result
            self.display_rf_results(rf_result)
            self.create_rf_plot(rf_result)
            
            self.update_status("✓ Random Forest complete")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"Random Forest failed:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error in Random Forest")
    
    def perform_rf_analysis(self, data, n_trees=500, top_n=20, exclude_qc=True, specific_groups=None, calc_roc=False):
        """Perform Random Forest feature selection"""
        sample_cols = [col for col in data.columns if '.mzML' in col]
        
        if exclude_qc or specific_groups:
            filtered_cols = []
            for col in sample_cols:
                base_name = re.sub(r'_(\d+|avg)\.mzML.*', '', col)
                group_name = self.group_mapping.get(base_name, "Ungrouped")
                
                if specific_groups and group_name not in specific_groups:
                    continue
                
                # Exclude if explicitly in qc_samples OR if its group name suggests it's QC
                if base_name not in getattr(self, 'qc_samples', []) and 'qc' not in group_name.lower():
                    filtered_cols.append(col)
            sample_cols = filtered_cols
            
        if len(sample_cols) < 2:
            raise ValueError("Not enough samples for Random Forest after excluding QCs. Please check your data.")
            
        X = data[sample_cols].T
        
        X = X.replace(0, np.nan)
        min_val = X.min().min() * 0.2
        if pd.isna(min_val): min_val = 0.1
        X = X.fillna(min_val)
        
        # Prepare labels
        y_labels = []
        for col in sample_cols:
            base_name = re.sub(r'_(\d+|avg)\.mzML.*', '', col)
            group = self.group_mapping.get(base_name, "Ungrouped")
            y_labels.append(group)
            
        unique_groups = set(y_labels)
        if len(unique_groups) < 2:
            raise ValueError("Random Forest requires at least 2 distinct groups. After excluding QCs, less than 2 groups remain.")
        
        # Encode labels
        le = LabelEncoder()
        y = le.fit_transform(y_labels)
        
        # Scale features
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        
        # Fit Random Forest (using oob_score for evaluation)
        try:
            rf = RandomForestClassifier(n_estimators=n_trees, oob_score=True, random_state=42)
            rf.fit(X_scaled, y)
            oob_error = 1.0 - getattr(rf, 'oob_score_', float('nan'))
        except Exception:
            # Fallback if too few samples for OOB scoring
            rf = RandomForestClassifier(n_estimators=n_trees, random_state=42)
            rf.fit(X_scaled, y)
            oob_error = float('nan')
        
        # Feature importance
        importances = rf.feature_importances_
        feature_ids = data.iloc[:, 0].values
        
        # Get top features
        importance_with_ids = list(zip(feature_ids, importances))
        importance_sorted = sorted(importance_with_ids, key=lambda x: x[1], reverse=True)
        
        top_features = importance_sorted[:top_n]
        
        # Cross-validation setup
        cv_method = min(5, min(np.bincount(y)))
        if cv_method < 2:
            from sklearn.model_selection import LeaveOneOut
            cv_folds = LeaveOneOut()
        else:
            from sklearn.model_selection import StratifiedKFold
            cv_folds = StratifiedKFold(n_splits=cv_method)
            
        cv_scores = cross_val_score(rf, X_scaled, y, cv=cv_folds)
        
        roc_data = None
        if calc_roc and len(unique_groups) == 2:
            from sklearn.metrics import roc_curve, auc
            y_prob = cross_val_predict(rf, X_scaled, y, cv=cv_folds, method="predict_proba")[:, 1]
            fpr, tpr, _ = roc_curve(y, y_prob)
            roc_data = {
                'fpr': fpr,
                'tpr': tpr,
                'auc': auc(fpr, tpr),
                'group1': list(le.inverse_transform([0]))[0],
                'group2': list(le.inverse_transform([1]))[0]
            }
        
        return {
            'importances': importances,
            'feature_ids': feature_ids,
            'top_features': top_features,
            'cv_accuracy': cv_scores.mean(),
            'cv_std': cv_scores.std(),
            'oob_error': oob_error,
            'model': rf,
            'roc_data': roc_data
        }

    # ==================== Random Forest ====================
    
    def run_rf(self):
        """Run Random Forest analysis"""
        if getattr(self, 'screened_data', None) is None:
            messagebox.showwarning("Warning", "Please run preprocessing (and optional Univariate Screening) first!")
            return
        
        try:
            self.update_status("Running Random Forest...")
            
            n_trees = self.ntrees_var.get()
            top_n = self.top_n_var.get()
            exclude_qc = getattr(self, 'rf_exclude_qc_var', tk.BooleanVar(value=True)).get()
            calc_roc = getattr(self, 'rf_roc_var', tk.BooleanVar(value=True)).get()
            
            specific_groups = None
            if getattr(self, 'rf_pairwise_var', tk.BooleanVar(value=False)).get():
                g1 = getattr(self, 'rf_group1_var', tk.StringVar()).get()
                g2 = getattr(self, 'rf_group2_var', tk.StringVar()).get()
                if not g1 or not g2:
                    messagebox.showwarning("Warning", "Please select both groups for pairwise comparison.")
                    return
                if g1 == g2:
                    messagebox.showwarning("Warning", "Groups must be different.")
                    return
                specific_groups = [g1, g2]
            
            rf_result = self.perform_rf_analysis(self.screened_data, n_trees, top_n, exclude_qc, specific_groups, calc_roc)
            
            self.rf_result = rf_result
            self.display_rf_results(rf_result)
            self.vis_manager.create_rf_plot(rf_result)
            
            self.update_status("✓ Random Forest complete")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"Random Forest failed:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error in Random Forest")
    
    def display_rf_results(self, result):
        """Display Random Forest results"""
        self.rf_results_text.delete('1.0', tk.END)
        
        text = []
        text.append("=" * 60)
        text.append("RANDOM FOREST RESULTS")
        text.append("=" * 60)
        text.append(f"\nCross-validation Accuracy: {result['cv_accuracy']:.3f} ± {result['cv_std']:.3f}")
        if not np.isnan(result.get('oob_error', float('nan'))):
            text.append(f"Out-of-Bag (OOB) Error Rate: {result['oob_error']:.3f}")
        else:
            text.append("Out-of-Bag (OOB) Error Rate: N/A (Sample size too small)")
        text.append(f"\nTop {len(result['top_features'])} Important Features:")
        text.append("-" * 60)
        
        for i, (feat_id, importance) in enumerate(result['top_features'], 1):
            text.append(f"{i:2d}. {feat_id}: {importance:.6f}")
        
        self.rf_results_text.insert('1.0', '\n'.join(text))
    
    def create_rf_plot(self, result):
        """Create Random Forest feature importance plot"""
        for widget in self.rf_plot_frame.winfo_children():
            widget.destroy()
        
        w, h, dpi = self.plot_width_var.get(), self.plot_height_var.get(), self.plot_dpi_var.get()
        
        has_curve = result.get('cv_curve') is not None
        has_roc = result.get('roc_data') is not None
        num_subplots = 1 + int(has_curve) + int(has_roc)
        
        fig = Figure(figsize=(w * num_subplots, h), dpi=dpi)
        self.generated_plots['Random Forest'] = fig
        ax = fig.add_subplot(1, num_subplots, 1)
        
        top_features = result['top_features']
        feature_names = [f[0] for f in top_features]
        importances = [f[1] for f in top_features]
        
        y_pos = np.arange(len(feature_names))
        
        ax.barh(y_pos, importances, align='center', color='steelblue', alpha=0.8)
        ax.set_yticks(y_pos)
        ax.set_yticklabels(feature_names, fontsize=8)
        ax.invert_yaxis()
        ax.set_xlabel('Feature Importance', fontsize=12)
        ax.set_title(f'Random Forest - Top {len(feature_names)} Features', 
                    fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3, axis='x')
        
        plot_idx = 2
        
        if has_curve:
            ax_curve = fig.add_subplot(1, num_subplots, plot_idx)
            curve = result['cv_curve']
            features = curve['features']
            scores = np.array(curve['scores'])
            stds = np.array(curve['stds'])
            
            ax_curve.plot(features, scores, marker='o', linestyle='-', color='darkorange', linewidth=2)
            ax_curve.fill_between(features, scores - stds, scores + stds, color='darkorange', alpha=0.2)
            ax_curve.set_xlabel('Number of Features', fontsize=12)
            ax_curve.set_ylabel('Cross-Validation Accuracy', fontsize=12)
            ax_curve.set_title('Accuracy vs Number of Features', fontsize=14, fontweight='bold')
            ax_curve.grid(True, alpha=0.3)
            plot_idx += 1
            
        if has_roc:
            ax_roc = fig.add_subplot(1, num_subplots, plot_idx)
            roc = result['roc_data']
            ax_roc.plot(roc['fpr'], roc['tpr'], color='darkorange', lw=2, label=f"AUC = {roc['auc']:.3f}")
            ax_roc.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
            ax_roc.set_xlim([0.0, 1.0])
            ax_roc.set_ylim([0.0, 1.05])
            ax_roc.set_xlabel('False Positive Rate', fontsize=12)
            ax_roc.set_ylabel('True Positive Rate', fontsize=12)
            ax_roc.set_title(f"ROC Curve ({roc['group2']} vs {roc['group1']})", fontsize=14, fontweight='bold')
            ax_roc.legend(loc="lower right", fontsize=10)
            ax_roc.grid(True, alpha=0.3)
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, master=self.rf_plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)
        
        toolbar = NavigationToolbar2Tk(canvas, self.rf_plot_frame)
        toolbar.update()
    
    # ==================== Heatmap Visualization ====================
    
    def run_heatmap(self):
        """Run Heatmap generation for Top Biomarkers"""
        if getattr(self, 'screened_data', None) is None:
            messagebox.showwarning("Warning", "Please run preprocessing (and optional Univariate Screening) first!")
            return
            
        if self.plsda_result is None or self.rf_result is None:
            messagebox.showwarning("Warning", "Please run PLS-DA and Random Forest first to extract the top features!")
            return
            
        try:
            self.update_status("Generating Top Features Heatmap...")
            
            top_n = self.heatmap_top_n_var.get()
            sort_method = getattr(self, 'heatmap_sort_var', tk.StringVar(value="Retention Time")).get()
            cmap = getattr(self, 'heatmap_cmap_var', tk.StringVar(value="coolwarm")).get()

            # Get top features from PLS-DA and RF
            vips = self.plsda_result['vip_scores']
            feature_ids = self.screened_data.iloc[:, 0].values
            top_vip_idx = np.argsort(vips)[-top_n:]
            top_vip_features = [feature_ids[i] for i in top_vip_idx]
            
            top_rf_features = [f[0] for f in self.rf_result['top_features'][:top_n]]
            
            # Union of top discriminative features
            combined_features = list(set(top_vip_features) | set(top_rf_features))
            
            def extract_rt(f_id):
                try:
                    # Expected format: ID_m/z_RT
                    parts = str(f_id).split('_')
                    if len(parts) >= 3:
                        return float(parts[-1])
                    return 0.0
                except:
                    return 0.0
            
            # Initial sort by Retention Time
            if sort_method == "Retention Time":
                combined_features.sort(key=extract_rt)
            
            if not combined_features:
                messagebox.showwarning("Warning", "No significant features found to plot.")
                return
                
            sample_cols = [col for col in self.screened_data.columns if '.mzML' in col]
            
            # Filter rows (features) based on combined top biomarkers
            heatmap_data = self.screened_data[self.screened_data.iloc[:, 0].isin(combined_features)]
            
            # Ensure the data matches the currently established list order
            heatmap_data = heatmap_data.set_index(heatmap_data.columns[0]).loc[combined_features].reset_index()
            self.heatmap_data = heatmap_data
            feature_labels = heatmap_data.iloc[:, 0].values
            
            bio_intensities = []
            bio_info = []
            qc_intensities = []
            qc_info = []
            
            # Locate the exact columns to plot: 
            # We want exactly ONE column per sample base name.
            for base_name, reps in self.replicate_mapping.items():
                group = self.group_mapping.get(base_name, "Ungrouped")
                is_qc = (base_name in self.qc_samples) or ('qc' in group.lower())
                
                # Check if an explicitly averaged column exists
                avg_col = f"{base_name}_avg.mzML Peak area"
                if avg_col in sample_cols:
                    vals = heatmap_data[avg_col].values
                else:
                    # Calculate mean on the fly if _avg missing
                    rep_cols = [col for col, _ in reps if col in sample_cols]
                    if rep_cols:
                        vals = heatmap_data[rep_cols].mean(axis=1).values
                    else:
                        continue
                
                if is_qc:
                    qc_intensities.append(vals)
                    qc_info.append({'name': base_name, 'group': group})
                else:
                    bio_intensities.append(vals)
                    bio_info.append({'name': base_name, 'group': group})
            
            if not bio_info and not qc_info:
                messagebox.showwarning("Warning", "No samples found to plot.")
                return

            # Combine all intensities for a unified Z-score scale
            all_intensities = bio_intensities + qc_intensities
            intensities = np.array(all_intensities).T
            
            # Standardize for visualization (Z-score by row)
            scaled_intensities = StandardScaler().fit_transform(intensities.T).T
            
            # Split back into biological and QC sets
            n_bio = len(bio_info)
            n_qc = len(qc_info)
            scaled_bio = scaled_intensities[:, :n_bio]
            scaled_qc = scaled_intensities[:, n_bio:]
            
            # Perform Hierarchical Clustering on Features (Rows) if selected
            row_linkage = None
            if sort_method == "Clustering" and len(combined_features) > 1:
                # Cluster using the biological samples to find co-regulated patterns
                data_for_row_clustering = scaled_bio if n_bio > 0 else scaled_qc
                row_linkage = sch.linkage(data_for_row_clustering, method='ward')
                row_dendro = sch.dendrogram(row_linkage, no_plot=True)
                row_idx = row_dendro['leaves']
                
                # Re-order features and data arrays
                scaled_bio = scaled_bio[row_idx, :]
                if n_qc > 0:
                    scaled_qc = scaled_qc[row_idx, :]
                feature_labels = [feature_labels[i] for i in row_idx]
            
            # Perform Hierarchical Clustering strictly on BIOLOGICAL samples
            ordered_data_list = []
            final_sample_names = []
            col_linkage = None
            
            if n_bio > 1:
                col_linkage = sch.linkage(scaled_bio.T, method='ward')
                
                col_dendro = sch.dendrogram(col_linkage, no_plot=True)
                col_idx = col_dendro['leaves']
                ordered_data_list.append(scaled_bio[:, col_idx])
                for i in col_idx:
                    info = bio_info[i]
                    final_sample_names.append(f"{info['name']}\n[{info['group']}]")
            elif n_bio == 1:
                ordered_data_list.append(scaled_bio)
                final_sample_names.append(f"{bio_info[0]['name']}\n[{bio_info[0]['group']}]")
                
            # Append QC samples to the far right (unclustered)
            if n_qc > 0:
                ordered_data_list.append(scaled_qc)
                for info in qc_info:
                    final_sample_names.append(f"{info['name']}\n[{info['group']}]")
            
            # Reconstruct final ordered data array
            ordered_data = np.hstack(ordered_data_list)
            
            self.vis_manager.create_heatmap_plot(
                ordered_data, feature_labels, final_sample_names, col_linkage, row_linkage, n_bio, n_qc, cmap
            )
            
            self.update_status("✓ Heatmap generated successfully")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"Heatmap failed:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error generating heatmap")

    # ==================== Venn Diagram Analysis ====================

    def calculate_venn_sets(self, selected_groups, threshold):
        """Calculate quantitative feature sets for Venn diagram based on detection frequency."""
        if getattr(self, 'preprocessed_data', None) is None:
            messagebox.showwarning("Warning", "Preprocessed data is not available. Please run preprocessing first.")
            return None

        data_source = self.preprocessed_data
        feature_ids = data_source.iloc[:, 0].values
        all_sample_cols = [col for col in data_source.columns if '.mzML' in col]

        final_sets = {g: set() for g in selected_groups}

        for group in selected_groups:
            # Collect sample columns for this group
            # Prefer _avg columns
            group_cols = []
            for base_name, reps in self.replicate_mapping.items():
                if self.group_mapping.get(base_name) == group:
                    avg_col = f"{base_name}_avg.mzML Peak area"
                    if avg_col in all_sample_cols:
                        group_cols.append(avg_col)
                    else:
                        rep_cols = [c for c, _ in reps if c in all_sample_cols]
                        group_cols.extend(rep_cols)

            if not group_cols:
                continue

            # Extract values, calculate frequency > 0
            vals = data_source[group_cols].values
            vals_zeroed = np.nan_to_num(vals, nan=0.0)
            detected_counts = np.sum(vals_zeroed > 0, axis=1)
            freqs = (detected_counts / len(group_cols)) * 100

            # Features meeting threshold
            pass_mask = freqs >= threshold
            passing_features = feature_ids[pass_mask]
            
            final_sets[group] = set(passing_features)

        return final_sets
    
    def run_venn_analysis(self):
        """Generate Venn Diagram"""
        selected_indices = self.venn_group_listbox.curselection()
        selected_groups = [self.venn_group_listbox.get(i) for i in selected_indices]
        
        if len(selected_groups) < 2 or len(selected_groups) > 3:
            messagebox.showwarning("Warning", "Venn diagrams only support 2 to 3 groups.\nFor more groups, please use the UpSet Plot tab.")
            return
            
        if getattr(self, 'preprocessed_data', None) is None:
            messagebox.showwarning("Warning", "Please run preprocessing first.")
            return
            
        try:
            self.update_status("Calculating Venn intersections...")
            threshold = self.venn_thresh_var.get()
            
            self.venn_sets = self.calculate_venn_sets(
                selected_groups, 
                threshold
            )
            
            if not self.venn_sets:
                messagebox.showerror("Error", "Could not calculate feature sets.")
                return
                
            self.vis_manager.create_venn_plot(self.venn_sets, selected_groups)
            
            # Populate intersections dropdown
            feature_map = defaultdict(list)
            for group_name, features in self.venn_sets.items():
                for f in features:
                    feature_map[f].append(group_name)
            
            self.venn_signature_map = defaultdict(list)
            for f, group_list in feature_map.items():
                group_list.sort()
                signature = " & ".join(group_list)
                self.venn_signature_map[signature].append(f)
                
            sorted_signatures = sorted(self.venn_signature_map.keys(), key=lambda s: (s.count('&'), s), reverse=True)
            if hasattr(self, 'venn_intersection_combo'):
                self.venn_intersection_combo['values'] = sorted_signatures
                if sorted_signatures:
                    self.venn_intersection_combo.set(sorted_signatures[0])
                    self.update_venn_table()
                else:
                    self.venn_intersection_combo.set('')
                    self.venn_tree.delete(*self.venn_tree.get_children())
            
            self.update_status("✓ Venn diagram generated")
            
        except Exception as e:
            import traceback
            messagebox.showerror("Error", f"Venn analysis failed:\n{str(e)}\n{traceback.format_exc()}")

    def update_venn_table(self, event=None):
        """Update the Venn features table based on the selected intersection"""
        if not hasattr(self, 'venn_tree'):
            return
            
        self.venn_tree.delete(*self.venn_tree.get_children())
        
        selected_sig = self.venn_intersection_var.get()
        if not selected_sig or not hasattr(self, 'venn_signature_map'):
            return
            
        features = self.venn_signature_map.get(selected_sig, [])
        for f in features:
            fid = str(f)
            f_id = fid
            f_mz = ""
            f_rt = ""
            try:
                parts = fid.rsplit('_', 2)
                if len(parts) == 3:
                    f_id = parts[0]
                    f_mz = parts[1]
                    f_rt = parts[2]
            except Exception:
                pass
                
            self.venn_tree.insert('', 'end', values=(f_id, f_mz, f_rt, fid))

    def on_venn_table_double_click(self, event):
        """Handle double-click on Venn table row to open in Feature Viewer"""
        selection = self.venn_tree.selection()
        if not selection:
            return
            
        item = selection[0]
        values = self.venn_tree.item(item, "values")
        if not values or len(values) < 4:
            return
            
        feature_id = values[3] # FeatureID is the 4th column
        
        for i in range(self.notebook.index("end")):
            if "Feature Viewer" in self.notebook.tab(i, "text"):
                self.notebook.select(i)
                self.feature_viewer_id_var.set(feature_id)
                self.run_feature_plot()
                break

    def create_feature_viewer_tab(self):
        """Tab for visualizing individual feature distributions."""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔬 Feature Viewer")

        # Controls
        control_frame = ttk.LabelFrame(tab, text="Feature Viewer Settings", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)

        # Left side: Feature search and select
        feat_frame = ttk.Frame(control_frame)
        feat_frame.pack(side='left', fill='y', padx=(0, 20))

        ttk.Label(feat_frame, text="Search ID:").grid(row=0, column=0, sticky='w', pady=2)
        self.feature_search_var = tk.StringVar()
        self.feature_search_var.trace_add('write', self.filter_feature_viewer_list)
        search_entry = ttk.Entry(feat_frame, textvariable=self.feature_search_var, width=20)
        search_entry.grid(row=0, column=1, sticky='w', pady=2)

        ttk.Label(feat_frame, text="Select Feature:").grid(row=1, column=0, sticky='w', pady=2)
        self.feature_viewer_id_var = tk.StringVar()
        self.feature_viewer_combo = ttk.Combobox(feat_frame, textvariable=self.feature_viewer_id_var, width=40)
        self.feature_viewer_combo.grid(row=1, column=1, sticky='w', pady=2)
        self.feature_viewer_combo.bind('<<ComboboxSelected>>', self.run_feature_plot)
        self.feature_viewer_combo.bind('<Return>', self.run_feature_plot)

        ttk.Label(feat_frame, text="Data Source:").grid(row=2, column=0, sticky='w', pady=2)
        self.fv_data_source_var = tk.StringVar(value="Preprocessed Data")
        ttk.Combobox(feat_frame, textvariable=self.fv_data_source_var, 
                     values=["Raw Data", "Preprocessed Data"], state='readonly', width=17).grid(row=2, column=1, sticky='w', pady=2)
                     
        self.fv_log_var = tk.BooleanVar(value=False) # Kept for session backwards compatibility
        trans_frame = ttk.Frame(feat_frame)
        trans_frame.grid(row=3, column=1, sticky='w', pady=2)
        ttk.Label(trans_frame, text="Transform:").pack(side='left')
        self.fv_transform_var = tk.StringVar(value="None")
        fv_trans_combo = ttk.Combobox(trans_frame, textvariable=self.fv_transform_var, 
                                      values=["None", "Log10", "Square Root", "Cube Root"], 
                                      state='readonly', width=11)
        fv_trans_combo.pack(side='left', padx=2)
        fv_trans_combo.bind('<<ComboboxSelected>>', self.run_feature_plot)

        self.feature_plot_exclude_qc_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            feat_frame,
            text="Exclude QC Samples",
            variable=self.feature_plot_exclude_qc_var,
            command=self.run_feature_plot
        ).grid(row=3, column=0, sticky='w', pady=2)

        self.fv_impute_lod_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            feat_frame,
            text="Apply LOD Imputation",
            variable=self.fv_impute_lod_var,
            command=self.run_feature_plot
        ).grid(row=4, column=0, sticky='w', pady=2)
        
        self.fv_lod_fraction_var = tk.StringVar(value="1/5 (20%)")
        fv_lod_combo = ttk.Combobox(
            feat_frame,
            textvariable=self.fv_lod_fraction_var,
            values=["1/2 (50%)", "1/3 (33%)", "1/4 (25%)", "1/5 (20%)"],
            state='readonly',
            width=10
        )
        fv_lod_combo.grid(row=4, column=1, sticky='w', pady=2)
        fv_lod_combo.bind('<<ComboboxSelected>>', self.run_feature_plot)

        self.fv_show_violin_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            feat_frame,
            text="Show Violin Plot",
            variable=self.fv_show_violin_var,
            command=self.run_feature_plot
        ).grid(row=5, column=0, sticky='w', pady=2)

        self.fv_show_boxplot_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            feat_frame,
            text="Show Box Plot",
            variable=self.fv_show_boxplot_var,
            command=self.run_feature_plot
        ).grid(row=5, column=1, sticky='w', pady=2)

        self.fv_show_points_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            feat_frame,
            text="Show Data Points",
            variable=self.fv_show_points_var,
            command=self.run_feature_plot
        ).grid(row=6, column=0, sticky='w', pady=2)

        btn_frame = ttk.Frame(feat_frame)
        btn_frame.grid(row=7, column=0, columnspan=2, sticky='w', pady=(10, 5))

        ttk.Button(
            btn_frame,
            text="▶️ Plot Feature",
            command=self.run_feature_plot
        ).pack(side='left', padx=(0, 5))
        
        ttk.Button(
            btn_frame,
            text="💾 Export Data",
            command=self.export_feature_data
        ).pack(side='left')

        # Right side: Group selection
        group_frame = ttk.LabelFrame(control_frame, text="Select Groups (All if none selected)", padding=5)
        group_frame.pack(side='left', fill='both', expand=True)

        self.fv_group_listbox = tk.Listbox(group_frame, selectmode='multiple', height=5, exportselection=0)
        self.fv_group_listbox.pack(side='left', fill='both', expand=True)
        sb = ttk.Scrollbar(group_frame, orient='vertical', command=self.fv_group_listbox.yview)
        sb.pack(side='right', fill='y')
        self.fv_group_listbox.config(yscrollcommand=sb.set)

        # Plot area
        self.feature_viewer_plot_frame = ttk.Frame(tab)
        self.feature_viewer_plot_frame.pack(fill='both', expand=True, padx=10, pady=5)

    def create_spectrum_tab(self):
        """Tab 10: Spectrum Search & Viewer"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔍 Spectrum Search")
        
        # Use a PanedWindow to split Search List and Plot
        paned = ttk.PanedWindow(tab, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # === Left Panel: Controls & Results ===
        left_panel = ttk.Frame(paned)
        paned.add(left_panel, weight=1)
        
        # 1. File Management
        control_frame = ttk.LabelFrame(left_panel, text="MGF File Management", padding=10)
        control_frame.pack(fill=tk.X, padx=5, pady=5)
        
    def create_spectrum_tab(self):
        """Tab 10: Spectrum Search & Viewer"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔍 Spectrum Search")
        
        # Use a PanedWindow to split Search List and Plot
        paned = ttk.PanedWindow(tab, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # === Left Panel: Controls & Results ===
        left_panel = ttk.Frame(paned)
        paned.add(left_panel, weight=1)
        
        # 1. File Management
        control_frame = ttk.LabelFrame(left_panel, text="MGF File Management", padding=10)
        control_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(
            control_frame,
            text="📂 Load MGF File(s)",
            command=self.load_mgf_files
        ).pack(side='left', padx=5)
        
        self.mgf_status_label = ttk.Label(control_frame, text="No MGF files loaded", foreground="gray")
        self.mgf_status_label.pack(side='left', padx=10)
        
        # 2. Search Frame
        search_frame = ttk.LabelFrame(left_panel, text="Search Spectra", padding=10)
        search_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Search Type
        ttk.Label(search_frame, text="Search by:").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.spectrum_search_type_var = tk.StringVar(value="Feature ID")
        search_type_combo = ttk.Combobox(
            search_frame,
            textvariable=self.spectrum_search_type_var,
            values=["Feature ID", "m/z", "RT range"],
            state="readonly",
            width=15
        )
        search_type_combo.grid(row=0, column=1, sticky='w', padx=5, pady=2)
        
        # Value
        ttk.Label(search_frame, text="Value:").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.spectrum_search_value_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.spectrum_search_value_var, width=25).grid(
            row=1, column=1, sticky='ew', padx=5, pady=2
        )
        
        # Tolerance
        ttk.Label(search_frame, text="Tolerance:").grid(row=2, column=0, sticky='w', padx=5, pady=2)
        self.spectrum_tolerance_var = tk.StringVar(value="0.01")
        ttk.Entry(search_frame, textvariable=self.spectrum_tolerance_var, width=10).grid(
            row=2, column=1, sticky='w', padx=5, pady=2
        )
        ttk.Label(search_frame, text="(Da for m/z)").grid(row=2, column=2, sticky='w')

        button_frame = ttk.Frame(search_frame)
        button_frame.grid(row=3, column=0, columnspan=3, pady=10)

        ttk.Button(button_frame, text="🔍 Search", command=self.search_spectra).pack(side='left', padx=5)
        ttk.Button(button_frame, text="📝 Annotate Important", 
                  command=self.annotate_features_with_spectra).pack(side='left', padx=5)
        
        # 3. Results List
        results_frame = ttk.LabelFrame(left_panel, text="Search Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ("ID", "m/z", "RT", "File")
        self.spectrum_tree = ttk.Treeview(results_frame, columns=columns, show='headings', selectmode='extended')
        self.spectrum_tree.heading("ID", text="Feature ID", command=lambda: self.treeview_sort_column(self.spectrum_tree, "ID", False))
        self.spectrum_tree.column("ID", width=120)
        self.spectrum_tree.heading("m/z", text="Precursor m/z", command=lambda: self.treeview_sort_column(self.spectrum_tree, "m/z", False))
        self.spectrum_tree.column("m/z", width=80)
        self.spectrum_tree.heading("RT", text="RT (s)", command=lambda: self.treeview_sort_column(self.spectrum_tree, "RT", False))
        self.spectrum_tree.column("RT", width=60)
        self.spectrum_tree.heading("File", text="Source File", command=lambda: self.treeview_sort_column(self.spectrum_tree, "File", False))
        self.spectrum_tree.column("File", width=100)
        
        sb = ttk.Scrollbar(results_frame, orient="vertical", command=self.spectrum_tree.yview)
        self.spectrum_tree.configure(yscrollcommand=sb.set)
        
        self.spectrum_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind Selection
        self.spectrum_tree.bind('<<TreeviewSelect>>', self.on_spectrum_select)
        
        # === Right Panel: Plot ===
        right_panel = ttk.Frame(paned)
        paned.add(right_panel, weight=2)
        
        # Plot Options
        plot_ops = ttk.LabelFrame(right_panel, text="Plot Options", padding=5)
        plot_ops.pack(fill=tk.X, padx=5, pady=5)
        
        self.spec_min_mz_var = tk.StringVar()
        ttk.Label(plot_ops, text="Min m/z:").pack(side=tk.LEFT, padx=2)
        ttk.Entry(plot_ops, textvariable=self.spec_min_mz_var, width=6).pack(side=tk.LEFT, padx=2)

        self.spec_max_mz_var = tk.StringVar()
        ttk.Label(plot_ops, text="Max m/z:").pack(side=tk.LEFT, padx=2)
        ttk.Entry(plot_ops, textvariable=self.spec_max_mz_var, width=6).pack(side=tk.LEFT, padx=2)

        self.spec_threshold_var = tk.DoubleVar(value=0.0)
        ttk.Label(plot_ops, text="Min Int:").pack(side=tk.LEFT, padx=2)
        ttk.Entry(plot_ops, textvariable=self.spec_threshold_var, width=6).pack(side=tk.LEFT, padx=2)
        
        self.spec_show_labels_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(plot_ops, text="Show Labels", variable=self.spec_show_labels_var, 
                       command=self.refresh_spectrum_plot).pack(side=tk.LEFT, padx=5)
        
        self.spec_normalize_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(plot_ops, text="Normalize %", variable=self.spec_normalize_var, 
                       command=self.refresh_spectrum_plot).pack(side=tk.LEFT, padx=5)
                       
        ttk.Button(plot_ops, text="🔄 Refresh", command=self.refresh_spectrum_plot).pack(side=tk.LEFT, padx=5)
        
        # Plot Canvas Area
        self.spectrum_plot_frame = ttk.Frame(right_panel)
        self.spectrum_plot_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Initialize internal state
        self.current_spectrum_matches = []
        self.selected_spectrum_data = None
        self.spec_draggable_labels = []
        self.spec_dragged_artist = None
        self.spec_drag_press_info = None

    def load_mgf_files(self):
        """Load MGF file(s) for spectrum search"""
        filepaths = filedialog.askopenfilenames(
            title="Select MGF file(s)",
            filetypes=[("MGF files", "*.mgf"), ("All files", "*.*")]
        )
        
        if not filepaths:
            return
        
        try:
            self.update_status("Loading MGF files...")
            num_spectra = self.mgf_engine.load_mgf_files(list(filepaths))
            self.mgf_status_label.config(
                text=f"Loaded {num_spectra} spectra from {len(filepaths)} file(s)",
                foreground="green"
            )
            self.update_status(f"✓ Loaded {num_spectra} spectra")
            messagebox.showinfo("Success", f"Loaded {num_spectra} spectra")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load MGF files:\n{str(e)}")
            self.update_status("Error loading MGF")

    def search_spectra(self):
        """Search loaded spectra"""
        if not self.mgf_engine.spectra:
            messagebox.showwarning("Warning", "Please load MGF files first!")
            return
        
        search_type = self.spectrum_search_type_var.get()
        search_value = self.spectrum_search_value_var.get().strip()
        
        if not search_value:
            messagebox.showwarning("Warning", "Please enter a search value!")
            return
        
        try:
            # Clear previous results
            for item in self.spectrum_tree.get_children():
                self.spectrum_tree.delete(item)
            self.current_spectrum_matches = []
            self.selected_spectrum_data = None
            
            self.update_status("Searching spectra...")
            
            matches = []
            
            if search_type == "Feature ID":
                feature_ids = search_value.split()
                matches = self.mgf_engine.search_by_feature_id(feature_ids)
                
            elif search_type == "m/z":
                target_mz = float(search_value)
                tolerance = float(self.spectrum_tolerance_var.get())
                matches = self.mgf_engine.search_by_mz(target_mz, tolerance)
                
            elif search_type == "RT range":
                if '-' in search_value:
                    rt_min, rt_max = map(float, search_value.split('-'))
                    matches = self.mgf_engine.search_by_rt(rt_min, rt_max)
                else:
                    messagebox.showwarning("Warning", "RT range format should be min-max (e.g., 5-10)")
                    return
            
            self.current_spectrum_matches = matches
            
            # Populate Treeview
            for i, spec in enumerate(matches):
                fid = spec.get('FEATUREID', spec.get('SCANS', 'N/A'))
                pepmass = spec.get('PEPMASS', 'N/A')
                rt = spec.get('RTINSECONDS', 'N/A')
                src = spec.get('SOURCEFILE', 'N/A')
                # Treeview values
                self.spectrum_tree.insert('', 'end', iid=str(i), values=(fid, pepmass, rt, src))
            
            self.update_status(f"Found {len(matches)} matches")
            
        except Exception as e:
            messagebox.showerror("Error", f"Search failed:\n{str(e)}")
            self.update_status("Error searching spectra")

    def on_spectrum_select(self, event):
        """Handle selection in spectrum list"""
        selected_items = self.spectrum_tree.selection()
        if not selected_items:
            return
            
        if len(selected_items) == 1:
            index = int(selected_items[0])
            if 0 <= index < len(self.current_spectrum_matches):
                self.selected_spectrum_data = [self.current_spectrum_matches[index]]
                self.refresh_spectrum_plot()
        elif len(selected_items) >= 2:
            idx1, idx2 = int(selected_items[0]), int(selected_items[1])
            self.selected_spectrum_data = [
                self.current_spectrum_matches[idx1], 
                self.current_spectrum_matches[idx2]
            ]
            self.refresh_spectrum_plot()

    def refresh_spectrum_plot(self):
        """Draw the selected spectrum in the plot frame"""
        if not self.selected_spectrum_data:
            return

        # Clear existing plot
        for widget in self.spectrum_plot_frame.winfo_children():
            widget.destroy()
            
        self.spec_draggable_labels.clear()

        # Settings
        try:
            threshold = float(self.spec_threshold_var.get())
        except ValueError:
            threshold = 0.0
            
        try:
            min_mz = float(self.spec_min_mz_var.get()) if self.spec_min_mz_var.get() else None
        except ValueError:
            min_mz = None
            self.spec_min_mz_var.set("")
            
        try:
            max_mz = float(self.spec_max_mz_var.get()) if self.spec_max_mz_var.get() else None
        except ValueError:
            max_mz = None
            self.spec_max_mz_var.set("")
            
        show_labels = self.spec_show_labels_var.get()
        normalize = self.spec_normalize_var.get()
        
        fig = Figure(figsize=(5, 4), dpi=100)
        self.generated_plots['Spectrum'] = fig # Store for export
        ax = fig.add_subplot(111)
        
        is_mirror = len(self.selected_spectrum_data) == 2
        all_mzs = []
        
        # Prepare annotation for hover
        annot = ax.annotate("", xy=(0,0), xytext=(0,10), textcoords="offset points",
                            bbox=dict(boxstyle="round", fc="yellow", alpha=0.8),
                            arrowprops=dict(arrowstyle="->"))
        annot.set_visible(False)
        self.spec_hover_annot = annot

        self.spec_plot_data = [] # To store plotted data for hover logic

        for idx, spec in enumerate(self.selected_spectrum_data):
            mzs = np.array(spec.get('mzs', []))
            intensities = np.array(spec.get('intensities', []))
            
            mask = np.ones_like(mzs, dtype=bool)
            if threshold > 0: mask &= (intensities >= threshold)
            if min_mz is not None: mask &= (mzs >= min_mz)
            if max_mz is not None: mask &= (mzs <= max_mz)

            mzs_plot = mzs[mask]
            ints_plot = intensities[mask]
            
            if normalize and len(ints_plot) > 0:
                max_int = ints_plot.max()
                if max_int > 0:
                    ints_plot = (ints_plot / max_int) * 100
            
            direction = -1 if (is_mirror and idx == 1) else 1
            color = 'red' if (is_mirror and idx == 1) else 'black'
            
            all_mzs.extend(mzs_plot)
            self.spec_plot_data.append((mzs_plot, ints_plot * direction))
            
            if len(mzs_plot) > 0:
                label = f"Spec {idx+1}: {spec.get('SCANS', spec.get('FEATUREID', ''))}"
                ax.vlines(mzs_plot, 0, ints_plot * direction, color=color, linewidth=1, label=label)
                
                # Highlight matched peaks if search was by m/z
                if 'matched_mzs' in spec:
                    matched = np.array(spec['matched_mzs'])
                    for mmz in matched:
                        diffs = np.abs(mzs_plot - mmz)
                        if diffs.size > 0:
                            closest_idx = diffs.argmin()
                            if diffs[closest_idx] < 0.01:
                                ax.vlines(mzs_plot[closest_idx], 0, ints_plot[closest_idx] * direction, color='blue', linewidth=2)
                                tol = float(self.spectrum_tolerance_var.get())
                                ax.axvspan(mmz - tol, mmz + tol, color='grey', alpha=0.2)

                # Labels
                if show_labels:
                    idx_sorted = np.argsort(ints_plot)[::-1]
                    for i in idx_sorted[:20]:
                        mz = mzs_plot[i]
                        inten = ints_plot[i] * direction
                        va = 'top' if direction == -1 else 'bottom'
                        txt = ax.text(mz, inten, f"{mz:.4f}", ha='center', va=va, fontsize=8, rotation=90)
                        self.spec_draggable_labels.append(txt)

        if is_mirror:
            ax.axhline(0, color='black', linewidth=0.5)

        ylabel = "Relative Intensity (%)" if normalize else "Intensity"
        ax.set_xlabel("m/z")
        ax.set_ylabel(ylabel)
        
        if is_mirror:
            title = "Mirror Plot Comparison"
            ax.legend(loc='best', fontsize=8)
            from matplotlib.ticker import FuncFormatter
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: f"{abs(x):.0f}"))
        else:
            spec = self.selected_spectrum_data[0]
            title = f"ID: {spec.get('FEATUREID', spec.get('SCANS', ''))} | Precursor: {spec.get('PEPMASS', '')}"
            
        ax.set_title(title, fontsize=10)
        ax.grid(True, alpha=0.3)
        
        if all_mzs:
            final_min_mz = min_mz if min_mz is not None else min(all_mzs) - 10
            final_max_mz = max_mz if max_mz is not None else max(all_mzs) + 10
            ax.set_xlim(final_min_mz, final_max_mz)

        # Embed
        canvas = FigureCanvasTkAgg(fig, master=self.spectrum_plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Connect events
        canvas.mpl_connect("motion_notify_event", self.on_spec_hover)
        canvas.mpl_connect("button_press_event", self.on_spec_press)
        canvas.mpl_connect("button_release_event", self.on_spec_release)
        canvas.mpl_connect("motion_notify_event", self.on_spec_drag)
        
        toolbar = NavigationToolbar2Tk(canvas, self.spectrum_plot_frame)
        toolbar.update()

    def on_spec_hover(self, event):
        """Handle mouse hover events on the spectrum plot canvas."""
        if not hasattr(self, 'spec_hover_annot') or not self.spec_hover_annot:
            return
            
        if event.inaxes != self.spec_hover_annot.axes:
            return

        if getattr(self, 'spec_dragged_artist', None): 
            return

        closest_mz = None
        closest_int = None
        min_dx = float('inf')
        
        for mzs_plot, ints_plot in self.spec_plot_data:
            if len(mzs_plot) == 0:
                continue
                
            dx = np.abs(mzs_plot - event.xdata)
            idx = np.argmin(dx)
            
            mz_range = mzs_plot.max() - mzs_plot.min()
            x_tolerance = mz_range * 0.005 if mz_range > 0 else 0.1
            
            if dx[idx] < x_tolerance and dx[idx] < min_dx:
                if ints_plot[idx] > 0 and 0 <= event.ydata <= ints_plot[idx]:
                    min_dx = dx[idx]
                    closest_mz = mzs_plot[idx]
                    closest_int = ints_plot[idx]
                elif ints_plot[idx] < 0 and ints_plot[idx] <= event.ydata <= 0:
                    min_dx = dx[idx]
                    closest_mz = mzs_plot[idx]
                    closest_int = ints_plot[idx]

        if closest_mz is not None:
            self.spec_hover_annot.set_text(f"m/z: {closest_mz:.4f}\nInt: {abs(closest_int):.1f}")
            self.spec_hover_annot.set_position((closest_mz, closest_int))
            self.spec_hover_annot.set_visible(True)
            event.canvas.draw_idle()
            self.update_status(f"Hovering over: m/z={closest_mz:.4f}, Intensity={abs(closest_int):.2f}")
        else:
            if self.spec_hover_annot.get_visible():
                self.spec_hover_annot.set_visible(False)
                event.canvas.draw_idle()
                self.update_status("Spectrum Plot: Hover over a peak to see details.")

    def on_spec_press(self, event):
        """Handle mouse button press for dragging labels."""
        for label in getattr(self, 'spec_draggable_labels', []):
            contains, _ = label.contains(event)
            if contains:
                if event.button == 1: # Left click
                    self.spec_dragged_artist = label
                    x0, y0 = label.get_position()
                    self.spec_drag_press_info = (x0, y0, event.xdata, event.ydata)
                    return
                elif event.button == 3: # Right click
                    label.remove()
                    self.spec_draggable_labels.remove(label)
                    event.canvas.draw_idle()
                    return

    def on_spec_release(self, event):
        """Handle mouse button release to stop dragging."""
        self.spec_dragged_artist = None
        self.spec_drag_press_info = None

    def on_spec_drag(self, event):
        """Handle mouse motion for dragging."""
        if getattr(self, 'spec_dragged_artist', None) is None or event.inaxes is None:
            return

        x0, y0, xpress, ypress = self.spec_drag_press_info
        dx = event.xdata - xpress
        dy = event.ydata - ypress

        self.spec_dragged_artist.set_position((x0 + dx, y0 + dy))
        event.canvas.draw_idle()

    def annotate_features_with_spectra(self):
        """Annotate important features (PLS-DA VIP + RF) with spectrum information"""
        if not self.mgf_engine.spectra:
            messagebox.showwarning("Warning", "Please load MGF files first!")
            return
        
        if self.plsda_result is None and self.rf_result is None:
            messagebox.showwarning("Warning", "Please run PLS-DA or Random Forest first!")
            return
        
        try:
            self.update_status("Annotating features...")
            important_features = set()
            
            if self.plsda_result:
                vips = self.plsda_result['vip_scores']
                feature_ids = self.screened_data.iloc[:, 0].values
                top_vip_idx = np.argsort(vips)[-50:]
                important_features.update(feature_ids[top_vip_idx])
            
            if self.rf_result:
                top_rf_features = [f[0] for f in self.rf_result['top_features'][:50]]
                important_features.update(top_rf_features)
            
            feature_df = pd.DataFrame({'FeatureID': list(important_features)})
            annotated_df = self.mgf_engine.annotate_features_with_spectra(feature_df)
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                initialfile="annotated_important_features.csv",
                filetypes=[("CSV files", "*.csv")]
            )
            
            if file_path:
                # Split FeatureID into separate columns
                annotated_df = ExportManager.process_feature_ids(annotated_df, 'FeatureID')
                
                annotated_df.to_csv(file_path, index=False)
                num_with_spectra = annotated_df['HasSpectrum'].sum()
                messagebox.showinfo(
                    "Success",
                    f"Annotated {len(annotated_df)} important features\n"
                    f"{num_with_spectra} have matching spectra ({num_with_spectra/len(annotated_df)*100:.1f}%)"
                )
                self.update_status("✓ Features annotated and exported")
        
        except Exception as e:
            messagebox.showerror("Error", f"Annotation failed:\n{str(e)}")
            self.update_status("Error annotating features")

    def export_venn_intersections(self):
        """Export the lists of features in each Venn intersection"""
        if not hasattr(self, 'venn_sets') or not self.venn_sets:
            messagebox.showwarning("Warning", "Please generate a Venn diagram first.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )
        
        if not file_path:
            return
            
        try:
            self.update_status("Exporting intersections...")
            
            feature_col = 'Feature_ID' if 'Feature_ID' in self.preprocessed_data.columns else 'FeatureID'
            
            feature_map = defaultdict(list)
            for group_name, features in self.venn_sets.items():
                for f in features:
                    feature_map[f].append(group_name)
            
            signature_map = defaultdict(list)
            for f, group_list in feature_map.items():
                group_list.sort()
                signature = " & ".join(group_list)
                signature_map[signature].append(f)
                
            export_data = {sig: feats for sig, feats in signature_map.items()}
            sorted_signatures = sorted(export_data.keys(), key=lambda s: (s.count('&'), s), reverse=True)
            
            if not sorted_signatures:
                self.update_status("Export cancelled: No features to export.")
                messagebox.showinfo("Info", "No features met the threshold criteria.")
                return

            if file_path.endswith('.xlsx'):
                venn_groups = list(self.venn_sets.keys())
                all_sample_cols = [c for c in self.preprocessed_data.columns if '.mzML' in c]
                group_cols_map = {}
                for g in venn_groups:
                    group_cols = []
                    for base_name, reps in self.replicate_mapping.items():
                        if self.group_mapping.get(base_name) == g:
                            avg_col = f"{base_name}_avg.mzML Peak area"
                            if avg_col in all_sample_cols:
                                group_cols.append(avg_col)
                            else:
                                rep_cols = [c for c, _ in reps if c in all_sample_cols]
                                group_cols.extend(rep_cols)
                    group_cols_map[g] = group_cols

                used_sheet_names = set()
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for name in sorted_signatures:
                        features = export_data[name]
                        
                        # Handle long or duplicate sheet names securely
                        base_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', name)[:31]
                        sheet_name = base_sheet_name
                        counter = 1
                        while sheet_name in used_sheet_names:
                            suffix = f"_{counter}"
                            sheet_name = base_sheet_name[:31 - len(suffix)] + suffix
                            counter += 1
                        used_sheet_names.add(sheet_name)
                        
                        if features:
                            mask = self.preprocessed_data[feature_col].isin(features)
                            df = self.preprocessed_data[mask].copy()
                            
                            # Safely align data_before_norm to avoid IndexingError
                            if self.data_before_norm is not None:
                                raw_mask = self.data_before_norm[feature_col].isin(features)
                                df_raw = self.data_before_norm[raw_mask].copy()
                                df_raw = df_raw.set_index(feature_col).loc[df[feature_col]].reset_index()
                            else:
                                df_raw = df.copy()

                            mean_cols = []
                            stat_cols = []
                            data_cols = []
                            
                            for g in venn_groups:
                                g_cols = group_cols_map[g]
                                data_cols.extend(g_cols)
                                
                                col_name = f'Mean_{g}'
                                if g_cols:
                                    df[col_name] = df[g_cols].mean(axis=1)
                                else:
                                    df[col_name] = 0
                                mean_cols.append(col_name)
                                
                                freq_col = f'Freq_pct_{g}'
                                if g_cols:
                                    # Use numeric arrays safely
                                    raw_vals = pd.to_numeric(df_raw[g_cols].values.flatten(), errors='coerce').reshape(df_raw[g_cols].shape)
                                    df[freq_col] = (np.sum((raw_vals > 0) & (~np.isnan(raw_vals)), axis=1) / len(g_cols)) * 100
                                else:
                                    df[freq_col] = 0
                                stat_cols.append(freq_col)
                            
                            if mean_cols:
                                df['Dominant_Group'] = df[mean_cols].idxmax(axis=1).apply(
                                    lambda x: str(x).replace('Mean_', '') if pd.notna(x) else 'None'
                                )
                            else:
                                df['Dominant_Group'] = "None"
                            
                            # Preserve original data columns too
                            cols = [feature_col, 'Dominant_Group'] + stat_cols + mean_cols + data_cols
                            # Filter unique columns maintaining order
                            final_cols = []
                            for c in cols:
                                if c in df.columns and c not in final_cols:
                                    final_cols.append(c)
                                    
                            df = df[final_cols]
                            
                            if 'Dominant_Group' in df.columns:
                                df = df.sort_values(by=['Dominant_Group', feature_col])
                            
                            df = ExportManager.process_feature_ids(df, feature_col)
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            
                            if PatternFill:
                                worksheet = writer.sheets[sheet_name]
                                # Identify Dominant_Group column index dynamically
                                dom_col_idx = df.columns.get_loc('Dominant_Group') + 1 if 'Dominant_Group' in df.columns else None
                                
                                for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
                                    if dom_col_idx:
                                        group = getattr(row_data, 'Dominant_Group', None)
                                        if group and group in self.group_colors:
                                            color_hex = self.group_colors[group].lstrip('#')
                                            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                                            for col_idx in range(1, len(df.columns) + 1):
                                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                                cell.fill = fill
                        else:
                            pd.DataFrame(columns=['ID', 'm/z', 'RT', feature_col]).to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # For CSV, we just dump them all in one long list with a "Type" column
                all_rows = []
                for name in sorted_signatures:
                    for f in export_data[name]:
                        row = {'Type': name, feature_col: f}
                        try:
                            parts = str(f).rsplit('_', 2)
                            if len(parts) == 3:
                                row['ID'] = parts[0]
                                row['m/z'] = parts[1]
                                row['RT'] = parts[2]
                        except:
                            pass
                        all_rows.append(row)
                
                df = pd.DataFrame(all_rows)
                cols = ['Type', 'ID', 'm/z', 'RT', feature_col]
                final_cols = [c for c in cols if c in df.columns]
                df[final_cols].to_csv(file_path, index=False)
                
            messagebox.showinfo("Success", f"Exported Venn intersections to:\n{file_path}")
            self.update_status("✓ Export complete")
            
        except Exception as e:
            import traceback
            messagebox.showerror("Error", f"Export failed:\n{str(e)}\n\nDetails:\n{traceback.format_exc()}")
            self.update_status("Export failed")

    def export_venn_data_subset(self, intersection_type):
        """
        Export data for a specific Venn intersection type (Common or Unique).
        Triggered from Venn Diagram context menu.
        """
        if not hasattr(self, 'venn_sets') or not self.venn_sets:
            messagebox.showwarning("Warning", "Please generate a Venn diagram first.")
            return
            
        groups = list(self.venn_sets.keys())
        if len(groups) < 2: return
        
        # Calculate intersections
        sets_values = list(self.venn_sets.values())
        common = set.intersection(*sets_values)
        
        unique_sets = {}
        for g in groups:
            others = [self.venn_sets[og] for og in groups if og != g]
            if others:
                unique_sets[g] = self.venn_sets[g] - set.union(*others)
            else:
                unique_sets[g] = self.venn_sets[g]
            
        final_features = set()
        suffix = ""
        
        if intersection_type == 'COMMON':
            final_features = common
            suffix = "common"
        elif intersection_type == 'UNIQUE_COMBINED':
            final_features = set.union(*unique_sets.values())
            suffix = "unique_all_groups"
        elif intersection_type.startswith('UNIQUE_'):
            g_name = intersection_type[7:] # Remove 'UNIQUE_'
            if g_name in unique_sets:
                final_features = unique_sets[g_name]
                suffix = f"unique_{g_name}"
        
        if not final_features:
            messagebox.showinfo("Info", f"No features found for {suffix}.")
            return

        if getattr(self, 'preprocessed_data', None) is None: return
        
        # Filter data for these features
        subset_df = self.preprocessed_data[self.preprocessed_data.iloc[:, 0].isin(final_features)]
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=f"venn_{suffix}_features_data.csv",
            filetypes=[("CSV files", "*.csv")]
        )
        
        if file_path:
            try:
                # Split FeatureID into separate columns
                if 'Feature_ID' in subset_df.columns:
                    subset_df = ExportManager.process_feature_ids(subset_df.copy(), 'Feature_ID')

                subset_df.to_csv(file_path, index=False)
                messagebox.showinfo("Success", f"Exported {len(subset_df)} features to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Export failed:\n{str(e)}")

    # ==================== UpSet Plot Analysis ====================

    def create_upset_tab(self):
        """Tab 10: UpSet Plot Analysis"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📊 UpSet Plot")
        
        control_frame = ttk.LabelFrame(tab, text="UpSet Plot Settings", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(control_frame, text="Select Groups (≥2):").grid(row=0, column=0, sticky='w', padx=5)
        
        list_frame = ttk.Frame(control_frame)
        list_frame.grid(row=1, column=0, rowspan=3, padx=5, pady=5, sticky='ns')
        
        self.upset_group_listbox = tk.Listbox(list_frame, selectmode='multiple', height=5, exportselection=0, width=25)
        self.upset_group_listbox.pack(side='left', fill='y')
        sb = ttk.Scrollbar(list_frame, orient='vertical', command=self.upset_group_listbox.yview)
        sb.pack(side='right', fill='y')
        self.upset_group_listbox.config(yscrollcommand=sb.set)
        
        ttk.Label(control_frame, text="Detection Frequency Threshold (%):").grid(row=1, column=1, sticky='w', padx=15)
        self.upset_thresh_var = tk.DoubleVar(value=80.0)
        ttk.Spinbox(control_frame, from_=0, to=100, increment=5, textvariable=self.upset_thresh_var, width=10).grid(row=2, column=1, sticky='nw', padx=15)
        
        self.upset_desc_label = ttk.Label(control_frame, text="(Feature must be > 0 in ≥ X% of group samples)", foreground="gray")
        self.upset_desc_label.grid(row=3, column=1, sticky='nw', padx=15)
        
        btn_frame = ttk.Frame(control_frame)
        btn_frame.grid(row=1, column=2, rowspan=3, padx=20)
        
        ttk.Button(btn_frame, text="▶️ Generate UpSet Plot", command=self.run_upset_analysis).pack(fill='x', pady=2)
        ttk.Button(btn_frame, text="💾 Export Intersections", command=self.export_upset_intersections).pack(fill='x', pady=2)
        
        self.upset_paned = ttk.PanedWindow(tab, orient=tk.VERTICAL)
        self.upset_paned.pack(fill='both', expand=True, padx=10, pady=5)

        self.upset_plot_frame = ttk.Frame(self.upset_paned)
        self.upset_paned.add(self.upset_plot_frame, weight=3)

        self.upset_table_frame = ttk.LabelFrame(self.upset_paned, text="Intersection Features")
        self.upset_paned.add(self.upset_table_frame, weight=1)

        table_control = ttk.Frame(self.upset_table_frame)
        table_control.pack(fill='x', padx=5, pady=5)
        ttk.Label(table_control, text="Select Intersection:").pack(side='left', padx=5)
        
        self.upset_intersection_var = tk.StringVar()
        self.upset_intersection_combo = ttk.Combobox(table_control, textvariable=self.upset_intersection_var, state='readonly', width=50)
        self.upset_intersection_combo.pack(side='left', padx=5)
        self.upset_intersection_combo.bind('<<ComboboxSelected>>', self.update_upset_table)

        columns = ("ID", "m/z", "RT", "FeatureID")
        self.upset_tree = ttk.Treeview(self.upset_table_frame, columns=columns, show='headings', selectmode='browse')
        self.upset_tree.heading("ID", text="ID", command=lambda: self.treeview_sort_column(self.upset_tree, "ID", False))
        self.upset_tree.column("ID", width=120)
        self.upset_tree.heading("m/z", text="m/z", command=lambda: self.treeview_sort_column(self.upset_tree, "m/z", False))
        self.upset_tree.column("m/z", width=100)
        self.upset_tree.heading("RT", text="RT", command=lambda: self.treeview_sort_column(self.upset_tree, "RT", False))
        self.upset_tree.column("RT", width=100)
        self.upset_tree.heading("FeatureID", text="Full Feature ID", command=lambda: self.treeview_sort_column(self.upset_tree, "FeatureID", False))
        self.upset_tree.column("FeatureID", width=250)

        sb2 = ttk.Scrollbar(self.upset_table_frame, orient="vertical", command=self.upset_tree.yview)
        self.upset_tree.configure(yscrollcommand=sb2.set)
        
        self.upset_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        sb2.pack(side=tk.RIGHT, fill=tk.Y, pady=5)
        
        self.upset_tree.bind("<Double-1>", self.on_upset_table_double_click)
        self.upset_tree.bind("<Button-3>", lambda e: self.show_tree_context_menu(e, self.upset_tree))

    def run_upset_analysis(self):
        """Generate UpSet Plot"""
        selected_indices = self.upset_group_listbox.curselection()
        selected_groups = [self.upset_group_listbox.get(i) for i in selected_indices]
        
        if len(selected_groups) < 2:
            messagebox.showwarning("Warning", "Please select at least 2 groups for UpSet plot analysis.")
            return
            
        if getattr(self, 'preprocessed_data', None) is None:
            messagebox.showwarning("Warning", "Please run preprocessing first.")
            return
            
        try:
            self.update_status("Calculating UpSet intersections...")
            threshold = self.upset_thresh_var.get()
            
            self.upset_sets = self.calculate_venn_sets(selected_groups, threshold)
            
            if not self.upset_sets:
                messagebox.showerror("Error", "Could not calculate feature sets.")
                return
                
            self.vis_manager.create_upset_plot(self.upset_sets, selected_groups)
            
            feature_map = defaultdict(list)
            for group_name, features in self.upset_sets.items():
                for f in features:
                    feature_map[f].append(group_name)
            
            self.upset_signature_map = defaultdict(list)
            for f, group_list in feature_map.items():
                group_list.sort()
                signature = " & ".join(group_list)
                self.upset_signature_map[signature].append(f)
                
            sorted_signatures = sorted(self.upset_signature_map.keys(), key=lambda s: (s.count('&'), s), reverse=True)
            if hasattr(self, 'upset_intersection_combo'):
                self.upset_intersection_combo['values'] = sorted_signatures
                if sorted_signatures:
                    self.upset_intersection_combo.set(sorted_signatures[0])
                    self.update_upset_table()
                else:
                    self.upset_intersection_combo.set('')
                    self.upset_tree.delete(*self.upset_tree.get_children())
            
            self.update_status("✓ UpSet plot generated")
            
        except Exception as e:
            import traceback
            messagebox.showerror("Error", f"UpSet analysis failed:\n{str(e)}\n{traceback.format_exc()}")

    def update_upset_table(self, event=None):
        """Update the UpSet features table based on the selected intersection"""
        if not hasattr(self, 'upset_tree'):
            return
            
        self.upset_tree.delete(*self.upset_tree.get_children())
        
        selected_sig = self.upset_intersection_var.get()
        if not selected_sig or not hasattr(self, 'upset_signature_map'):
            return
            
        features = self.upset_signature_map.get(selected_sig, [])
        for f in features:
            fid = str(f)
            f_id, f_mz, f_rt = fid, "", ""
            try:
                parts = fid.rsplit('_', 2)
                if len(parts) == 3:
                    f_id, f_mz, f_rt = parts[0], parts[1], parts[2]
            except Exception:
                pass
            self.upset_tree.insert('', 'end', values=(f_id, f_mz, f_rt, fid))

    def on_upset_table_double_click(self, event):
        """Handle double-click on UpSet table row to open in Feature Viewer"""
        selection = self.upset_tree.selection()
        if not selection: return
        values = self.upset_tree.item(selection[0], "values")
        if not values or len(values) < 4: return
        
        for i in range(self.notebook.index("end")):
            if "Feature Viewer" in self.notebook.tab(i, "text"):
                self.notebook.select(i)
                self.feature_viewer_id_var.set(values[3])
                self.run_feature_plot()
                break

    def export_upset_intersections(self):
        """Export the lists of features in each UpSet intersection"""
        if not hasattr(self, 'upset_sets') or not self.upset_sets:
            messagebox.showwarning("Warning", "Please generate an UpSet plot first.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )
        if not file_path: return
            
        try:
            self.update_status("Exporting intersections...")
            feature_col = 'Feature_ID' if 'Feature_ID' in self.preprocessed_data.columns else 'FeatureID'
            
            feature_map = defaultdict(list)
            for group_name, features in self.upset_sets.items():
                for f in features:
                    feature_map[f].append(group_name)
            
            signature_map = defaultdict(list)
            for f, group_list in feature_map.items():
                group_list.sort()
                signature_map[" & ".join(group_list)].append(f)
                
            export_data = {sig: feats for sig, feats in signature_map.items()}
            sorted_signatures = sorted(export_data.keys(), key=lambda s: (s.count('&'), s), reverse=True)
            
            if not sorted_signatures:
                self.update_status("Export cancelled: No features to export.")
                messagebox.showinfo("Info", "No features met the criteria.")
                return

            if file_path.endswith('.xlsx'):
                upset_groups = list(self.upset_sets.keys())
                all_sample_cols = [c for c in self.preprocessed_data.columns if '.mzML' in c]
                group_cols_map = {}
                for g in upset_groups:
                    group_cols = []
                    for base_name, reps in self.replicate_mapping.items():
                        if self.group_mapping.get(base_name) == g:
                            avg_col = f"{base_name}_avg.mzML Peak area"
                            if avg_col in all_sample_cols: group_cols.append(avg_col)
                            else: group_cols.extend([c for c, _ in reps if c in all_sample_cols])
                    group_cols_map[g] = group_cols

                used_sheet_names = set()
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for name in sorted_signatures:
                        features = export_data[name]
                        base_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', name)[:31]
                        sheet_name = base_sheet_name
                        counter = 1
                        while sheet_name in used_sheet_names:
                            suffix = f"_{counter}"
                            sheet_name = base_sheet_name[:31 - len(suffix)] + suffix
                            counter += 1
                        used_sheet_names.add(sheet_name)
                        
                        if features:
                            df = self.preprocessed_data[self.preprocessed_data[feature_col].isin(features)].copy()
                            df = ExportManager.process_feature_ids(df, feature_col)
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                        else:
                            pd.DataFrame(columns=['ID', 'm/z', 'RT', feature_col]).to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                all_rows = []
                for name in sorted_signatures:
                    for f in export_data[name]:
                        row = {'Type': name, feature_col: f}
                        try:
                            parts = str(f).rsplit('_', 2)
                            if len(parts) == 3: row['ID'], row['m/z'], row['RT'] = parts[0], parts[1], parts[2]
                        except: pass
                        all_rows.append(row)
                df = pd.DataFrame(all_rows)
                df[[c for c in ['Type', 'ID', 'm/z', 'RT', feature_col] if c in df.columns]].to_csv(file_path, index=False)
                
            messagebox.showinfo("Success", f"Exported UpSet intersections to:\n{file_path}")
            self.update_status("✓ Export complete")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{str(e)}")
            self.update_status("Export failed")

    def export_upset_data_subset(self, intersection_type):
        """Export data for a specific UpSet intersection type (triggered via context menu)."""
        if not hasattr(self, 'upset_sets') or not self.upset_sets: return
        groups = list(self.upset_sets.keys())
        if len(groups) < 2: return
        
        common = set.intersection(*list(self.upset_sets.values()))
        unique_sets = {}
        for g in groups:
            others = [self.upset_sets[og] for og in groups if og != g]
            unique_sets[g] = self.upset_sets[g] - set.union(*others) if others else self.upset_sets[g]
            
        final_features = set()
        suffix = ""
        if intersection_type == 'COMMON':
            final_features = common
            suffix = "common"
        elif intersection_type == 'UNIQUE_COMBINED':
            final_features = set.union(*unique_sets.values())
            suffix = "unique_all_groups"
        elif intersection_type.startswith('UNIQUE_'):
            g_name = intersection_type[7:]
            if g_name in unique_sets:
                final_features = unique_sets[g_name]
                suffix = f"unique_{g_name}"
        
        if not final_features: return messagebox.showinfo("Info", f"No features found for {suffix}.")
        if getattr(self, 'preprocessed_data', None) is None: return
        
        subset_df = self.preprocessed_data[self.preprocessed_data.iloc[:, 0].isin(final_features)]
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", initialfile=f"upset_{suffix}_features_data.csv", filetypes=[("CSV files", "*.csv")])
        if file_path:
            if 'Feature_ID' in subset_df.columns:
                subset_df = ExportManager.process_feature_ids(subset_df.copy(), 'Feature_ID')
            subset_df.to_csv(file_path, index=False)
            messagebox.showinfo("Success", f"Exported {len(subset_df)} features.")

    # ==================== Feature Viewer ====================

    def update_feature_viewer_options(self):
        """Populate the feature selection combobox in the Feature Viewer tab."""
        if getattr(self, 'data', None) is not None and hasattr(self, 'feature_viewer_combo'):
            feature_ids = self.data.iloc[:, 0].tolist()
            self.all_feature_ids = feature_ids
            self.feature_viewer_combo['values'] = feature_ids
            if feature_ids and not self.feature_viewer_id_var.get():
                self.feature_viewer_id_var.set(feature_ids[0])

    def filter_feature_viewer_list(self, var_name, index, mode):
        """Filter the feature combobox based on search entry"""
        if not hasattr(self, 'all_feature_ids') or not self.all_feature_ids:
            return
        search_term = self.feature_search_var.get().lower()
        if search_term == '':
            self.feature_viewer_combo['values'] = self.all_feature_ids
        else:
            filtered = [f for f in self.all_feature_ids if search_term in str(f).lower()]
            self.feature_viewer_combo['values'] = filtered

    def run_feature_plot(self, event=None):
        """Generate a box plot for the selected feature across groups."""
        data_source = getattr(self, 'fv_data_source_var', tk.StringVar(value="Raw Data")).get()
        target_df = self.preprocessed_data if data_source == "Preprocessed Data" else self.data
        
        if target_df is None:
            messagebox.showwarning("Warning", f"{data_source} is not available. Please run the required steps first.")
            return

        feature_id = self.feature_viewer_id_var.get()
        if not feature_id:
            return

        transform_type = getattr(self, 'fv_transform_var', tk.StringVar(value="None")).get()
        use_log = getattr(self, 'fv_log_var', tk.BooleanVar(value=False)).get()
        if use_log and transform_type == "None":
            transform_type = "Log10" # Fallback for backwards compatibility with old sessions
        impute_lod = getattr(self, 'fv_impute_lod_var', tk.BooleanVar(value=False)).get()
        show_violin = getattr(self, 'fv_show_violin_var', tk.BooleanVar(value=True)).get()
        show_boxplot = getattr(self, 'fv_show_boxplot_var', tk.BooleanVar(value=True)).get()
        show_points = getattr(self, 'fv_show_points_var', tk.BooleanVar(value=True)).get()
        
        lod_fraction_str = getattr(self, 'fv_lod_fraction_var', tk.StringVar(value="1/5 (20%)")).get()
        if "1/2" in lod_fraction_str: lod_fraction = 0.5
        elif "1/3" in lod_fraction_str: lod_fraction = 1.0 / 3.0
        elif "1/4" in lod_fraction_str: lod_fraction = 0.25
        else: lod_fraction = 0.2

        try:
            self.update_status(f"Plotting feature: {feature_id}...")
            
            feature_row = target_df[target_df.iloc[:, 0] == feature_id]
            if feature_row.empty:
                messagebox.showerror("Error", f"Feature '{feature_id}' not found in {data_source}.")
                self.update_status(f"Error: Feature {feature_id} not found.")
                return

            feature_row = feature_row.iloc[0]

            # Prepare data for plotting
            plot_data, plot_labels, plot_sample_names = [], [], []
            
            # Prepare storage for the Export button
            self.current_feature_plot_data = {
                'FeatureID': feature_id,
                'Data': []
            }
            
            # Calculate LOD value if imputation is enabled
            lod_value = None
            if impute_lod:
                all_sample_cols = [col for col in target_df.columns if '.mzML' in col]
                all_intensities = pd.to_numeric(feature_row[all_sample_cols], errors='coerce').values.astype(float)
                pos_intensities = all_intensities[all_intensities > 0]
                if len(pos_intensities) > 0:
                    lod_value = np.min(pos_intensities) * lod_fraction
                else:
                    lod_value = 0.0
            
            unique_groups = sorted(set(self.group_mapping.values()))
            
            exclude_qc = getattr(self, 'feature_plot_exclude_qc_var', tk.BooleanVar(value=True)).get()
            if exclude_qc:
                unique_groups = [g for g in unique_groups if 'qc' not in g.lower()]
                
            selected_indices = self.fv_group_listbox.curselection()
            selected_groups = [self.fv_group_listbox.get(i) for i in selected_indices]
            
            if selected_groups:
                unique_groups = [g for g in unique_groups if g in selected_groups]
                
            special_groups = [g for g in ['QC', 'Ungrouped'] if g in unique_groups]
            for g in special_groups:
                unique_groups.remove(g)
            unique_groups.extend(special_groups)

            ylabel = "Normalized Peak Intensity" if data_source == "Preprocessed Data" else "Peak Intensity"
            if transform_type == "Log10":
                ylabel = f"Log10({ylabel})"
            elif transform_type == "Square Root":
                ylabel = f"Sqrt({ylabel})"
            elif transform_type == "Cube Root":
                ylabel = f"Cbrt({ylabel})"

            for group in unique_groups:
                group_cols = []
                for base_name, reps in self.replicate_mapping.items():
                    if self.group_mapping.get(base_name) == group:
                        avg_col = f"{base_name}_avg.mzML Peak area"
                        if avg_col in target_df.columns:
                            group_cols.append(avg_col)
                        else:
                            rep_cols = [c for c, _ in reps if c in target_df.columns]
                            group_cols.extend(rep_cols)
                            
                if not group_cols: continue

                intensities = pd.to_numeric(feature_row[group_cols], errors='coerce').values.astype(float)
                
                if impute_lod and lod_value is not None:
                    # Replace 0 or NaN with lod_value
                    missing_mask = np.isnan(intensities) | (intensities <= 0)
                    intensities[missing_mask] = lod_value

                # Clip any accidental negative baselines before applying math transforms
                intensities[intensities < 0] = 0
                
                if transform_type == "Log10":
                    # Use log10(x + 1) so 0 intensity remains 0 and isn't dropped
                    intensities = np.log10(intensities + 1)
                elif transform_type == "Square Root":
                    intensities = np.sqrt(intensities)
                elif transform_type == "Cube Root":
                    intensities = np.cbrt(intensities)

                # Explicitly missing values (NaN) will be ignored and excluded from the plot
                valid_mask = ~np.isnan(intensities)
                plot_data.append(intensities[valid_mask])
                plot_labels.append(group)
                
                # Extract clean sample names
                clean_names = [re.sub(r'\.mzML.*', '', c) for c in group_cols]
                valid_names = np.array(clean_names)[valid_mask]
                plot_sample_names.append(valid_names.tolist())
                
                # Save to export array
                for idx_val, val in enumerate(intensities[valid_mask]):
                    self.current_feature_plot_data['Data'].append({
                        'Sample': valid_names[idx_val],
                        'Group': group,
                        'Value': val
                    })

            self.vis_manager.create_feature_distribution_plot(
                feature_id, plot_data, plot_labels, plot_sample_names, 
                ylabel=ylabel, show_points=show_points, show_violin=show_violin, show_boxplot=show_boxplot
            )
            self.update_status(f"✓ Plot for {feature_id} generated.")

        except Exception as e:
            import traceback
            messagebox.showerror("Error", f"Failed to plot feature:\n{str(e)}\n\nDetails:\n{traceback.format_exc()}")
            self.update_status("Error plotting feature")

    # ==================== Export Functions ====================
    
    def export_preprocessed(self):
        """Export preprocessed data to CSV"""
        self.export_manager.export_preprocessed_data(self.preprocessed_data)
    
    def export_pca(self):
        """Export PCA results to CSV"""
        self.export_manager.export_pca_results(self.pca_result)
    
    def export_volcano_results(self):
        """Export significant features from volcano plot analysis to CSV"""
        self.export_manager.export_volcano_results(
            self.volcano_result,
            self.preprocessed_data,
            self.pvalue_var.get(),
            self.fc_var.get()
        )

    def export_volcano_subset(self, subset_type="BOTH"):
        """
        Export specific subset of volcano plot results (UP, DOWN, or BOTH).
        Triggered from Volcano Plot context menu.
        """
        self.export_manager.export_volcano_subset(
            self.volcano_result,
            self.pvalue_var.get(),
            self.fc_var.get(),
            subset_type
        )

    def export_rf_features(self):
        """Export Random Forest feature importance to CSV"""
        self.export_manager.export_rf_results(self.rf_result)
        
    def export_consensus_features(self):
        """Find and export exact overlap of significant features between Volcano, PLS-DA, and RF"""
        if not all([self.volcano_result, self.plsda_result, self.rf_result]):
            messagebox.showwarning("Warning", "Please run Volcano Plot, PLS-DA, and Random Forest first!")
            return
            
        try:
            self.update_status("Calculating consensus biomarkers...")
            
            # 1. Volcano Significant Features
            pval_thresh = self.pvalue_var.get()
            fc_thresh = self.fc_var.get()
            v_fc = self.volcano_result['fold_changes']
            v_pvals = self.volcano_result['pvalues']
            
            sig_up = (v_fc >= np.log2(fc_thresh)) & (v_pvals <= pval_thresh)
            sig_down = (v_fc <= -np.log2(fc_thresh)) & (v_pvals <= pval_thresh)
            volcano_sig_mask = sig_up | sig_down
            
            volcano_features = set(self.volcano_result['feature_ids'][volcano_sig_mask])
            volcano_dict = {f_id: (fc, p) for f_id, fc, p in zip(self.volcano_result['feature_ids'], v_fc, v_pvals) if f_id in volcano_features}

            # 2. PLS-DA VIP > 1.0 Features
            vip_scores = self.plsda_result['vip_scores']
            plsda_features_all = self.plsda_result.get('feature_ids', self.screened_data.iloc[:, 0].values)
            
            plsda_features = set(plsda_features_all[vip_scores > 1.0])
            plsda_dict = {f_id: vip for f_id, vip in zip(plsda_features_all, vip_scores) if f_id in plsda_features}

            # 3. Random Forest Top Features
            rf_dict = {f[0]: f[1] for f in self.rf_result['top_features']}
            rf_features = set(rf_dict.keys())

            # 4. Find Intersection
            consensus_features = volcano_features & plsda_features & rf_features
            
            if not consensus_features:
                messagebox.showinfo("Consensus Biomarkers", "No features met all three criteria:\n- Volcano (Significant)\n- PLS-DA (VIP > 1.0)\n- Random Forest (Top N)")
                self.update_status("No consensus features found.")
                return
                
            # Build Export Data
            data = [{'FeatureID': f, 'Log2FC': volcano_dict[f][0], 'P-value': volcano_dict[f][1], 'PLS-DA_VIP': plsda_dict[f], 'RF_Importance': rf_dict[f]} for f in consensus_features]
            df = pd.DataFrame(data).sort_values('RF_Importance', ascending=False)
            df = ExportManager.process_feature_ids(df, 'FeatureID')

            filepath = filedialog.asksaveasfilename(defaultextension=".csv", initialfile="consensus_biomarkers.csv", filetypes=[("CSV files", "*.csv")])
            
            if filepath:
                df.to_csv(filepath, index=False)
                messagebox.showinfo("Success", f"Found and exported {len(consensus_features)} robust consensus biomarkers to:\n{filepath}")
                self.update_status(f"✓ Exported {len(consensus_features)} consensus biomarkers")
                
        except Exception as e:
            import traceback
            messagebox.showerror("Error", f"Failed to calculate consensus features:\n{str(e)}\n\nDetails:\n{traceback.format_exc()}")
            self.update_status("Error calculating consensus features")

    def export_plsda_vips(self):
        """Export PLS-DA VIP scores to CSV"""
        self.export_manager.export_plsda_results(self.plsda_result, self.screened_data)
        
    def export_heatmap_data(self):
        """Export heatmap data to CSV"""
        self.export_manager.export_heatmap_data(self.heatmap_data)
        
    def export_feature_data(self):
        """Export the exact sample data points for the currently viewed feature."""
        if not hasattr(self, 'current_feature_plot_data') or not self.current_feature_plot_data.get('Data'):
            messagebox.showwarning("Warning", "No feature data to export. Please plot a feature first.")
            return
            
        feature_id = self.current_feature_plot_data['FeatureID']
        clean_id = str(feature_id).replace('/', '_').replace(':', '_')
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=f"FeatureData_{clean_id}.csv",
            filetypes=[("CSV files", "*.csv")]
        )
        
        if filepath:
            try:
                df = pd.DataFrame(self.current_feature_plot_data['Data'])
                df.to_csv(filepath, index=False)
                messagebox.showinfo("Success", f"Feature data exported successfully to:\n{filepath}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export feature data:\n{str(e)}")

    def export_current_plot(self):
        """Export the plot currently visible in the active tab"""
        current_tab = self.notebook.index(self.notebook.select())
        tab_text = self.notebook.tab(current_tab, "text")
        
        plot_key = None
        default_name = "plot.png"
        
        # Map tabs to stored plots
        if "Prepro" in tab_text:
            plot_key = 'Preprocessing'
            default_name = "distribution_plot.png"
        elif "PCA" in tab_text:
            plot_key = 'PCA'
            default_name = "pca_score_plot.png"
        elif "Screening" in tab_text: # Univariate Screening
            plot_key = 'Univariate Screening'
            default_name = "volcano_plot.png"
        elif "PLS-DA" in tab_text:
            plot_key = 'PLS-DA'
            default_name = "plsda_plot.png"
        elif "Validation" in tab_text:
            plot_key = 'PLS-DA Validation'
            default_name = "plsda_validation_plot.png"
        elif "Random Forest" in tab_text:
            plot_key = 'Random Forest'
            default_name = "rf_importance_plot.png"
        elif "Heatmap" in tab_text:
            plot_key = 'Heatmap'
            default_name = "heatmap.png"
        elif "Comparative" in tab_text or "Venn" in tab_text:
            plot_key = 'Venn Diagram'
            default_name = "venn_diagram.png"
        elif "UpSet" in tab_text:
            plot_key = 'UpSet Plot'
            default_name = "upset_plot.png"
        elif "Feature Viewer" in tab_text:
            plot_key = 'Feature_Viewer'
            default_name = f"feature_plot_{self.feature_viewer_id_var.get().replace('/', '_').replace(':', '_')}.png"
        elif "Spectrum" in tab_text:
            plot_key = 'Spectrum'
            default_name = "spectrum_plot.png"
            
        fig = self.generated_plots.get(plot_key)
        
        if fig:
            self.export_manager.save_plot_high_res(fig, default_name, parent=self)
        else:
            messagebox.showinfo("Info", "No plot found for the current tab.\nMake sure you have generated the plot first.")

    def export_pdf_report(self):
        """Export all generated plots to a single PDF file with customizable sizes."""
        file_label_text = self.file_label.cget("text") if self.data is not None and hasattr(self, 'file_label') else ""
        
        self.export_manager.export_pdf_report(
            self.generated_plots,
            self.plsda_result,
            self.rf_result,
            self.heatmap_data,
            getattr(self, 'screened_data', None),
            file_label_text
        )

    # ==================== Session Management ====================

    def save_session(self):
        """Save the entire analysis session to a file."""
        if self.data is None:
            messagebox.showwarning("Warning", "No data loaded. Nothing to save.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".hmas",
            initialfile="analysis_session.hmas",
            filetypes=[("Herbal Metabolomics Analysis Session", "*.hmas"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            self.update_status("Saving session...")

            session_state = {
                'data_state': {
                    'data': self.data,
                    'preprocessed_data': self.preprocessed_data,
                    'screened_data': self.screened_data,
                    'data_before_norm': self.data_before_norm,
                    'replicate_mapping': self.replicate_mapping,
                    'group_mapping': self.group_mapping,
                    'group_colors': self.group_colors,
                    'qc_samples': self.qc_samples,
                    'sample_names': self.sample_names,
                    'file_label_text': self.file_label.cget("text")
                },
                'results_state': {
                    'pca_result': self.pca_result,
                    'plsda_result': self.plsda_result,
                    'volcano_result': self.volcano_result,
                    'rf_result': self.rf_result,
                    'heatmap_data': self.heatmap_data,
                    'venn_sets': getattr(self, 'venn_sets', None),
                    'upset_sets': getattr(self, 'upset_sets', None),
                },
                'gui_vars': {
                    'avg_replicates_var': self.avg_replicates_var.get(),
                    'impute_lod_var': getattr(self, 'impute_lod_var', tk.BooleanVar(value=False)).get(),
                    'lod_fraction_var': getattr(self, 'lod_fraction_var', tk.StringVar(value="1/5 (20%)")).get(),
                    'filter_detection_var': self.filter_detection_var.get(),
                    'detection_threshold_var': self.detection_threshold_var.get(),
                    'filter_intensity_var': self.filter_intensity_var.get(),
                    'min_intensity_var': self.min_intensity_var.get(),
                    'filter_rsd_var': self.filter_rsd_var.get(),
                    'rsd_threshold_var': self.rsd_threshold_var.get(),
                    'filter_iqr_var': self.filter_iqr_var.get(),
                    'iqr_factor_var': self.iqr_factor_var.get(),
                    'norm_method_var': self.norm_method_var.get(),
                    'pca_components_var': self.pca_components_var.get(),
                    'pca_log_var': self.pca_log_var.get(),
                    'show_ellipses_var': self.show_ellipses_var.get(),
                    'show_labels_var': self.show_labels_var.get(),
                    'pca_permanova_var': getattr(self, 'pca_permanova_var', tk.BooleanVar(value=False)).get(),
                    'pca_permdisp_var': getattr(self, 'pca_permdisp_var', tk.BooleanVar(value=False)).get(),
                    'volcano_group1_var': self.volcano_group1_var.get(),
                    'volcano_group2_var': self.volcano_group2_var.get(),
                    'pvalue_var': self.pvalue_var.get(),
                    'fc_var': self.fc_var.get(),
                    'plsda_components_var': self.plsda_components_var.get(),
                    'plsda_log_var': self.plsda_log_var.get(),
                    'plsda_exclude_qc_var': self.plsda_exclude_qc_var.get(),
                    'n_perms_var': self.n_perms_var.get(),
                    'pls_val_exclude_qc_var': self.pls_val_exclude_qc_var.get(),
                    'ntrees_var': self.ntrees_var.get(),
                    'top_n_var': self.top_n_var.get(),
                    'rf_exclude_qc_var': self.rf_exclude_qc_var.get(),
                    'rf_pairwise_var': getattr(self, 'rf_pairwise_var', tk.BooleanVar(value=False)).get(),
                    'rf_group1_var': getattr(self, 'rf_group1_var', tk.StringVar()).get(),
                    'rf_group2_var': getattr(self, 'rf_group2_var', tk.StringVar()).get(),
                    'rf_roc_var': getattr(self, 'rf_roc_var', tk.BooleanVar(value=True)).get(),
                    'heatmap_top_n_var': self.heatmap_top_n_var.get(),
                    'heatmap_sort_var': getattr(self, 'heatmap_sort_var', tk.StringVar(value="Clustering")).get(),
                    'heatmap_cmap_var': getattr(self, 'heatmap_cmap_var', tk.StringVar(value="coolwarm")).get(),
                    'venn_thresh_var': self.venn_thresh_var.get(),
                    'upset_thresh_var': getattr(self, 'upset_thresh_var', tk.DoubleVar(value=80.0)).get(),
                    'feature_search_var': getattr(self, 'feature_search_var', tk.StringVar()).get(),
                    'feature_viewer_id_var': self.feature_viewer_id_var.get(),
                    'fv_data_source_var': getattr(self, 'fv_data_source_var', tk.StringVar(value="Preprocessed Data")).get(),
                    'fv_log_var': getattr(self, 'fv_log_var', tk.BooleanVar(value=False)).get(),
                    'fv_transform_var': getattr(self, 'fv_transform_var', tk.StringVar(value="None")).get(),
                    'feature_plot_exclude_qc_var': getattr(self, 'feature_plot_exclude_qc_var', tk.BooleanVar(value=True)).get(),
                    'fv_impute_lod_var': getattr(self, 'fv_impute_lod_var', tk.BooleanVar(value=False)).get(),
                    'fv_lod_fraction_var': getattr(self, 'fv_lod_fraction_var', tk.StringVar(value="1/5 (20%)")).get(),
                    'fv_show_violin_var': getattr(self, 'fv_show_violin_var', tk.BooleanVar(value=True)).get(),
                    'fv_show_boxplot_var': getattr(self, 'fv_show_boxplot_var', tk.BooleanVar(value=True)).get(),
                    'fv_show_points_var': getattr(self, 'fv_show_points_var', tk.BooleanVar(value=True)).get(),
                },
                'gui_texts': {
                    'summary_text': self.summary_text.get('1.0', tk.END),
                    'preprocess_text': self.preprocess_text.get('1.0', tk.END),
                    'plsda_results_text': self.plsda_results_text.get('1.0', tk.END),
                    'rf_results_text': self.rf_results_text.get('1.0', tk.END),
                }
            }

            with open(file_path, 'wb') as f:
                pickle.dump(session_state, f)

            self.update_status("✓ Session saved successfully.")
            messagebox.showinfo("Success", f"Session saved to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save session:\n{str(e)}")
            self.update_status("Error saving session.")

    def load_session(self):
        """Load a previously saved analysis session."""
        file_path = filedialog.askopenfilename(
            title="Select Session File",
            filetypes=[("Herbal Metabolomics Analysis Session", "*.hmas"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            self.update_status("Loading session...")
            with open(file_path, 'rb') as f:
                session_state = pickle.load(f)
            
            self._restore_session_state(session_state)

            self.update_status("✓ Session loaded successfully.")
            messagebox.showinfo("Success", f"Session loaded from:\n{file_path}")

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            messagebox.showerror("Error", f"Failed to load session:\n{str(e)}\n\nDetails:\n{error_detail}")
            self.update_status("Error loading session.")

    def configure_plot_settings(self):
        """Dialog to configure global plot dimensions and DPI."""
        settings_window = tk.Toplevel(self)
        settings_window.title("Global Plot Settings")
        settings_window.geometry("380x250")
        settings_window.transient(self)
        settings_window.grab_set()

        main_frame = ttk.Frame(settings_window, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="Default Plot Dimensions", font=('Arial', 12, 'bold')).grid(row=0, column=0, columnspan=2, pady=(0, 15))

        # Width
        ttk.Label(main_frame, text="Base Width (inches):").grid(row=1, column=0, sticky='w', pady=5)
        width_spinbox = ttk.Spinbox(main_frame, from_=4, to=20, increment=0.5, textvariable=self.plot_width_var, width=10)
        width_spinbox.grid(row=1, column=1, sticky='e', pady=5)

        # Height
        ttk.Label(main_frame, text="Base Height (inches):").grid(row=2, column=0, sticky='w', pady=5)
        height_spinbox = ttk.Spinbox(main_frame, from_=4, to=20, increment=0.5, textvariable=self.plot_height_var, width=10)
        height_spinbox.grid(row=2, column=1, sticky='e', pady=5)

        # DPI
        ttk.Label(main_frame, text="DPI (dots per inch):").grid(row=3, column=0, sticky='w', pady=5)
        dpi_spinbox = ttk.Spinbox(main_frame, from_=75, to=300, increment=25, textvariable=self.plot_dpi_var, width=10)
        dpi_spinbox.grid(row=3, column=1, sticky='e', pady=5)

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=(20, 0))

        def apply_and_refresh():
            settings_window.destroy()
            if self.generated_plots:
                response = messagebox.askyesno(
                    "Refresh Plots?",
                    "Apply these settings and refresh all existing plots?",
                    parent=self
                )
                if response:
                    self.update_status("Refreshing plots with new dimensions...")
                    self._recreate_all_plots()
                    self.update_status("✓ Plots refreshed.")

        ttk.Button(button_frame, text="Apply & Refresh", command=apply_and_refresh).pack(side='left', padx=10)
        ttk.Button(button_frame, text="Cancel", command=settings_window.destroy).pack(side='left', padx=10)

    def _restore_session_state(self, state):
        """Helper function to apply the loaded state to the application."""
        # Restore data state
        data = state.get('data_state', {})
        for attr, value in data.items():
            if attr == 'file_label_text':
                self.file_label.config(text=value, foreground="green")
            else:
                setattr(self, attr, value)

        # Restore GUI variables
        gui_vars = state.get('gui_vars', {})
        for var_name, value in gui_vars.items():
            if hasattr(self, var_name):
                getattr(self, var_name).set(value)

        # Restore analysis results
        results = state.get('results_state', {})
        for attr, value in results.items():
            setattr(self, attr, value)

        # Restore text widgets
        texts = state.get('gui_texts', {})
        self._repopulate_result_texts(texts)

        # Update UI elements that depend on loaded data
        self.display_data_preview()
        self.update_group_dropdowns()
        self.update_qc_status_label()
        self.on_rsd_filter_toggle()
        self.show_dist_button.config(state='normal' if self.preprocessed_data is not None else 'disabled')
        self.update_venn_groups()
        self.update_upset_groups()
        self.update_feature_viewer_groups()
        self.update_feature_viewer_options()

        if getattr(self, 'plsda_pairwise_var', tk.BooleanVar(value=False)).get() and hasattr(self, 'plsda_group_frame'):
            self.plsda_group_frame.grid()
        if getattr(self, 'rf_pairwise_var', tk.BooleanVar(value=False)).get() and hasattr(self, 'rf_group_frame'):
            self.rf_group_frame.grid()

        # Recreate all plots
        self._recreate_all_plots()

    def _repopulate_result_texts(self, texts):
        """Helper to fill text boxes from loaded session."""
        if 'summary_text' in texts:
            self.summary_text.delete('1.0', tk.END)
            self.summary_text.insert('1.0', texts['summary_text'])
        if 'preprocess_text' in texts:
            self.preprocess_text.delete('1.0', tk.END)
            self.preprocess_text.insert('1.0', texts['preprocess_text'])
        if 'plsda_results_text' in texts:
            self.plsda_results_text.delete('1.0', tk.END)
            self.plsda_results_text.insert('1.0', texts['plsda_results_text'])
        if 'rf_results_text' in texts:
            self.rf_results_text.delete('1.0', tk.END)
            self.rf_results_text.insert('1.0', texts['rf_results_text'])

    def _recreate_all_plots(self):
        """Helper to redraw all plots after loading a session."""
        self.generated_plots = {} # Clear old plot objects
        if self.preprocessed_data is not None: self.show_distribution_plots()
        if self.pca_result is not None: self.create_pca_plot_with_groups(self.pca_result)
        if self.volcano_result is not None: self.create_volcano_plot(self.volcano_result, self.pvalue_var.get(), self.fc_var.get())
        if self.plsda_result is not None: self.create_plsda_plot(self.plsda_result)
        if self.preprocessed_data is not None: self.vis_manager.create_distribution_plots()
        if self.pca_result is not None: self.vis_manager.create_pca_plot(self.pca_result)
        if self.volcano_result is not None: self.vis_manager.create_volcano_plot(self.volcano_result, self.pvalue_var.get(), self.fc_var.get())
        if self.plsda_result is not None: self.vis_manager.create_plsda_plot(self.plsda_result)
        if self.plsda_result is not None:
            # Check if a validation plot was previously generated
            if 'PLS-DA Validation' in self.generated_plots:
                self.run_pls_val() # Rerunning is the most reliable way to regenerate this plot
        if self.rf_result is not None: self.create_rf_plot(self.rf_result)
        if self.rf_result is not None: self.vis_manager.create_rf_plot(self.rf_result)
        if self.heatmap_data is not None: self.run_heatmap()
        if getattr(self, 'venn_sets', None) is not None:
            self.vis_manager.create_venn_plot(self.venn_sets, list(self.venn_sets.keys()))
        if getattr(self, 'upset_sets', None) is not None:
            self.vis_manager.create_upset_plot(self.upset_sets, list(self.upset_sets.keys()))
        if self.feature_viewer_id_var.get():
            self.run_feature_plot()

    # ==================== Utility Functions ====================
    
    def treeview_sort_column(self, tv, col, reverse):
        """Sort treeview contents when a column header is clicked."""
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        
        try:
            # Sort numerically if possible (treat empty strings as -infinity for sorting)
            l.sort(key=lambda t: float(t[0]) if t[0] != "" else float('-inf'), reverse=reverse)
        except ValueError:
            # Fallback to alphabetical string sorting
            l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        # Update the heading command to reverse the sort direction on the next click
        tv.heading(col, command=lambda _col=col: self.treeview_sort_column(tv, _col, not reverse))

    def update_status(self, message):
        """Update status bar"""
        self.status_bar.config(text=message)
        self.update_idletasks()
        
    def show_help(self):
        """Show quick start guide"""
        help_window = tk.Toplevel(self)
        help_window.title("Quick Start Guide")
        help_window.geometry("700x600")
        
        help_text = scrolledtext.ScrolledText(help_window, wrap=tk.WORD, font=('Arial', 10))
        help_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        guide = """
HERBAL METABOLOMICS ANALYZER - QUICK START GUIDE
=================================================

1. LOAD DATA
   - Click "Load CSV File" 
   - Select your MZmine quantification CSV
   - Works with any sample naming convention

2. RENAME SAMPLES (OPTIONAL BUT RECOMMENDED)
   - Click "Rename Samples" button
   - Replace MZmine IDs (e.g. JW1-119-10) with biological names (e.g. Control_A)
   - The app safely updates the underlying columns.

3. CONFIGURE GROUPS (IMPORTANT!)
   - Click "Configure Groups" button
   - Assign each sample to a group
   - Type new group names in dropdown
   - 💡 TIP: Name QC group as "QC" for auto-detection
   - Customize group colors
   - Click "Save Configuration"

4. SELECT QC SAMPLES (For RSD Filtering)
   - Click "Select QC Samples" button
   - Check samples to use as QC
   - Or use "Auto-Detect" for groups containing "QC"
   - Click "Save Selection"

5. PREPROCESSING
   - Keep original + add averaged (3 replicates total)
   
   FEATURE FILTERING:
   • Detection rate: Remove features not detected in ≥X% samples
   • Min intensity: Remove low-abundance features below threshold
   • QC RSD: Remove features with high variability in QC samples
     → Select QC samples first!
   • IQR filter: Remove features with low variance
   
   - Select normalization method (TIC recommended)
   - Click "Run Preprocessing"
   - Click "Show Distribution Plots" to verify normalization

6. PCA ANALYSIS
   - Enable 95% confidence ellipses
   - Set number of components
   - Check the "Log10 Transform" box if your data is skewed into a single cluster
   - Click "Run PCA"
   - Hover over dots in the PCA plot to identify outliers and sample IDs

7. UNIVARIATE SCREENING (Volcano Plot)
   - Use the dropdown menus to explicitly select which two groups to compare.
   - Set p-value & fold-change thresholds to find obvious initial differences.
   - Hover over significant data points to see feature IDs.
   - Click "Filter Data for PLS-DA & RF" to reduce the dataset exclusively 
     to these significant features for sharper downstream ML modeling.

8. PLS-DA ANALYSIS & VALIDATION
   - Run standard PLS-DA to extract LV scores and calculate VIP metrics
   - Use the "PLS Validation" tab to run a rigorous permutation test
     (essential for preventing model overfitting)

9. RANDOM FOREST & HEATMAP
   - Run Random Forest to score feature importance and calculate OOB error
   - Use the "Heatmap (HCA)" tab to visually validate the Top VIP & RF 
     metabolites across your actual sample groups!
        """
        
        help_text.insert('1.0', guide)
        help_text.config(state='disabled')
        
    def show_about(self):
        """Show about dialog"""
        messagebox.showinfo(
            "About",
            "Herbal Metabolomics Analyzer\n\n"
            "Version: 2.14.1\n"
            "Built with Python + Tkinter + scikit-learn\n\n"
            "Complete integrated workflow for untargeted\n"
            "metabolomics analysis of herbal medicine samples\n\n"
            "Features:\n"
            "• Works with ANY sample naming convention\n"
            "• Advanced preprocessing & technical duplicate averaging\n"
            "• Smart Sample Renaming across dataset\n"
            "• QC sample RSD filtering\n"
            "• Interactive PCA with hover tooltips and ellipses\n"
            "• Univariate Screening (Volcano subsetting for ML models)\n"
            "• PLS-DA modeling with Variable Importance (VIP) scores\n"
            "• 100x Permutation Testing (R2/Q2) for model validation\n"
            "• Random Forest feature evaluation & OOB Error testing\n"
            "• Single-HCA Hierarchical Abundance Heatmaps\n\n"
            "© 2026 Herbal Metabolomics Research Lab"
        )


def main():
    """Main entry point"""
    app = MetabolomicsApp()
    app.mainloop()


if __name__ == "__main__":
    main()